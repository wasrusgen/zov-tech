"""
Парсер таблицы занятости сборщиков (Excel).

Формат файла:
  Col A: зона (север, охта, ...)
  Col B: ФИО + телефоны (всё в одной ячейке)
  Col C: тип строки ('сборка' / 'замер' / 'только замеры')
  Col D+: ячейки заказов, одна колонка = один день

  Строка 1 (2026+): даты
  Строка 2 (2026+): дни недели
  --- VS ---
  Строка 1 (до 2026): дни недели
  Строка 2 (до 2026): даты

  Стоимость сборки — последнее число в тексте ячейки:
    '1322Б Парголово ул.Заречная 10 кв105 89110064400 Алексеев А.К. 63900'
    Составная: '6030+20100' → 26130
    Доделка без суммы: '' или 'Доделка ...' → 0
"""
from __future__ import annotations
import os
import re
import json
import time
import logging
from pathlib import Path
from datetime import date, datetime
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

log = logging.getLogger("zov.assembler_parser")

# Кэш: {path: {mtime, data}}
_cache: dict[str, dict] = {}

_AMOUNT_RE = re.compile(r"([\d]+(?:[+][\d]+)*)\s*$")
_COMPOUND_RE = re.compile(r"^\d+(?:[+]\d+)+$")


def _extract_amount(text: str) -> int:
    """Извлекает стоимость из конца текста ячейки."""
    text = (text or "").strip()
    m = _AMOUNT_RE.search(text)
    if not m:
        return 0
    raw = m.group(1)
    if _COMPOUND_RE.match(raw):
        return sum(int(x) for x in raw.split("+"))
    return int(raw)


def _extract_name(cell_b: str) -> str:
    """Извлекает ФИО из первой строки или до первого телефона."""
    s = (cell_b or "").strip()
    if not s:
        return ""
    # Первая строка
    first_line = s.split("\n")[0].strip()
    # Обрезаем по телефону (8-9xx, +7)
    m = re.search(r"[\s,](?:8[-\s]?9|(?:\+7))\d", first_line)
    if m:
        first_line = first_line[: m.start()].strip()
    return first_line or s.split("\n")[0]


def _is_date_row(row_vals: list) -> bool:
    """Проверяет, содержит ли строка datetime-объекты (ряд с датами)."""
    dates = [v for v in row_vals if isinstance(v, (datetime, date))]
    return len(dates) >= 5


def parse_sheet(ws) -> list[dict]:
    """
    Парсит один лист. Возвращает список записей:
      {name, zone, date, order_text, amount, sheet_title}
    """
    if openpyxl is None:
        return []

    title = ws.title
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 3:
        return []

    # Определяем строку с датами (row index 0 или 1)
    date_row_idx = None
    for i in (0, 1):
        if _is_date_row(rows[i]):
            date_row_idx = i
            break
    if date_row_idx is None:
        return []

    date_row = rows[date_row_idx]
    # Строим словарь col_index → date
    col_to_date: dict[int, date] = {}
    for ci, v in enumerate(date_row):
        if ci < 3:  # A, B, C — не даты
            continue
        if isinstance(v, datetime):
            col_to_date[ci] = v.date()
        elif isinstance(v, date):
            col_to_date[ci] = v

    if not col_to_date:
        return []

    records: list[dict] = []
    current_assembler: dict | None = None  # {name, zone}

    for ri, row in enumerate(rows):
        if ri <= date_row_idx:
            continue

        col_a = str(row[0] or "").strip().lower()
        col_b = str(row[1] or "").strip()
        col_c = str(row[2] or "").strip().lower()

        # Строка сборщика
        if col_b and col_c in ("сборка", "сборка "):
            name = _extract_name(col_b)
            zone = col_a or ""
            current_assembler = {"name": name, "zone": zone}

        elif col_c in ("замер", "замер ", "только замеры"):
            # Строка замеров — пропускаем (не относится к стоимости)
            pass

        else:
            current_assembler = None  # Разрыв блока
            continue

        if not current_assembler:
            continue

        # Собираем заказы из ячеек (cols D+)
        for ci, cell_val in enumerate(row):
            if ci < 3:
                continue
            if ci not in col_to_date:
                continue
            text = str(cell_val or "").strip()
            if not text:
                continue
            amount = _extract_amount(text)
            if amount == 0 and not text:
                continue
            records.append({
                "assembler_name": current_assembler["name"],
                "assembler_zone": current_assembler["zone"],
                "date": col_to_date[ci].isoformat(),
                "order_text": text[:120],
                "amount": amount,
                "sheet": title,
            })

    return records


def parse_file(xlsx_path: str, sheets_filter: list[str] | None = None) -> dict[str, Any]:
    """
    Парсит Excel файл. Кэширует по mtime.
    sheets_filter: список названий листов для парсинга; None = все.
    """
    if openpyxl is None:
        return {"error": "openpyxl not installed", "records": []}

    path = str(xlsx_path)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return {"error": f"file_not_found: {path}", "records": []}

    if path in _cache and _cache[path]["mtime"] == mtime:
        return _cache[path]["data"]

    log.info("Parsing assembler schedule: %s", path)
    t0 = time.time()
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return {"error": str(e), "records": []}

    all_records: list[dict] = []
    parsed_sheets = []
    for sname in wb.sheetnames:
        if sheets_filter and sname not in sheets_filter:
            continue
        # Пропускаем служебные листы
        if sname.lower().startswith("лист") or sname.lower() == "образец":
            continue
        try:
            ws = wb[sname]
            recs = parse_sheet(ws)
            all_records.extend(recs)
            parsed_sheets.append({"sheet": sname, "records": len(recs)})
        except Exception as e:
            log.warning("Sheet %s parse error: %s", sname, e)

    wb.close()
    elapsed = round(time.time() - t0, 2)
    log.info("Parsed %d records in %.2fs", len(all_records), elapsed)

    data = {
        "records": all_records,
        "parsed_sheets": parsed_sheets,
        "elapsed_s": elapsed,
        "parsed_at": datetime.utcnow().isoformat(),
    }
    _cache[path] = {"mtime": mtime, "data": data}
    return data


def aggregate(records: list[dict]) -> dict[str, Any]:
    """
    Агрегирует записи по сборщику и месяцу.
    Возвращает:
      by_assembler: {name: {year_month: {orders, total_amount}}}
      by_month:     {year_month: {total_amount, order_count, assemblers}}
    """
    by_assembler: dict[str, dict] = {}
    by_month: dict[str, dict] = {}

    for r in records:
        name = r["assembler_name"]
        dt = r["date"][:7]  # YYYY-MM
        amount = r["amount"]

        # by_assembler
        if name not in by_assembler:
            by_assembler[name] = {}
        if dt not in by_assembler[name]:
            by_assembler[name][dt] = {"orders": 0, "total_amount": 0, "zone": r.get("assembler_zone", "")}
        by_assembler[name][dt]["orders"] += 1
        by_assembler[name][dt]["total_amount"] += amount

        # by_month
        if dt not in by_month:
            by_month[dt] = {"total_amount": 0, "order_count": 0, "assemblers": set()}
        by_month[dt]["total_amount"] += amount
        by_month[dt]["order_count"] += 1
        by_month[dt]["assemblers"].add(name)

    # Сериализуем sets
    for k in by_month:
        by_month[k]["assemblers"] = sorted(by_month[k]["assemblers"])

    return {"by_assembler": by_assembler, "by_month": by_month}
