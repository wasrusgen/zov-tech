"""ЗОВ Backend — FastAPI app. Полный порт Apps Script Code.gs."""
from __future__ import annotations
import base64
import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timezone, timedelta
import secrets
from pathlib import Path
from typing import Any
from fastapi import FastAPI, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse

from .config import get_config
from .auth import verify_init_data
from . import sheets, ai, telegram as tg, proxy_pool, catalog, geocoder, drive
from . import parsers
from . import proposals as proposals_mod
from . import assembler_parser
from .parsers import dns as parser_dns, wb as parser_wb, ozon as parser_ozon, yamarket as parser_ym, citilink as parser_cl

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
log = logging.getLogger("zov.backend")

app = FastAPI(title="ZOV Backend", version="2.0")

# Каталог под фото замеров (монтируется как volume в docker-compose)
PHOTOS_DIR = Path(os.environ.get("PHOTOS_DIR", "/app/photos"))
try:
    PHOTOS_DIR.mkdir(parents=True, exist_ok=True)
except Exception as _e:
    logging.getLogger("zov.backend").warning("Не удалось создать PHOTOS_DIR=%s: %s", PHOTOS_DIR, _e)

_SAFE_ID_RE = re.compile(r"^[A-Za-z0-9_\-]{1,64}$")
_SAFE_FILE_RE = re.compile(r"^[A-Za-z0-9_\-.]{1,80}$")

# CORS — MiniApp хостится на github.io, бэкенд на api.wasrusgen1.pro.
# Простые запросы (text/plain или без Content-Type) не триггерят preflight.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)


# =================================================================
# Startup: канонизация схем таблиц
# =================================================================

@app.on_event("startup")
async def _on_startup() -> None:
    """При запуске бэкенда канонизируем схему Measurements один раз.
    Это исправляет рассинхронизацию порядка колонок без ручного вмешательства."""
    import asyncio
    try:
        await asyncio.to_thread(_ensure_measurements_sheet)
        log.info("Startup: Measurements schema OK")
    except Exception as e:
        log.warning("Startup: Measurements schema check failed (non-fatal): %s", e)


# =================================================================
# Health & ping
# =================================================================

@app.get("/healthz")
async def healthz():
    return {"ok": True, "service": "zov-tech-backend", "time": _now_iso()}


@app.get("/")
async def root():
    return {
        "status": "ok",
        "service": "zov-tech-backend",
        "version": "2.0",
        "available_paths": ["me", "measurement", "podbor", "ping", "test_ai", "test_telegram", "seed_admin"],
    }


# =================================================================
# Compatibility layer with Apps Script (?path=X)
# =================================================================

@app.post("/")
async def root_post(request: Request):
    return await _dispatch_post(request)


@app.api_route("/exec", methods=["GET", "POST"])
async def apps_script_compat(request: Request):
    """Эмулируем поведение `/exec?path=X` чтобы старый MiniApp-код тоже работал."""
    if request.method == "GET":
        path = request.query_params.get("path", "")
        if path == "ping":
            return {"pong": True, "time": _now_iso()}
        if path == "seed_admin":
            return JSONResponse(_handle_seed_admin())
        if path == "test_ai" or path == "test_claude":
            return JSONResponse(_handle_test_ai())
        if path == "test_telegram":
            return JSONResponse(_handle_test_telegram())
        return {"status": "ok", "service": "zov-tech-backend"}
    return await _dispatch_post(request)


async def _dispatch_post(request: Request):
    path = request.query_params.get("path", "")
    try:
        body = await request.json()
    except Exception:
        body = {}

    handlers = {
        "me":            _handle_me,
        "measurement":   _handle_measurement,
        "measurements":  _handle_measurements_list,
        "measurement_detail": _handle_measurement_detail,
        "podbor":        _handle_podbor,
        "clients":       _handle_clients,
        "lead":          _handle_lead,
        "grant_role":    _handle_grant_role,
        "staff_list":    _handle_staff_list,
        "measurement_request":   _handle_measurement_request,
        "measurement_inbox":     _handle_measurement_inbox,
        "measurement_schedule":  _handle_measurement_schedule,
        "measurement_next_no":   _handle_measurement_next_no,
        "measurement_logistics": _handle_measurement_logistics,
        "geocode":               _handle_geocode,
        "client_note":           _handle_client_note,
        "client_create":         _handle_client_create,
        "client_update":         _handle_client_update,
        "client_delete":         _handle_client_delete,
        "measurement_design_upload": _handle_measurement_design_upload,
        "measurement_add_photos":    _handle_measurement_add_photos,
        "measurement_decision":  _handle_measurement_decision,
        "measurement_set_status": _handle_measurement_set_status,
        "manager_pending":       _handle_manager_pending,
        "assembly_create":       _handle_assembly_create,
        "assembly_list":         _handle_assembly_list,
        "assembly_detail":       _handle_assembly_detail,
        "assembly_set_kitchen_price": _handle_assembly_set_kitchen_price,
        "sign_request_create":   _handle_sign_request_create,
        "sign_request_submit":   _handle_sign_request_submit,
        "managers_list":         _handle_managers_list,
        "assembly_rates_list":   _handle_assembly_rates_list,
        "assembly_rate_save":    _handle_assembly_rate_save,
        "assembly_rate_delete":  _handle_assembly_rate_delete,
        "assembler_analytics":   _handle_assembler_analytics,
        "assembler_earnings":    _handle_assembler_earnings,
        "staff_clients":         _handle_staff_clients,
        "assembly_schedule":     _handle_assembly_schedule,
        "measurement_schedule":  _handle_measurement_schedule,
        "contract_preview":      _handle_contract_preview,
        "contract_save":         _handle_contract_save,
        "invoice_create":        _handle_invoice_create,
        "equipment_save":        _handle_equipment_save,
        "measurer_earnings":     _handle_measurer_earnings,
        "assembler_client_podbor": _handle_assembler_client_podbor,
        "act4_preview":          _handle_act4_preview,
        "act4_save":             _handle_act4_save,
        "assembly_set_status":   _handle_assembly_set_status,
        "assembly_set_expeditor": _handle_assembly_set_expeditor,
        "assembly_photo_upload":    _handle_assembly_photo_upload,
        "assembler_set_probation":  _handle_assembler_set_probation,
        "assembly_notes_save":      _handle_assembly_notes_save,
        "assembly_invoice_create":  _handle_assembly_invoice_create,
        "assembly_extras_list":     _handle_assembly_extras_list,
        "assembly_extra_add":       _handle_assembly_extra_add,
        "assembly_extra_delete":    _handle_assembly_extra_delete,
        "assembly_extra_approve":   _handle_assembly_extra_approve,
        "assembly_receipt_parse":   _handle_assembly_receipt_parse,
        "staff_roster":             _handle_staff_roster,
        "client_order_timeline":    _handle_client_order_timeline,
        "manager_finance_summary":  _handle_manager_finance_summary,
        "feedback_submit":           _handle_feedback_submit,
        "feedback_my":              _handle_feedback_my,
        "assembly_suggest_slots":   _handle_assembly_suggest_slots,
        "assembly_propose_date":    _handle_assembly_propose_date,
        "assembly_date_confirm":    _handle_assembly_date_confirm,
        "assembly_date_decline":    _handle_assembly_date_decline,
        "proposal_brief":        proposals_mod.handle_brief,
        "proposal_create":       proposals_mod.handle_create,
        "proposal_upsert_variant": proposals_mod.handle_upsert_variant,
        "proposal_remove_variant": proposals_mod.handle_remove_variant,
        "proposal_send":         proposals_mod.handle_send,
        "proposal_list":         proposals_mod.handle_list,
        "proposal_detail":       proposals_mod.handle_detail,
        "proposal_vote":         proposals_mod.handle_vote,
        "proposal_client_submit": proposals_mod.handle_client_submit,
        "contract_review":        _handle_contract_review,
        "ping":          lambda b: {"pong": True, "time": _now_iso()},
        "seed_admin":    lambda b: _handle_seed_admin(),
        "test_ai":       lambda b: _handle_test_ai(),
        "test_claude":   lambda b: _handle_test_ai(),
        "test_telegram": lambda b: _handle_test_telegram(),
    }
    fn = handlers.get(path)
    if not fn:
        return JSONResponse({"error": "unknown_path", "path": path}, status_code=404)

    try:
        # podbor использует Playwright sync API → выполняем в threadpool
        if path == "podbor":
            import asyncio
            result = await asyncio.to_thread(fn, body)
        else:
            result = fn(body)
        return JSONResponse(result)
    except Exception as e:
        log.exception("api error on path=%s", path)
        sheets.log_event("api_error", None, {"path": path, "error": str(e)})
        return JSONResponse({"error": str(e)}, status_code=500)


# =================================================================
# Native /api/* routes (preferred for new MiniApp)
# =================================================================

@app.post("/api/me")
async def api_me(request: Request):
    body = await _safe_json(request)
    return _handle_me(body)


@app.post("/api/measurement")
async def api_measurement(request: Request):
    body = await _safe_json(request)
    return _handle_measurement(body)


@app.post("/api/podbor")
async def api_podbor(request: Request):
    body = await _safe_json(request)
    # _handle_podbor использует Playwright sync API — выполняем в threadpool
    import asyncio
    return await asyncio.to_thread(_handle_podbor, body)


@app.post("/api/clients")
async def api_clients(request: Request):
    body = await _safe_json(request)
    return _handle_clients(body)


@app.post("/api/lead")
async def api_lead(request: Request):
    body = await _safe_json(request)
    return _handle_lead(body)


@app.post("/api/measurements")
async def api_measurements(request: Request):
    body = await _safe_json(request)
    return _handle_measurements_list(body)


@app.post("/api/measurement_detail")
async def api_measurement_detail(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_detail(body)


@app.post("/api/measurement_request")
async def api_measurement_request(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_request(body)


@app.post("/api/measurement_inbox")
async def api_measurement_inbox(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_inbox(body)


@app.post("/api/measurement_schedule")
async def api_measurement_schedule(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_schedule(body)


@app.post("/api/measurement_next_no")
async def api_measurement_next_no(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_next_no(body)


@app.post("/api/measurement_logistics")
async def api_measurement_logistics(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_logistics(body)


@app.post("/api/geocode")
async def api_geocode(request: Request):
    body = await _safe_json(request)
    return _handle_geocode(body)


@app.post("/api/client_note")
async def api_client_note(request: Request):
    body = await _safe_json(request)
    return _handle_client_note(body)


@app.post("/api/client_create")
async def api_client_create(request: Request):
    body = await _safe_json(request)
    try:
        return _handle_client_create(body)
    except Exception as e:
        log.exception("client_create error")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/client_delete")
async def api_client_delete(request: Request):
    body = await _safe_json(request)
    try:
        return _handle_client_delete(body)
    except Exception as e:
        log.exception("client_delete error")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/client_update")
async def api_client_update(request: Request):
    body = await _safe_json(request)
    try:
        return _handle_client_update(body)
    except Exception as e:
        log.exception("client_update error")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/measurement_design_upload")
async def api_measurement_design_upload(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_design_upload(body)


@app.post("/api/measurement_decision")
async def api_measurement_decision(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_decision(body)


@app.post("/api/measurement_set_status")
async def api_measurement_set_status(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_set_status(body)


@app.post("/api/measurement_add_photos")
async def api_measurement_add_photos(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_add_photos(body)


@app.post("/api/manager_pending")
async def api_manager_pending(request: Request):
    body = await _safe_json(request)
    return _handle_manager_pending(body)


@app.post("/api/assembly_create")
async def api_assembly_create(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_create(body)


@app.post("/api/assembly_list")
async def api_assembly_list(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_list(body)


@app.post("/api/assembly_detail")
async def api_assembly_detail(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_detail(body)


@app.post("/api/assembly_set_kitchen_price")
async def api_assembly_set_kitchen_price(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_set_kitchen_price(body)


@app.post("/api/assembly_rates_list")
async def api_assembly_rates_list(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_rates_list(body)


@app.post("/api/assembly_rate_save")
async def api_assembly_rate_save(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_rate_save(body)


@app.post("/api/assembly_rate_delete")
async def api_assembly_rate_delete(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_rate_delete(body)


@app.post("/api/grant_role")
async def api_grant_role(request: Request):
    """Админ выдаёт роль другому пользователю.
    body: {initData, target_tg_id, role: 'measurer'|'assembler'|'manager'|'client', action: 'grant'|'revoke'}"""
    body = await _safe_json(request)
    return _handle_grant_role(body)


@app.post("/api/staff_list")
async def api_staff_list(request: Request):
    body = await _safe_json(request)
    return _handle_staff_list(body)


@app.get("/api/photo/{measurement_id}/{filename}")
async def api_photo(measurement_id: str, filename: str):
    """Отдаёт фото замера. Защита от path traversal — только разрешённые id и имена."""
    if not _SAFE_ID_RE.match(measurement_id) or not _SAFE_FILE_RE.match(filename):
        return JSONResponse({"error": "bad_path"}, status_code=400)
    if filename.startswith(".") or ".." in filename:
        return JSONResponse({"error": "bad_path"}, status_code=400)
    p = PHOTOS_DIR / measurement_id / filename
    try:
        p_resolved = p.resolve()
        if PHOTOS_DIR.resolve() not in p_resolved.parents:
            return JSONResponse({"error": "bad_path"}, status_code=400)
    except Exception:
        return JSONResponse({"error": "bad_path"}, status_code=400)
    if not p.exists() or not p.is_file():
        return JSONResponse({"error": "not_found"}, status_code=404)
    ext = filename.rsplit(".", 1)[-1].lower()
    media = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "png": "image/png", "webp": "image/webp",
        "pdf": "application/pdf",
        "dwg": "application/acad",
        "dxf": "application/dxf",
    }.get(ext, "application/octet-stream")
    return FileResponse(str(p), media_type=media)


@app.get("/api/daily_reminders")
async def api_daily_reminders(request: Request):
    """Внутренний эндпоинт для бота: клиенты с годовщиной договора сегодня (МСК).
    Защищён через заголовок Authorization: Bearer <INTERNAL_SECRET>."""
    cfg = get_config()
    secret = cfg.internal_secret
    auth_header = request.headers.get("Authorization", "")
    if not secret or auth_header != f"Bearer {secret}":
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    return JSONResponse(_handle_daily_reminders())


@app.post("/api/shipments")
async def api_shipments(request: Request):
    """Отгрузки из ОТГРУЗКИ.xlsx (Google Drive). Только для менеджера."""
    body = await _safe_json(request)
    import asyncio
    return JSONResponse(await asyncio.to_thread(_handle_shipments, body))


@app.post("/api/arrivals")
async def api_arrivals(request: Request):
    """Поступления на склад СПб из «Поступление заказов на склад СПб.xlsx». Только для менеджера."""
    body = await _safe_json(request)
    import asyncio
    return JSONResponse(await asyncio.to_thread(_handle_arrivals, body))


@app.post("/api/proposal_brief")
async def api_proposal_brief(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_brief(body))

@app.post("/api/proposal_create")
async def api_proposal_create(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_create(body))

@app.post("/api/proposal_upsert_variant")
async def api_proposal_upsert_variant(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_upsert_variant(body))

@app.post("/api/proposal_remove_variant")
async def api_proposal_remove_variant(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_remove_variant(body))

@app.post("/api/proposal_send")
async def api_proposal_send(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_send(body))

@app.post("/api/proposal_list")
async def api_proposal_list(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_list(body))

@app.post("/api/proposal_detail")
async def api_proposal_detail(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_detail(body))

@app.post("/api/proposal_vote")
async def api_proposal_vote(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_vote(body))

@app.post("/api/proposal_client_submit")
async def api_proposal_client_submit(request: Request):
    body = await _safe_json(request)
    return JSONResponse(proposals_mod.handle_client_submit(body))


@app.post("/api/contract_review")
async def api_contract_review(request: Request):
    """AI-анализ текста договора. GigaChat → структурированный разбор на русском."""
    body = await _safe_json(request)
    import asyncio
    return JSONResponse(await asyncio.to_thread(_handle_contract_review, body))


def _handle_daily_reminders() -> dict[str, Any]:
    """Находит клиентов с годовщиной договора сегодня по МСК.
    Дедуплицирует: один менеджер + один клиент = одно уведомление,
    даже если по клиенту несколько строк в Measurements."""
    from datetime import timedelta
    moscow_now = datetime.now(timezone.utc) + timedelta(hours=3)
    today_md = (moscow_now.month, moscow_now.day)
    current_year = moscow_now.year

    reminders: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()  # (manager_tg_id, client_key)

    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception as e:
        log.exception("daily_reminders: не удалось прочитать Measurements")
        return {"error": str(e)}

    if not rows or len(rows) < 2:
        return {"ok": True, "reminders": [], "date": moscow_now.strftime("%d.%m.%Y")}

    headers = rows[0]
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * max(0, len(headers) - len(r))))

        if row.get("archived_at"):
            continue

        contract_date_raw = (row.get("contract_date") or "").strip()
        if not contract_date_raw:
            continue

        manager_tg_id = (row.get("manager_tg_id") or "").strip()
        if not manager_tg_id:
            continue

        # Парсим дату договора: ISO YYYY-MM-DD или DD.MM.YYYY
        cd = None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                cd = datetime.strptime(contract_date_raw[:10], fmt)
                break
            except ValueError:
                continue
        if cd is None:
            continue

        # Годовщина — месяц и день совпадают с сегодня
        if (cd.month, cd.day) != today_md:
            continue

        # Договор должен быть из прошлых лет, не из нынешнего
        if cd.year >= current_year:
            continue

        client_tg_id = (row.get("client_tg_id") or "").strip()
        client_name = (row.get("client_name") or "Без имени").strip()
        client_key = client_tg_id or client_name.lower()

        dedup_key = (manager_tg_id, client_key)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        years = current_year - cd.year
        reminders.append({
            "manager_tg_id": manager_tg_id,
            "client_name": client_name,
            "contract_date": contract_date_raw,
            "years": years,
        })

    log.info("daily_reminders: %d годовщин на %s", len(reminders), moscow_now.strftime("%d.%m.%Y"))
    return {"ok": True, "reminders": reminders, "date": moscow_now.strftime("%d.%m.%Y")}


@app.get("/api/test_ai")
async def api_test_ai():
    return _handle_test_ai()


@app.get("/api/test_telegram")
async def api_test_telegram():
    return _handle_test_telegram()


@app.get("/api/seed_admin")
async def api_seed_admin():
    return _handle_seed_admin()


@app.get("/api/parse_dns")
def api_parse_dns(q: str = "", limit: int = 1):  # sync — для threadpool
    """Тест парсера DNS."""
    if not q:
        return {"error": "missing_query", "hint": "use ?q=<search>"}
    try:
        results = parser_dns.search_dns(q, limit=min(max(1, limit), 5))
        return {"ok": True, "query": q, "count": len(results), "results": results}
    except Exception as e:
        return {"ok": False, "error": str(e), "query": q}


@app.get("/api/parse_wb")
def api_parse_wb(q: str = "", limit: int = 3):
    if not q:
        return {"error": "missing_query"}
    try:
        results = parser_wb.search_wb(q, limit=min(max(1, limit), 10))
        return {"ok": True, "query": q, "count": len(results), "results": results}
    except Exception as e:
        return {"ok": False, "error": str(e), "query": q}


@app.get("/api/parse_ozon")
def api_parse_ozon(q: str = "", limit: int = 3):
    if not q:
        return {"error": "missing_query"}
    try:
        results = parser_ozon.search_ozon(q, limit=min(max(1, limit), 10))
        return {"ok": True, "query": q, "count": len(results), "results": results}
    except Exception as e:
        return {"ok": False, "error": str(e), "query": q}


@app.get("/api/parse_yamarket")
def api_parse_yamarket(q: str = "", limit: int = 3):
    if not q:
        return {"error": "missing_query"}
    try:
        results = parser_ym.search_yamarket(q, limit=min(max(1, limit), 10))
        return {"ok": True, "query": q, "count": len(results), "results": results}
    except Exception as e:
        return {"ok": False, "error": str(e), "query": q}


@app.get("/api/parse_citilink")
def api_parse_citilink(q: str = "", limit: int = 3):
    if not q:
        return {"error": "missing_query"}
    try:
        results = parser_cl.search_citilink(q, limit=min(max(1, limit), 10))
        return {"ok": True, "query": q, "count": len(results), "results": results}
    except Exception as e:
        return {"ok": False, "error": str(e), "query": q}


@app.get("/api/parse_all")
def api_parse_all(q: str = ""):
    """Спрашивает все источники и возвращает агрегированный результат."""
    if not q:
        return {"error": "missing_query"}
    try:
        data = parsers.enrich_one(q)
        return {"ok": True, "query": q, "data": data}
    except Exception as e:
        return {"ok": False, "error": str(e), "query": q}


@app.get("/api/proxy_status")
async def api_proxy_status():
    """Диагностика: показывает текущее состояние proxy-пула."""
    return proxy_pool.pool_status()


from fastapi import BackgroundTasks


_CATALOG_REFRESH_STATUS = {"running": False, "last_result": None, "started_at": None}


def _bg_refresh(categories, per_brand, delay):
    """Фоновая задача обновления каталога — пишет статус в глобал."""
    import datetime as _dt
    _CATALOG_REFRESH_STATUS["running"] = True
    _CATALOG_REFRESH_STATUS["started_at"] = _dt.datetime.now(_dt.timezone.utc).isoformat()
    try:
        result = catalog.refresh_catalog(
            categories=categories,
            per_brand=per_brand,
            delay_sec=delay,
        )
        _CATALOG_REFRESH_STATUS["last_result"] = result
    except Exception as e:
        log.exception("bg catalog refresh failed")
        _CATALOG_REFRESH_STATUS["last_result"] = {"ok": False, "error": str(e)}
    finally:
        _CATALOG_REFRESH_STATUS["running"] = False


@app.post("/api/catalog/refresh")
def api_catalog_refresh(background: BackgroundTasks,
                        cat: str = "", per_brand: int = 2, delay: float = 1.0):
    """Запускает refresh в ФОНЕ. Возвращает сразу, статус смотри в /api/catalog/refresh_status.

    Параметры:
      cat: одна категория или пусто = все 8 (очень долго)
      per_brand: сколько моделей на (brand × category) — default 2
      delay: задержка между запросами, сек — default 1.0
    """
    if _CATALOG_REFRESH_STATUS["running"]:
        return {"ok": False, "error": "already running", "started_at": _CATALOG_REFRESH_STATUS["started_at"]}

    categories = [cat] if cat else None
    background.add_task(
        _bg_refresh,
        categories,
        max(1, min(per_brand, 5)),
        max(0.0, min(delay, 10.0)),
    )
    return {
        "ok": True,
        "queued": True,
        "categories": categories or "all",
        "hint": "GET /api/catalog/refresh_status — узнать прогресс",
    }


@app.get("/api/catalog/refresh_status")
def api_catalog_refresh_status():
    """Статус последнего/текущего refresh'а каталога."""
    return _CATALOG_REFRESH_STATUS


@app.post("/api/catalog/clear")
def api_catalog_clear(cat: str = ""):
    """Удаляет всё содержимое каталога (или одной категории)."""
    removed = catalog.clear_catalog(category=cat or None)
    return {"ok": True, "removed": removed, "category": cat or "all"}


@app.get("/api/catalog/list")
def api_catalog_list(cat: str = "", tier: str = "", brand: str = "", limit: int = 100):
    """Читает каталог моделей из Sheets с фильтрами."""
    items = catalog.list_catalog(
        category=cat or None,
        tier=tier or None,
        brand=brand or None,
        limit=min(limit, 500),
    )
    return {
        "ok": True,
        "filters": {"category": cat, "tier": tier, "brand": brand},
        "count": len(items),
        "items": items,
    }


@app.get("/api/catalog/preview_ai")
def api_catalog_preview_ai(cats: str = "fridge", tiers: str = ""):
    """Превью того, что AI получит в prompt (для отладки)."""
    cat_list = [c.strip() for c in cats.split(",") if c.strip()]
    tier_list = [t.strip() for t in tiers.split(",") if t.strip()] or None
    text = catalog.list_for_ai(cat_list, tiers=tier_list, limit_per_cat=30)
    return {"text": text, "length_chars": len(text)}


# =================================================================
# Handlers
# =================================================================

def _handle_me(body: dict[str, Any]) -> dict[str, Any]:
    cfg = get_config()
    init_data = body.get("initData") or ""
    auth = verify_init_data(init_data, cfg.bot_token)

    # Fallback для Telegram Desktop side-panel — initData может приходить пустым.
    # Доверяем initDataUnsafe.user (НЕпроверенным данным) — только для UI-режима.
    # Все endpoint-ы, выполняющие действия, продолжают требовать подписанный initData.
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        unsafe_user = unsafe.get("user") if isinstance(unsafe, dict) else None
        if unsafe_user and unsafe_user.get("id"):
            auth = {
                "user": unsafe_user,
                "auth_date": int(time.time()),
                "start_param": unsafe.get("start_param"),
                "_unsafe": True,
            }
        else:
            return {"error": "invalid_init_data"}

    tg_user = auth["user"]
    tg_id = tg_user["id"]
    start_param = body.get("startParam") or auth.get("start_param")
    explicit_role = body.get("role") if body.get("role") in ("manager", "client", "staff") else None
    user = sheets.get_or_create_user(tg_user, start_param, explicit_role)
    # Берём roles из словаря если они уже распарсены (после grant_role),
    # иначе fallback на парсинг сырой CSV-колонки
    roles = user.get("roles") or sheets.parse_roles(user.get("role", ""))

    # Staff (замерщик / сборщик) — отдельный кабинет, доступен только тем у кого роль выдана
    if explicit_role == "staff":
        has_measurer  = "measurer"  in roles
        has_assembler = "assembler" in roles
        has_expeditor = "expeditor" in roles
        if not (has_measurer or has_assembler or has_expeditor):
            return {
                "role": "staff",
                "roles": roles,
                "error": "no_staff_role",
                "user": {
                    "tg_id": tg_id,
                    "full_name": user.get("full_name", ""),
                    "avatar_initial": _initial(user.get("full_name") or tg_user.get("first_name", "")),
                },
            }
        full_name = user.get("full_name", "") or tg_user.get("first_name", "")
        # Оборудование замерщика
        equipment_raw = user.get("equipment", "")
        equipment_list = [x.strip() for x in equipment_raw.split(",") if x.strip()] if equipment_raw else []
        equipment_ok = _equipment_complete(equipment_list) if has_measurer else True
        return {
            "role": "staff",
            "roles": roles,
            "user": {
                "tg_id": tg_id,
                "full_name": full_name,
                "avatar_initial": _initial(full_name),
            },
            "capabilities": {
                "measurer":  has_measurer,
                "assembler": has_assembler,
                "expeditor": has_expeditor,
                "dispatcher": sheets.has_role(user, "dispatcher"),
            },
            "equipment": equipment_list,
            "equipment_ok": equipment_ok,
        }

    if "manager" in roles:
        m = sheets.get_manager_profile(tg_id) or {
            "full_name": user.get("full_name", ""), "salon": "",
            "is_zov_employee": False, "status": "lapsed", "active_until": None,
        }
        return {
            "role": "manager",
            "roles": roles,
            "user": {
                "tg_id": tg_id,
                "full_name": m.get("full_name") or user.get("full_name", ""),
                "salon": m.get("salon", ""),
                "avatar_initial": _initial(m.get("full_name") or tg_user.get("first_name", "")),
            },
            "status": m.get("status", "lapsed"),
            "status_until": _format_date(m.get("active_until")),
        }

    # client
    c = sheets.get_client_profile(tg_id) or {}
    manager = None
    mgr_id = c.get("manager_tg_id")
    if mgr_id:
        try:
            mgr_id_int = int(mgr_id)
            mp = sheets.get_manager_profile(mgr_id_int)
            if mp:
                manager = {"full_name": mp.get("full_name"), "salon": mp.get("salon")}
        except (TypeError, ValueError):
            pass

    full_name = c.get("full_name") or user.get("full_name", "")
    return {
        "role": "client",
        "roles": roles,
        "user": {
            "tg_id": tg_id,
            "full_name": full_name,
            "avatar_initial": _initial(full_name or tg_user.get("first_name", "")),
        },
        "manager": manager,
    }


# Обязательный набор оборудования замерщика (ключи)
EQUIPMENT_REQUIRED = {"tablet", "laser_tape", "angle_meter", "tape", "laser_level"}


def _equipment_complete(equipment_list: list[str]) -> bool:
    return EQUIPMENT_REQUIRED.issubset(set(equipment_list))


def _handle_equipment_save(body: dict[str, Any]) -> dict[str, Any]:
    """Замерщик сохраняет свой набор оборудования.
    body: {initData, equipment: ["tablet","laser_tape",...]}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not sheets.has_role(user, "measurer"):
        return {"error": "only_measurer"}

    raw = body.get("equipment") or []
    if not isinstance(raw, list):
        return {"error": "invalid_equipment"}
    # Принимаем только известные ключи
    valid_keys = {"tablet", "laser_tape", "angle_meter", "tape", "laser_level"}
    clean = [k for k in raw if k in valid_keys]
    equipment_str = ",".join(clean)

    # Убедимся что колонка equipment есть в Users
    try:
        ws = sheets.sheet("Users")
        headers = ws.row_values(1)
        if "equipment" not in headers:
            ws.update_cell(1, len(headers) + 1, "equipment")
    except Exception as e:
        log.warning("equipment col ensure error: %s", e)

    sheets.update_cell_by_key("Users", "tg_id", tg_id, "equipment", equipment_str)
    equipment_ok = _equipment_complete(clean)
    return {"ok": True, "equipment": clean, "equipment_ok": equipment_ok}


@app.post("/api/equipment_save")
async def api_equipment_save(request: Request):
    body = await _safe_json(request)
    return _handle_equipment_save(body)


# =================================================================
# Заработки замерщика — по листу Measurements
# =================================================================

def _handle_measurer_earnings(body: dict[str, Any]) -> dict[str, Any]:
    """Личная статистика замерщика: количество замеров и сумма по месяцам.
    body: {initData, year?}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "measurer") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    year_filter = str(body.get("year") or "").strip()

    _ensure_measurements_sheet()
    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception as e:
        return {"error": f"sheet_error: {e}"}

    if not rows or len(rows) < 2:
        return {"ok": True, "months": {}, "total_amount": 0, "total_measurements": 0}

    headers = rows[0]
    months: dict[str, dict] = {}

    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if str(row.get("assigned_to_tg_id", "")) != str(tg_id):
            continue
        if row.get("archived_at"):
            continue

        # Дата — из scheduled_at или zamer_date или ts
        date_str = row.get("scheduled_at") or row.get("zamer_date") or row.get("ts") or ""
        if not date_str:
            continue
        try:
            ym = date_str[:7]  # "2026-05"
            if year_filter and not ym.startswith(year_filter):
                continue
        except Exception:
            continue

        fee_raw = row.get("measurement_fee", "")
        try:
            fee = float(fee_raw) if fee_raw else 0.0
        except (ValueError, TypeError):
            fee = 0.0

        status = row.get("status", "")
        if ym not in months:
            months[ym] = {"total_amount": 0.0, "measurements": 0, "paid": 0}
        months[ym]["measurements"] += 1
        months[ym]["total_amount"] += fee
        if fee > 0:
            months[ym]["paid"] += 1

    total_amount = sum(m["total_amount"] for m in months.values())
    total_meas = sum(m["measurements"] for m in months.values())
    months_sorted = dict(sorted(months.items(), reverse=True))

    return {
        "ok": True,
        "months": months_sorted,
        "total_amount": total_amount,
        "total_measurements": total_meas,
    }


@app.post("/api/measurer_earnings")
async def api_measurer_earnings(request: Request):
    body = await _safe_json(request)
    return _handle_measurer_earnings(body)


# =================================================================
# Подбор техники для сборщика — по клиенту из замера
# =================================================================

def _handle_assembler_client_podbor(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает сводку выбранной техники для сборщика/замерщика.
    body: {initData, measurement_id}
    Доступно: назначенный замерщик/сборщик, менеджер."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    is_staff = sheets.has_role(user, "assembler") or sheets.has_role(user, "measurer")
    is_manager = sheets.has_role(user, "manager")
    if not (is_staff or is_manager):
        return {"error": "forbidden"}

    measurement_id = (body.get("measurement_id") or "").strip()
    if not measurement_id:
        return {"error": "missing_measurement_id"}

    _ensure_measurements_sheet()
    mrow = sheets.find_row("Measurements", "id", measurement_id)
    if not mrow:
        return {"error": "measurement_not_found"}

    # Проверка доступа: только назначенный или менеджер
    if is_staff and not is_manager:
        if str(mrow.get("assigned_to_tg_id", "")) != str(tg_id):
            return {"error": "not_assigned"}

    podbor_lead_id = (mrow.get("podbor_lead_id") or "").strip()
    client_name = mrow.get("client_name", "")
    client_phone = mrow.get("client_phone", "")

    if not podbor_lead_id:
        return {"ok": True, "has_podbor": False, "client_name": client_name}

    # Ищем proposal по client_key
    client_key = client_name.lower() if client_name else ""
    try:
        import proposals as proposals_mod
    except ImportError:
        from . import proposals as proposals_mod

    try:
        ws_p = sheets.sheet("Proposals")
        rows_p = ws_p.get_all_values()
    except Exception:
        return {"ok": True, "has_podbor": False, "client_name": client_name, "error_detail": "proposals_unavailable"}

    if not rows_p or len(rows_p) < 2:
        return {"ok": True, "has_podbor": False, "client_name": client_name}

    headers_p = rows_p[0]
    proposal = None
    for r in rows_p[1:]:
        rd = dict(zip(headers_p, r + [""] * (len(headers_p) - len(r))))
        if rd.get("client_key", "").lower() == client_key and rd.get("manager_tg_id") == str(mrow.get("manager_tg_id", "")):
            proposal = rd
            break

    if not proposal:
        return {"ok": True, "has_podbor": False, "client_name": client_name}

    # Парсим positions
    try:
        positions = json.loads(proposal.get("positions_json") or "[]")
    except (ValueError, TypeError):
        positions = []

    # Формируем сводку: только выбранные (voted yes) или все варианты
    summary = []
    for pos in positions:
        category = pos.get("label") or pos.get("category", "")
        variants = pos.get("variants") or []
        chosen = [v for v in variants if v.get("client_vote") == "yes"]
        if not chosen:
            # Если голосов нет — берём первый вариант как предложенный
            chosen = variants[:1]
        for v in chosen:
            summary.append({
                "category": category,
                "name": v.get("name") or v.get("title") or "",
                "price": v.get("price") or v.get("final_price") or 0,
                "image_url": v.get("image_url") or "",
                "voted": v.get("client_vote") == "yes",
            })

    return {
        "ok": True,
        "has_podbor": True,
        "client_name": client_name,
        "proposal_status": proposal.get("status", ""),
        "items": summary,
        "total_items": len(summary),
    }


@app.post("/api/assembler_client_podbor")
async def api_assembler_client_podbor(request: Request):
    body = await _safe_json(request)
    return _handle_assembler_client_podbor(body)


# =================================================================
# Акт №4 — приёмка товара (экспедитор)
# =================================================================

def _act4_columns() -> list[str]:
    return [
        "id", "assembly_id", "act_num", "act_date", "supplier",
        "items_json", "notes", "total_items", "damaged_count",
        "signed_by_name", "signed_by_phone", "signed_via", "signed_at",
        "signature_b64", "otp_code", "otp_expires_at",
        "created_at", "created_by_tg_id", "updated_at",
    ]


def _ensure_act4_sheet() -> None:
    want = _act4_columns()
    try:
        ws = sheets.sheet("Act4s")
        existing = ws.row_values(1)
        if not existing:
            ws.update("A1", [want])
            return
        missing = [c for c in want if c not in existing]
        if missing:
            ws.update("A1", [existing + missing])
    except Exception:
        sheets.ensure_sheet("Act4s", want)


def _handle_act4_preview(body: dict[str, Any]) -> dict[str, Any]:
    """Загружает данные для Акта №4.
    body: {initData, assembly_id}
    Доступно: expeditor, assembler, manager."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    is_exp  = sheets.has_role(user, "expeditor")
    is_asm  = sheets.has_role(user, "assembler")
    is_mgr  = sheets.has_role(user, "manager")
    if not (is_exp or is_asm or is_mgr):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    # Проверка доступа к сборке
    is_owner = (str(asm.get("manager_tg_id")) == str(tg_id) or
                str(asm.get("assigned_to_tg_id")) == str(tg_id) or
                is_mgr)
    if not is_owner and not is_exp:
        return {"error": "forbidden"}

    _ensure_act4_sheet()
    act4 = sheets.find_row("Act4s", "assembly_id", assembly_id)

    # Номер акта: asm-id + "-4" если не задан вручную
    default_act_num = f"{assembly_id}-4"
    default_date = _now_iso()[:10]

    return {
        "ok": True,
        "assembly_id": assembly_id,
        "client_name":  asm.get("client_name", ""),
        "client_phone": asm.get("client_phone", ""),
        "address":      asm.get("address", ""),
        "manager_tg_id": asm.get("manager_tg_id", ""),
        # Данные акта (если уже сохранён)
        "act_num":      act4.get("act_num", default_act_num) if act4 else default_act_num,
        "act_date":     act4.get("act_date", default_date)   if act4 else default_date,
        "supplier":     act4.get("supplier", "")             if act4 else "",
        "items":        json.loads(act4["items_json"]) if act4 and act4.get("items_json") else [],
        "notes":        act4.get("notes", "")                if act4 else "",
        "signed_by_name":  act4.get("signed_by_name", "")   if act4 else "",
        "signed_by_phone": act4.get("signed_by_phone", "")  if act4 else "",
        "signed_via":      act4.get("signed_via", "")        if act4 else "",
        "signed_at":       act4.get("signed_at", "")         if act4 else "",
        "is_signed":    bool(act4 and act4.get("signed_by_name")) if act4 else False,
    }


def _handle_act4_save(body: dict[str, Any]) -> dict[str, Any]:
    """Сохраняет / обновляет Акт №4.
    body: {initData, assembly_id, act_num, act_date, supplier, items, notes,
           signed_by_name?, signed_by_phone?, signed_via?}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "expeditor") or
            sheets.has_role(user, "assembler") or
            sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id  = (body.get("assembly_id") or "").strip()
    act_num      = (body.get("act_num") or f"{assembly_id}-4").strip()
    act_date     = (body.get("act_date") or _now_iso()[:10]).strip()
    supplier     = (body.get("supplier") or "").strip()
    notes        = (body.get("notes") or "").strip()
    items        = body.get("items") or []
    signed_by_name  = (body.get("signed_by_name") or "").strip()
    signed_by_phone = (body.get("signed_by_phone") or "").strip()
    signed_via      = (body.get("signed_via") or "").strip()

    if not isinstance(items, list):
        return {"error": "invalid_items"}

    # Подсчёт
    total_items   = sum(int(it.get("qty", 1)) for it in items)
    damaged_count = sum(int(it.get("qty", 1)) for it in items if it.get("condition") == "damaged")
    items_json    = json.dumps(items, ensure_ascii=False)
    now_iso       = _now_iso()

    _ensure_act4_sheet()
    existing = sheets.find_row("Act4s", "assembly_id", assembly_id)

    if existing:
        for col, val in [
            ("act_num", act_num), ("act_date", act_date), ("supplier", supplier),
            ("items_json", items_json), ("notes", notes),
            ("total_items", str(total_items)), ("damaged_count", str(damaged_count)),
            ("updated_at", now_iso),
        ]:
            sheets.update_cell_by_key("Act4s", "assembly_id", assembly_id, col, val)
        if signed_by_name:
            for col, val in [
                ("signed_by_name", signed_by_name),
                ("signed_by_phone", signed_by_phone),
                ("signed_via", signed_via or "manual"),
                ("signed_at", now_iso),
            ]:
                sheets.update_cell_by_key("Act4s", "assembly_id", assembly_id, col, val)
    else:
        act4_id = str(uuid.uuid4())[:8]
        signed_at = now_iso if signed_by_name else ""
        sheets.append_row("Act4s", [
            act4_id, assembly_id, act_num, act_date, supplier,
            items_json, notes, str(total_items), str(damaged_count),
            signed_by_name, signed_by_phone, signed_via or ("manual" if signed_by_name else ""),
            signed_at, now_iso, str(tg_id), now_iso,
        ])

    # Автоматика: подписали акт №4 → сборка переходит в in_progress
    if signed_by_name:
        try:
            _ensure_assemblies_sheet()
            asm = sheets.find_row("Assemblies", "id", assembly_id)
            if asm and asm.get("status") in ("created", "scheduled"):
                sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status", "in_progress")
                sheets.update_cell_by_key("Assemblies", "id", assembly_id, "started_at", now_iso)
                log.info("act4 signed → assembly %s in_progress", assembly_id)
            # Уведомить менеджера
            mgr_id = asm.get("manager_tg_id") if asm else None
            if mgr_id:
                dmg_text = f"⚠️ Повреждений: {damaged_count}" if damaged_count else "✅ Без повреждений"
                tg.send_message(int(mgr_id),
                    f"📦 <b>Акт №4 подписан — сборка началась</b>\n"
                    f"Сборка: <code>{assembly_id}</code>\n"
                    f"Клиент: {asm.get('client_name','')}\n"
                    f"Позиций: {total_items} · {dmg_text}\n"
                    f"Подписал: {signed_by_name}")
        except Exception as e:
            log.warning("act4 notify error: %s", e)

    return {"ok": True, "total_items": total_items, "damaged_count": damaged_count}


def _handle_expeditor_inbox(body):
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    is_exp = sheets.has_role(user, "expeditor")
    is_mgr = sheets.has_role(user, "manager")
    if not (is_exp or is_mgr):
        return {"error": "forbidden"}
    _ensure_assemblies_sheet()
    _ensure_act4_sheet()
    try:
        ws = sheets.sheet("Assemblies")
        rows = ws.get_all_values()
    except Exception:
        return {"ok": True, "assemblies": []}
    if not rows or len(rows) < 2:
        return {"ok": True, "assemblies": []}
    headers = rows[0]
    try:
        act_ws = sheets.sheet("Act4s")
        act_rows = act_ws.get_all_values()
        act_headers = act_rows[0] if act_rows else []
        acts = {}
        if act_headers and "assembly_id" in act_headers:
            aidx = act_headers.index("assembly_id")
            for r in act_rows[1:]:
                if r and len(r) > aidx:
                    acts[r[aidx]] = dict(zip(act_headers, r + [""] * max(0, len(act_headers) - len(r))))
    except Exception:
        acts = {}
    out = []
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * max(0, len(headers) - len(r))))
        if row.get("archived_at") or row.get("status") in ("cancelled",):
            continue
        visible = is_mgr or str(row.get("expeditor_tg_id", "")) == str(tg_id)
        if not visible:
            continue
        act = acts.get(row.get("id", ""), {})
        out.append({
            "id": row.get("id",""), "client_name": row.get("client_name",""),
            "client_phone": row.get("client_phone",""), "address": row.get("address",""),
            "scheduled_at": row.get("scheduled_at",""), "status": row.get("status",""),
            "is_signed": bool(act.get("signed_by_name")),
            "signed_at": act.get("signed_at",""), "act_num": act.get("act_num",""),
        })
    return {"ok": True, "assemblies": out}


def _handle_act4_request_otp(body):
    import random, datetime as dt
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "expeditor") or sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}
    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}
    code = str(random.randint(100000, 999999))
    expires = (dt.datetime.utcnow() + dt.timedelta(minutes=10)).isoformat()
    _ensure_act4_sheet()
    existing = sheets.find_row("Act4s", "assembly_id", assembly_id)
    if existing:
        sheets.update_cell_by_key("Act4s", "assembly_id", assembly_id, "otp_code", code)
        sheets.update_cell_by_key("Act4s", "assembly_id", assembly_id, "otp_expires_at", expires)
    else:
        act4_id = str(uuid.uuid4())[:8]
        sheets.append_row("Act4s", [act4_id, assembly_id, assembly_id+"-4", _now_iso()[:10],
            "", "[]", "", "0", "0", "", "", "", "", "", code, expires, _now_iso(), str(tg_id), _now_iso()])
    try:
        asm = sheets.find_row("Assemblies", "id", assembly_id)
        client = asm.get("client_name", "") if asm else ""
        tg.send_message(int(tg_id),
            "<b>Код подписи акта</b>\n\nКлиент: " + client + "\nКод: <code>" + code + "</code>\n\nДействителен 10 минут.")
    except Exception as e:
        log.warning("otp send: %s", e)
        return {"error": "send_failed"}
    return {"ok": True, "sent": True}


def _handle_act4_verify_otp(body):
    import datetime as dt
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    assembly_id = (body.get("assembly_id") or "").strip()
    code_input = str(body.get("code") or "").strip()
    signer_name = (body.get("signed_by_name") or "").strip()
    if not assembly_id or not code_input:
        return {"error": "missing_fields"}
    _ensure_act4_sheet()
    act4 = sheets.find_row("Act4s", "assembly_id", assembly_id)
    if not act4:
        return {"error": "act_not_found"}
    stored = act4.get("otp_code", "")
    expires_str = act4.get("otp_expires_at", "")
    if not stored or stored != code_input:
        return {"error": "invalid_code"}
    if expires_str:
        try:
            exp = dt.datetime.fromisoformat(expires_str)
            if dt.datetime.utcnow() > exp:
                return {"error": "code_expired"}
        except Exception:
            pass
    now_iso = _now_iso()
    name = signer_name or (user.get("name") or user.get("first_name") or str(tg_id))
    for col, val in [("signed_by_name",name),("signed_via","telegram_otp"),
                     ("signed_at",now_iso),("otp_code",""),("otp_expires_at","")]:
        sheets.update_cell_by_key("Act4s", "assembly_id", assembly_id, col, val)
    return {"ok": True, "signed": True, "signed_by_name": name}


def _handle_act4_save_signature(body):
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "expeditor") or sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}
    assembly_id = (body.get("assembly_id") or "").strip()
    signature_b64 = (body.get("signature_b64") or "").strip()
    signer_name = (body.get("signed_by_name") or "").strip()
    if not assembly_id or not signature_b64:
        return {"error": "missing_fields"}
    name = signer_name or (user.get("name") or user.get("first_name") or str(tg_id))
    now_iso = _now_iso()
    _ensure_act4_sheet()
    existing = sheets.find_row("Act4s", "assembly_id", assembly_id)
    if existing:
        for col, val in [("signature_b64",signature_b64),("signed_by_name",name),
                         ("signed_via","canvas"),("signed_at",now_iso)]:
            sheets.update_cell_by_key("Act4s", "assembly_id", assembly_id, col, val)
    else:
        act4_id = str(uuid.uuid4())[:8]
        sheets.append_row("Act4s", [act4_id, assembly_id, assembly_id+"-4", now_iso[:10],
            "", "[]", "", "0", "0", name, "", "canvas", now_iso, signature_b64,
            "", "", now_iso, str(tg_id), now_iso])
    return {"ok": True, "signed": True, "signed_by_name": name}


def _handle_dispatcher_inbox(body):
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    is_disp = sheets.has_role(user, "dispatcher")
    is_mgr  = sheets.has_role(user, "manager")
    if not (is_disp or is_mgr):
        return {"error": "forbidden"}

    _ensure_assemblies_sheet()
    try:
        ws = sheets.sheet("Assemblies")
        rows = ws.get_all_values()
    except Exception:
        return {"ok": True, "assemblies": []}
    if not rows or len(rows) < 2:
        return {"ok": True, "assemblies": []}

    headers = rows[0]
    STATUS_ORDER = {"created": 0, "shipped": 1, "arrived": 2, "scheduled": 3,
                    "in_progress": 4, "completed": 5, "cancelled": 9}
    out = []
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * max(0, len(headers) - len(r))))
        if row.get("archived_at"):
            continue
        st = row.get("status", "created")
        if st == "cancelled":
            continue
        out.append({
            "id":                    row.get("id", ""),
            "client_name":           row.get("client_name", ""),
            "client_phone":          row.get("client_phone", ""),
            "address":               row.get("address", ""),
            "scope_of_work":         row.get("scope_of_work", ""),
            "status":                st,
            "scheduled_at":          row.get("scheduled_at", ""),
            "shipment_date":         row.get("shipment_date", ""),
            "packages_count":        row.get("packages_count", ""),
            "arrival_date":          row.get("arrival_date", ""),
            "arrival_packages_count": row.get("arrival_packages_count", ""),
            "expeditor_tg_id":       row.get("expeditor_tg_id", ""),
            "manager_note":          row.get("manager_note", ""),
            "ts":                    row.get("ts", ""),
        })
    out.sort(key=lambda x: (STATUS_ORDER.get(x["status"], 9), x.get("ts", "")))
    return {"ok": True, "assemblies": out}


def _handle_assembly_set_shipment(body):
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "dispatcher") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id    = (body.get("assembly_id") or "").strip()
    shipment_date  = (body.get("shipment_date") or "").strip()
    packages_count = str(body.get("packages_count") or "").strip()
    if not assembly_id or not shipment_date:
        return {"error": "missing_fields"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "shipment_date",  shipment_date)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "packages_count", packages_count)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status",         "shipped")

    # Уведомить менеджера
    try:
        mgr_id = asm.get("manager_tg_id")
        if mgr_id:
            tg.send_message(int(mgr_id),
                f"🚚 <b>Отгрузка с фабрики зафиксирована</b>\n"
                f"Клиент: {asm.get('client_name','')}\n"
                f"Дата отгрузки: {shipment_date}\n"
                f"Упаковок: {packages_count or '—'}")
    except Exception as e:
        log.warning("dispatcher notify error: %s", e)

    return {"ok": True, "status": "shipped"}


def _handle_assembly_set_arrival(body):
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "dispatcher") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id              = (body.get("assembly_id") or "").strip()
    arrival_date             = (body.get("arrival_date") or "").strip()
    arrival_packages_count   = str(body.get("arrival_packages_count") or "").strip()
    if not assembly_id or not arrival_date:
        return {"error": "missing_fields"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "arrival_date",            arrival_date)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "arrival_packages_count",  arrival_packages_count)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "arrival_confirmed_by_tg_id", str(tg_id))
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status",                  "arrived")

    # Проверяем расхождение упаковок
    try:
        expected = int(asm.get("packages_count") or 0)
        actual   = int(arrival_packages_count or 0)
        if expected > 0 and actual != expected:
            mgr_id = asm.get("manager_tg_id")
            if mgr_id:
                tg.send_message(int(mgr_id),
                    f"⚠️ <b>Расхождение упаковок на складе</b>\n"
                    f"Клиент: {asm.get('client_name','')}\n"
                    f"Ожидалось: {expected} · Принято: {actual}")
    except Exception as e:
        log.warning("arrival check error: %s", e)

    return {"ok": True, "status": "arrived"}


def _handle_assembly_assign_dispatch(body):
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "dispatcher") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id    = (body.get("assembly_id") or "").strip()
    scheduled_at   = (body.get("scheduled_at") or "").strip()
    expeditor_tg_id = str(body.get("expeditor_tg_id") or "").strip()
    if not assembly_id or not scheduled_at:
        return {"error": "missing_fields"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "scheduled_at",    scheduled_at)
    if expeditor_tg_id:
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "expeditor_tg_id", expeditor_tg_id)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status", "scheduled")

    # Уведомить экспедитора
    try:
        exp_id = int(expeditor_tg_id) if expeditor_tg_id else None
        if exp_id:
            tg.send_message(exp_id,
                f"📋 <b>Назначена сборка</b>\n"
                f"Клиент: {asm.get('client_name','')}\n"
                f"Адрес: {asm.get('address','')}\n"
                f"Дата: {scheduled_at[:10]}")
        mgr_id = asm.get("manager_tg_id")
        if mgr_id:
            tg.send_message(int(mgr_id),
                f"✅ <b>Сборка назначена</b>\n"
                f"Клиент: {asm.get('client_name','')}\n"
                f"Дата: {scheduled_at[:10]}\n"
                f"Экспедитор: tg:{expeditor_tg_id or '—'}")
    except Exception as e:
        log.warning("dispatch notify error: %s", e)

    return {"ok": True, "status": "scheduled"}


@app.post("/api/dispatcher_inbox")
async def api_dispatcher_inbox(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_dispatcher_inbox(body))


@app.post("/api/assembly_set_shipment")
async def api_assembly_set_shipment(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_assembly_set_shipment(body))


@app.post("/api/assembly_set_arrival")
async def api_assembly_set_arrival(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_assembly_set_arrival(body))


@app.post("/api/assembly_assign_dispatch")
async def api_assembly_assign_dispatch(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_assembly_assign_dispatch(body))


@app.post("/api/expeditor_inbox")
async def api_expeditor_inbox(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_expeditor_inbox(body))

@app.post("/api/act4_request_otp")
async def api_act4_request_otp(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_act4_request_otp(body))

@app.post("/api/act4_verify_otp")
async def api_act4_verify_otp(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_act4_verify_otp(body))

@app.post("/api/act4_save_signature")
async def api_act4_save_signature(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_act4_save_signature(body))


@app.post("/api/act4_preview")
async def api_act4_preview(request: Request):
    body = await _safe_json(request)
    return _handle_act4_preview(body)


@app.post("/api/act4_save")
async def api_act4_save(request: Request):
    body = await _safe_json(request)
    return _handle_act4_save(body)


# =================================================================
# Смена статуса сборки — сборщик меняет статус прямо из карточки
# =================================================================

_ASSEMBLY_STATUS_TRANSITIONS = {
    # текущий → допустимые следующие
    "created":     ["in_progress", "cancelled"],
    "scheduled":   ["in_progress", "cancelled"],
    "in_progress": ["done", "cancelled"],
}


def _handle_assembly_set_status(body: dict[str, Any]) -> dict[str, Any]:
    """Сборщик / менеджер меняет статус сборки.
    body: {initData, assembly_id, status}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    is_assembler = sheets.has_role(user, "assembler")
    is_manager   = sheets.has_role(user, "manager")
    if not (is_assembler or is_manager):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    new_status  = (body.get("status") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    # Assembler — только назначенный; менеджер — любая своя
    if is_assembler and not is_manager:
        if str(row.get("assigned_to_tg_id", "")) != str(tg_id):
            return {"error": "not_assigned"}

    current = (row.get("status") or "created").strip()
    allowed = _ASSEMBLY_STATUS_TRANSITIONS.get(current, [])
    if new_status not in allowed:
        return {"error": "invalid_transition",
                "msg": f"Из «{current}» нельзя перейти в «{new_status}»",
                "allowed": allowed}

    now_iso = _now_iso()
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status", new_status)

    # Временны́е метки
    if new_status == "in_progress":
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "started_at", now_iso)
    elif new_status == "done":
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "completed_at", now_iso)

    # Уведомить менеджера если меняет сборщик
    if is_assembler and not is_manager:
        try:
            mgr_id = row.get("manager_tg_id")
            if mgr_id:
                labels = {"in_progress": "🔨 Сборка началась", "done": "✅ Сборка завершена", "cancelled": "❌ Сборка отменена"}
                tg.send_message(int(mgr_id),
                    f"{labels.get(new_status, new_status)}\n"
                    f"Сборка: <code>{assembly_id}</code>\n"
                    f"Клиент: {row.get('client_name','')}")
        except Exception as e:
            log.warning("assembly_set_status notify: %s", e)

    # Уведомить клиента
    client_tg_id_str = (row.get("client_tg_id") or "").strip()
    if client_tg_id_str:
        try:
            client_msgs = {
                "in_progress": (
                    f"🔨 <b>Сборка вашей кухни началась!</b>\n"
                    f"Адрес: {row.get('address','')}\n"
                    f"Мастер уже на объекте."
                ),
                "done": (
                    f"✅ <b>Сборка завершена!</b>\n"
                    f"Адрес: {row.get('address','')}\n"
                    f"Пожалуйста, проверьте работу и подпишите акт."
                ),
                "cancelled": (
                    f"❌ <b>Сборка отменена.</b>\n"
                    f"Свяжитесь с менеджером для уточнения деталей."
                ),
            }
            if new_status in client_msgs:
                tg.send_message(int(client_tg_id_str), client_msgs[new_status])
        except Exception as e:
            log.warning("assembly_set_status notify client: %s", e)

    sheets.log_event("assembly_status_changed", tg_id, {
        "id": assembly_id, "from": current, "to": new_status,
    })
    return {"ok": True, "status": new_status, "prev_status": current}


@app.post("/api/assembly_set_status")
async def api_assembly_set_status(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_set_status(body)


# =================================================================
# Назначить экспедитора на сборку
# =================================================================

def _handle_assembly_set_expeditor(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер назначает экспедитора на сборку.
    body: {initData, assembly_id, expeditor_tg_id}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    assembly_id    = (body.get("assembly_id") or "").strip()
    exp_tg_id      = str(body.get("expeditor_tg_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "expeditor_tg_id", exp_tg_id)

    # Уведомить экспедитора
    if exp_tg_id:
        try:
            exp_user = sheets.find_user(int(exp_tg_id))
            exp_name = exp_user.get("full_name", "") if exp_user else ""
            tg.send_message(int(exp_tg_id),
                f"📦 <b>Вам назначена приёмка товара</b>\n"
                f"Сборка: <code>{assembly_id}</code>\n"
                f"Клиент: {row.get('client_name','')}\n"
                f"Адрес: {row.get('address','')}\n\n"
                f"Оформите Акт №4 при доставке.")
        except Exception as e:
            log.warning("expeditor notify: %s", e)
        return {"ok": True, "expeditor_tg_id": exp_tg_id}
    return {"ok": True, "expeditor_tg_id": ""}


@app.post("/api/assembly_set_expeditor")
async def api_assembly_set_expeditor(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_set_expeditor(body)


_DATA_URL_RE = re.compile(r"^data:image/(jpeg|jpg|png|webp);base64,(.+)$", re.DOTALL)


# =================================================================
# Фото-отчёт сборки (сборщик / менеджер)
# =================================================================

def _handle_assembly_photo_upload(body: dict[str, Any]) -> dict[str, Any]:
    """Сохраняет фото сборки.
    body: {initData, assembly_id, photo_b64, kind: 'before'|'in_progress'|'after'}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    is_assembler = sheets.has_role(user, "assembler")
    is_manager   = sheets.has_role(user, "manager")
    if not (is_assembler or is_manager):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    kind        = (body.get("kind") or "after").strip()
    if kind not in ("before", "in_progress", "after"):
        kind = "after"
    if not assembly_id or not _SAFE_ID_RE.match(assembly_id):
        return {"error": "missing_assembly_id"}

    photo_b64 = (body.get("photo_b64") or "").strip()
    if not photo_b64:
        return {"error": "missing_photo"}

    m = _DATA_URL_RE.match(photo_b64)
    if not m:
        return {"error": "invalid_photo_format", "msg": "Ожидается data:image/...;base64,..."}
    ext = "jpg" if m.group(1) in ("jpeg", "jpg") else m.group(1)
    try:
        raw = base64.b64decode(m.group(2), validate=False)
    except Exception:
        return {"error": "invalid_photo_base64"}
    if len(raw) > 10 * 1024 * 1024:
        return {"error": "photo_too_large", "msg": "Максимум 10 МБ"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    if is_assembler and not is_manager:
        if str(row.get("assigned_to_tg_id", "")) != str(tg_id):
            return {"error": "not_assigned"}

    col_name = f"photos_{kind}"
    target_dir = PHOTOS_DIR / assembly_id
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        existing = [x for x in (row.get(col_name) or "").split(",") if x.strip()]
        n = len(existing) + 1
        filename = f"{kind}_{n}.{ext}"
        (target_dir / filename).write_bytes(raw)
    except Exception as e:
        log.warning("assembly photo save failed: %s", e)
        return {"error": "save_failed"}

    existing_str = (row.get(col_name) or "").strip().strip(",")
    new_val = (existing_str + "," + filename).lstrip(",")
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, col_name, new_val)

    log.info("Assembly photo saved: %s/%s", assembly_id, filename)
    return {"ok": True, "filename": filename, "kind": kind}


# =================================================================
# Испытательный срок сборщика (менеджер включает/выключает)
# =================================================================

def _ensure_users_probation_col() -> None:
    """Добавляет колонку on_probation в Users если её нет."""
    try:
        ws = sheets.sheet("Users")
        headers = ws.row_values(1)
        if "on_probation" not in headers:
            ws.update_cell(1, len(headers) + 1, "on_probation")
            log.info("Users: добавили колонку on_probation")
    except Exception as e:
        log.warning("_ensure_users_probation_col: %s", e)


def _handle_assembler_set_probation(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер устанавливает / снимает испытательный срок у сборщика.
    body: {initData, assembler_tg_id, on_probation: true|false}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    assembler_tg_id = str(body.get("assembler_tg_id") or "").strip()
    if not assembler_tg_id:
        return {"error": "missing_assembler_tg_id"}
    on_prob = bool(body.get("on_probation"))

    target = sheets.find_user(int(assembler_tg_id))
    if not target:
        return {"error": "user_not_found"}
    if not sheets.has_role(target, "assembler"):
        return {"error": "not_assembler"}

    _ensure_users_probation_col()
    ok = sheets.update_cell_by_key("Users", "tg_id", assembler_tg_id, "on_probation", "1" if on_prob else "")
    if not ok:
        return {"error": "update_failed"}

    # Уведомить сборщика
    try:
        if on_prob:
            tg.send_message(int(assembler_tg_id),
                "📋 <b>Вы переведены на испытательный срок.</b>\n"
                "Для каждого заказа требуется фото-отчёт «До / После сборки».")
        else:
            tg.send_message(int(assembler_tg_id),
                "✅ <b>Испытательный срок завершён.</b> Поздравляем!")
    except Exception as e:
        log.warning("probation notify: %s", e)

    return {"ok": True, "assembler_tg_id": assembler_tg_id, "on_probation": on_prob}


# =================================================================
# Заметки сборщика
# =================================================================

def _handle_assembly_notes_save(body: dict[str, Any]) -> dict[str, Any]:
    """Сборщик (назначенный) сохраняет заметки по ходу сборки.
    body: {initData, assembly_id, notes}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    notes       = (body.get("notes") or "").strip()[:2000]
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    if sheets.has_role(user, "assembler") and not sheets.has_role(user, "manager"):
        if str(row.get("assigned_to_tg_id", "")) != str(tg_id):
            return {"error": "not_assigned"}

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "assembler_notes", notes)
    return {"ok": True}


# =================================================================
# Счёт клиенту на сборку
# =================================================================

def _handle_assembly_invoice_create(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер / сборщик создаёт счёт клиенту на оплату сборки.
    body: {initData, assembly_id, amount?}
    Если amount не передан — берём assembly_price_for_client."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    # Вычисляем цену сборки для клиента через общий хелпер
    prices = _calc_assembly_prices(row, tg_id)
    auto_amount = prices.get("assembly_price_for_client") or 0

    amount_raw = body.get("amount")
    if amount_raw is not None:
        try:
            amount = float(amount_raw)
            if amount <= 0:
                raise ValueError
        except (TypeError, ValueError):
            return {"error": "invalid_amount"}
    else:
        amount = float(auto_amount or 0)
    if amount <= 0:
        return {"error": "amount_required", "msg": "Укажите сумму или задайте стоимость кухни"}

    address = row.get("address", "")
    purpose = f"Оплата услуг по сборке кухни {address or assembly_id}"
    try:
        qr_b64 = _invoice_qr_b64(amount, purpose)
    except Exception as e:
        log.warning("assembly invoice qr error: %s", e)
        qr_b64 = ""

    now_date = _now_iso()[:10]
    try:
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "assembly_invoice_amount", str(amount))
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "assembly_invoice_date", now_date)
    except Exception as e:
        log.warning("assembly_invoice_create: save error: %s", e)

    return {
        "ok": True,
        "assembly_id":  assembly_id,
        "client_name":  row.get("client_name", "Клиент"),
        "client_phone": row.get("client_phone", ""),
        "address":      address,
        "date":         now_date,
        "amount":       amount,
        "purpose":      purpose,
        "ip_name":      _IP_NAME,
        "ip_inn":       _IP_INN,
        "bank_name":    _IP_BANK,
        "bic":          _IP_BIC,
        "rs":           _IP_RS,
        "ks":           _IP_KS,
        "qr_b64":       qr_b64,
    }


# =================================================================
# Доп работы (AssemblyExtras) — чеки из магазина
# =================================================================

def _assembly_extras_columns() -> list[str]:
    return ["id", "ts", "assembly_id", "added_by_tg_id", "added_by_name",
            "description", "amount", "receipt_photo",
            "status",               # pending | approved | rejected
            "approved_by_tg_id", "approved_at"]


def _ensure_extras_sheet() -> None:
    want = _assembly_extras_columns()
    try:
        ws = sheets.sheet("AssemblyExtras")
        existing = ws.row_values(1)
        if not existing:
            ws.update("A1", [want])
            return
        missing = [c for c in want if c not in existing]
        if missing:
            ws.update("A1", [existing + missing])
    except Exception:
        sheets.ensure_sheet("AssemblyExtras", want)


def _assembly_extra_auth(body: dict) -> tuple[dict | None, int | None]:
    """Возвращает (user, tg_id) или (None, None) при ошибке."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return None, None
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    return user, tg_id


def _handle_assembly_extras_list(body: dict[str, Any]) -> dict[str, Any]:
    """Список доп работ по сборке."""
    user, tg_id = _assembly_extra_auth(body)
    if not user:
        return {"error": "invalid_init_data"}
    if not (sheets.has_role(user, "assembler") or sheets.has_role(user, "manager") or sheets.has_role(user, "client")):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_extras_sheet()
    try:
        ws = sheets.sheet("AssemblyExtras")
        rows = ws.get_all_values()
    except Exception:
        return {"ok": True, "extras": []}
    if not rows or len(rows) < 2:
        return {"ok": True, "extras": []}
    headers = rows[0]
    out = []
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if row.get("assembly_id") != assembly_id:
            continue
        out.append({
            "id":            row.get("id", ""),
            "ts":            row.get("ts", ""),
            "description":   row.get("description", ""),
            "amount":        row.get("amount", ""),
            "receipt_photo": row.get("receipt_photo", ""),
            "added_by_name": row.get("added_by_name", ""),
            "status":        row.get("status", "pending") or "pending",
            "approved_at":   row.get("approved_at", ""),
        })
    return {"ok": True, "extras": out}


def _handle_assembly_extra_add(body: dict[str, Any]) -> dict[str, Any]:
    """Добавляет доп работу. receipt_b64 — фото чека (опционально)."""
    user, tg_id = _assembly_extra_auth(body)
    if not user:
        return {"error": "invalid_init_data"}
    if not (sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    description = (body.get("description") or "").strip()[:300]
    amount_raw  = body.get("amount")
    receipt_b64 = (body.get("receipt_b64") or "").strip()

    if not assembly_id or not description:
        return {"error": "missing_fields"}
    try:
        amount = float(amount_raw) if amount_raw else 0.0
    except (TypeError, ValueError):
        amount = 0.0

    _ensure_assemblies_sheet()
    if not sheets.find_row("Assemblies", "id", assembly_id):
        return {"error": "assembly_not_found"}

    # Сохраняем фото чека если есть
    receipt_fn = ""
    if receipt_b64:
        m = _DATA_URL_RE.match(receipt_b64)
        if m and _SAFE_ID_RE.match(assembly_id):
            ext = "jpg" if m.group(1) in ("jpeg", "jpg") else m.group(1)
            try:
                raw = base64.b64decode(m.group(2), validate=False)
                if len(raw) <= 10 * 1024 * 1024:
                    target_dir = PHOTOS_DIR / assembly_id
                    target_dir.mkdir(parents=True, exist_ok=True)
                    existing_cnt = len([f for f in target_dir.iterdir() if f.name.startswith("receipt_")])
                    receipt_fn = f"receipt_{existing_cnt + 1}.{ext}"
                    (target_dir / receipt_fn).write_bytes(raw)
            except Exception as e:
                log.warning("extra receipt save: %s", e)

    extra_id = _short_id()
    full_name = user.get("full_name") or f"{user.get('first_name','')} {user.get('last_name','')}".strip()

    _ensure_extras_sheet()
    sheets.append_named_row("AssemblyExtras", {
        "id": extra_id, "ts": _now_iso(), "assembly_id": assembly_id,
        "added_by_tg_id": str(tg_id), "added_by_name": full_name,
        "description": description, "amount": str(amount) if amount else "",
        "receipt_photo": receipt_fn,
        "status": "pending",
        "approved_by_tg_id": "", "approved_at": "",
    })

    # Уведомить менеджера о новой доп работе
    try:
        asm_row = sheets.find_row("Assemblies", "id", assembly_id)
        if asm_row:
            mgr_id = asm_row.get("manager_tg_id")
            if mgr_id and str(mgr_id) != str(tg_id):
                amt_str = f"{amount:,.0f} ₽".replace(",", " ") if amount else "сумма не указана"
                tg.send_message(int(mgr_id),
                    f"🧾 <b>Доп работа на согласование</b>\n"
                    f"Клиент: {asm_row.get('client_name','')}\n"
                    f"{description} — {amt_str}")
    except Exception as e:
        log.warning("extra_add notify manager: %s", e)

    return {"ok": True, "extra": {
        "id": extra_id, "ts": _now_iso(), "description": description,
        "amount": str(amount) if amount else "",
        "receipt_photo": receipt_fn, "added_by_name": full_name,
    }}


def _handle_assembly_extra_delete(body: dict[str, Any]) -> dict[str, Any]:
    """Удаляет запись доп работы (менеджер или автор)."""
    user, tg_id = _assembly_extra_auth(body)
    if not user:
        return {"error": "invalid_init_data"}
    if not (sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    extra_id    = (body.get("extra_id") or "").strip()
    if not extra_id:
        return {"error": "missing_extra_id"}

    _ensure_extras_sheet()
    try:
        ws = sheets.sheet("AssemblyExtras")
        rows = ws.get_all_values()
        if not rows:
            return {"error": "not_found"}
        headers = rows[0]
        id_idx    = headers.index("id")       if "id"       in headers else -1
        auth_idx  = headers.index("added_by_tg_id") if "added_by_tg_id" in headers else -1
        for i, r in enumerate(rows[1:], start=2):
            if len(r) > id_idx and r[id_idx] == extra_id:
                is_author  = auth_idx >= 0 and len(r) > auth_idx and str(r[auth_idx]) == str(tg_id)
                is_manager = sheets.has_role(user, "manager")
                if not (is_author or is_manager):
                    return {"error": "forbidden"}
                ws.delete_rows(i)
                return {"ok": True}
    except Exception as e:
        log.warning("extra_delete: %s", e)
    return {"error": "not_found"}


def _handle_assembly_receipt_parse(body: dict[str, Any]) -> dict[str, Any]:
    """Парсит сумму из фото чека через GigaChat Vision."""
    user, tg_id = _assembly_extra_auth(body)
    if not user:
        return {"error": "invalid_init_data"}
    if not (sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    photo_b64 = (body.get("photo_b64") or "").strip()
    if not photo_b64:
        return {"error": "missing_photo"}

    result = ai.parse_receipt_amount(photo_b64)
    return {
        "ok":     not result.get("error"),
        "amount": result.get("amount"),
        "raw":    result.get("raw", ""),
    }


def _handle_assembly_extra_approve(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер согласует или отклоняет доп работу.
    body: {initData, assembly_id, extra_id, action: 'approve'|'reject'}"""
    user, tg_id = _assembly_extra_auth(body)
    if not user:
        return {"error": "invalid_init_data"}
    if not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    assembly_id = (body.get("assembly_id") or "").strip()
    extra_id    = (body.get("extra_id") or "").strip()
    action      = (body.get("action") or "").strip()
    if action not in ("approve", "reject"):
        return {"error": "bad_action"}
    if not extra_id:
        return {"error": "missing_extra_id"}

    _ensure_extras_sheet()
    try:
        ws = sheets.sheet("AssemblyExtras")
        rows = ws.get_all_values()
        if not rows:
            return {"error": "not_found"}
        headers = rows[0]
        id_idx     = headers.index("id")     if "id"     in headers else -1
        status_idx = headers.index("status") if "status" in headers else -1
        appr_idx   = headers.index("approved_by_tg_id") if "approved_by_tg_id" in headers else -1
        at_idx     = headers.index("approved_at")       if "approved_at"       in headers else -1
        author_idx = headers.index("added_by_tg_id")    if "added_by_tg_id"    in headers else -1
        desc_idx   = headers.index("description")       if "description"       in headers else -1

        new_status = "approved" if action == "approve" else "rejected"
        now_iso = _now_iso()

        for i, r in enumerate(rows[1:], start=2):
            if len(r) > id_idx and r[id_idx] == extra_id:
                if status_idx >= 0:
                    ws.update_cell(i, status_idx + 1, new_status)
                if appr_idx >= 0:
                    ws.update_cell(i, appr_idx + 1, str(tg_id))
                if at_idx >= 0:
                    ws.update_cell(i, at_idx + 1, now_iso)

                # Уведомить сборщика
                try:
                    author_tg_id = r[author_idx] if author_idx >= 0 and len(r) > author_idx else ""
                    desc = r[desc_idx] if desc_idx >= 0 and len(r) > desc_idx else ""
                    if author_tg_id and str(author_tg_id) != str(tg_id):
                        emoji = "✅" if action == "approve" else "❌"
                        label = "согласована" if action == "approve" else "отклонена"
                        tg.send_message(int(author_tg_id),
                            f"{emoji} <b>Доп работа {label}</b>\n{desc}")
                except Exception as e:
                    log.warning("extra_approve notify: %s", e)

                return {"ok": True, "status": new_status}
    except Exception as e:
        log.warning("extra_approve: %s", e)
    return {"error": "not_found"}


# =================================================================
# Обзор команды (менеджер)
# =================================================================

def _handle_staff_roster(body: dict[str, Any]) -> dict[str, Any]:
    """Полный список сотрудников с нагрузкой, статусом оборудования и испытательным сроком."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    # Считаем активные сборки по сборщику
    active_by_assembler: dict[str, int] = {}
    try:
        _ensure_assemblies_sheet()
        ws_asm = sheets.sheet("Assemblies")
        asm_rows = ws_asm.get_all_values()
        if asm_rows and len(asm_rows) > 1:
            hdrs = asm_rows[0]
            for r in asm_rows[1:]:
                row = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                if row.get("status") in ("created", "scheduled", "in_progress"):
                    atg = (row.get("assigned_to_tg_id") or "").strip()
                    if atg:
                        active_by_assembler[atg] = active_by_assembler.get(atg, 0) + 1
    except Exception:
        pass

    # Счётчик замеров за текущий месяц по замерщику
    month_prefix = _now_iso()[:7]  # "2026-05"
    measures_by_measurer: dict[str, int] = {}
    try:
        ws_m = sheets.sheet("Measurements")
        m_rows = ws_m.get_all_values()
        if m_rows and len(m_rows) > 1:
            hdrs = m_rows[0]
            for r in m_rows[1:]:
                row = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                if (row.get("ts") or "").startswith(month_prefix):
                    atg = (row.get("assigned_to_tg_id") or "").strip()
                    if atg:
                        measures_by_measurer[atg] = measures_by_measurer.get(atg, 0) + 1
    except Exception:
        pass

    EQUIPMENT_REQUIRED = {"tablet", "laser_tape", "angle_meter", "tape", "laser_level"}
    out: list[dict] = []

    try:
        ws_u = sheets.sheet("Users")
        u_rows = ws_u.get_all_values()
        if not u_rows or len(u_rows) < 2:
            return {"ok": True, "staff": []}
        hdrs = u_rows[0]
        for r in u_rows[1:]:
            row = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
            roles = sheets.parse_roles(row.get("role", ""))
            if not any(rl in roles for rl in ("assembler", "measurer", "expeditor")):
                continue
            tg_id_str = (row.get("tg_id") or "").strip()
            full_name = (f"{row.get('first_name','')} {row.get('last_name','')}".strip()
                         or row.get("tg_username", "") or tg_id_str)

            eq_raw = row.get("equipment", "")
            eq_list = [x.strip() for x in eq_raw.split(",") if x.strip()] if eq_raw else []
            equipment_ok = EQUIPMENT_REQUIRED.issubset(set(eq_list)) if "measurer" in roles else None
            on_probation = str(row.get("on_probation", "")).lower() in ("1", "true", "yes")

            out.append({
                "tg_id":        tg_id_str,
                "full_name":    full_name,
                "tg_username":  row.get("tg_username", ""),
                "roles":        roles,
                "equipment_ok": equipment_ok,
                "on_probation": on_probation,
                "avg_stars":    _get_avg_stars(tg_id_str),
                "active_assemblies": active_by_assembler.get(tg_id_str, 0),
                "month_measures":    measures_by_measurer.get(tg_id_str, 0),
            })
    except Exception as e:
        log.warning("staff_roster: %s", e)
        return {"error": "sheets_error"}

    # Сортировка: сначала сборщики, потом замерщики
    out.sort(key=lambda x: (0 if "assembler" in x["roles"] else 1, x["full_name"]))
    return {"ok": True, "staff": out}


# =================================================================
# Таймлайн заказа для клиента
# =================================================================

def _handle_client_order_timeline(body: dict[str, Any]) -> dict[str, Any]:
    """Визуальный таймлайн заказа: замер → сборка → акт.
    Доступен клиенту, менеджеру и назначенному сборщику."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    is_authorized = (
        str(asm.get("client_tg_id", "")) == str(tg_id) or
        str(asm.get("manager_tg_id", "")) == str(tg_id) or
        str(asm.get("assigned_to_tg_id", "")) == str(tg_id)
    )
    if not is_authorized:
        return {"error": "forbidden"}

    milestones: list[dict] = []

    # --- Замер ---
    measurement = None
    measurement_id = (asm.get("measurement_id") or "").strip()
    if measurement_id:
        try:
            measurement = sheets.find_row("Measurements", "id", measurement_id)
        except Exception:
            pass

    if measurement:
        milestones.append({
            "key": "request_created",
            "icon": "📋",
            "title": "Заявка создана",
            "ts": measurement.get("ts") or measurement.get("created_at", ""),
            "done": True,
            "detail": None,
        })

        measurer_name = ""
        if measurement.get("assigned_to_tg_id"):
            try:
                m_user = sheets.find_user(int(measurement["assigned_to_tg_id"]))
                if m_user:
                    measurer_name = (m_user.get("full_name") or
                        f"{m_user.get('first_name','')} {m_user.get('last_name','')}".strip())
            except Exception:
                pass

        milestones.append({
            "key": "measure_scheduled",
            "icon": "📐",
            "title": "Замер назначен",
            "ts": measurement.get("scheduled_at") or None,
            "done": bool(measurement.get("scheduled_at")),
            "detail": measurer_name or None,
        })

        meas_done = measurement.get("status") == "completed"
        milestones.append({
            "key": "measure_done",
            "icon": "✅",
            "title": "Замер выполнен",
            "ts": measurement.get("completed_at") or (measurement.get("scheduled_at") if meas_done else None),
            "done": meas_done,
            "detail": None,
        })
    else:
        milestones.append({
            "key": "request_created",
            "icon": "📋",
            "title": "Заявка создана",
            "ts": asm.get("ts", ""),
            "done": True,
            "detail": None,
        })

    # --- Сборка создана ---
    milestones.append({
        "key": "assembly_created",
        "icon": "🔨",
        "title": "Сборка создана",
        "ts": asm.get("ts", ""),
        "done": True,
        "detail": asm.get("address") or None,
    })

    # --- Товар принят (Акт №4) ---
    act4_signed = False
    act4_signed_at = ""
    act4_signed_by = ""
    try:
        _ensure_act4_sheet()
        act4_row = sheets.find_row("Act4s", "assembly_id", assembly_id)
        if act4_row and act4_row.get("signed_by_name"):
            act4_signed = True
            act4_signed_at = act4_row.get("signed_at", "")
            act4_signed_by = act4_row.get("signed_by_name", "")
    except Exception:
        pass

    milestones.append({
        "key": "goods_accepted",
        "icon": "📦",
        "title": "Товар принят",
        "ts": act4_signed_at or None,
        "done": act4_signed,
        "detail": f"Принял: {act4_signed_by}" if act4_signed_by else None,
    })

    # --- Сборка началась ---
    asm_status = asm.get("status", "")
    in_progress_done = asm_status in ("in_progress", "done")
    milestones.append({
        "key": "assembly_started",
        "icon": "🔧",
        "title": "Сборка началась",
        "ts": asm.get("started_at") or None,
        "done": in_progress_done,
        "detail": None,
    })

    # --- Доп работы ---
    extras_count = 0
    extras_approved = 0.0
    try:
        _ensure_assembly_extras_sheet()
        ws_ex = sheets.sheet("AssemblyExtras")
        ex_rows = ws_ex.get_all_values()
        if ex_rows and len(ex_rows) > 1:
            hdrs = ex_rows[0]
            for r in ex_rows[1:]:
                rd = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                if rd.get("assembly_id") == assembly_id:
                    extras_count += 1
                    if rd.get("status") == "approved":
                        try:
                            extras_approved += float(rd.get("amount") or 0)
                        except (ValueError, TypeError):
                            pass
    except Exception:
        pass

    if extras_count > 0:
        detail_str = f"{extras_count} поз."
        if extras_approved > 0:
            detail_str += f" · одобрено {int(extras_approved):,} ₽".replace(",", " ")
        milestones.append({
            "key": "extras",
            "icon": "🧾",
            "title": "Доп работы",
            "ts": None,
            "done": True,
            "detail": detail_str,
        })

    # --- Сборка завершена ---
    asm_done = asm_status == "done"
    milestones.append({
        "key": "assembly_done",
        "icon": "✅",
        "title": "Сборка завершена",
        "ts": asm.get("completed_at") or None,
        "done": asm_done,
        "detail": None,
    })

    # --- Акт подписан ---
    signed = bool(asm.get("signed_by_name"))
    milestones.append({
        "key": "act_signed",
        "icon": "✍️",
        "title": "Акт сдачи-приёмки подписан",
        "ts": asm.get("signed_at") or None,
        "done": signed,
        "detail": asm.get("signed_by_name") or None,
    })

    return {
        "ok": True,
        "assembly_id": assembly_id,
        "client_name": asm.get("client_name", ""),
        "address": asm.get("address", ""),
        "status": asm_status,
        "milestones": milestones,
    }


# =================================================================
# Финансовая сводка для менеджера
# =================================================================

_MONTHS_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]
_MONTHS_RU_GEN = [
    "", "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
]


def _month_prefixes(period: str, now: datetime) -> tuple[list[str], str]:
    """Возвращает (список префиксов YYYY-MM, человеко-читаемый лейбл)."""
    y, m = now.year, now.month
    if period == "prev_month":
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        return [f"{y:04d}-{m:02d}"], f"{_MONTHS_RU[m]} {y}"
    elif period == "quarter":
        prefixes = []
        labels = []
        for i in range(3):
            cm, cy = m - i, y
            if cm <= 0:
                cm += 12
                cy -= 1
            prefixes.append(f"{cy:04d}-{cm:02d}")
            labels.append(_MONTHS_RU_GEN[cm])
        return prefixes, f"{labels[-1]} – {labels[0]} {y}"
    else:  # current_month
        return [f"{y:04d}-{m:02d}"], f"{_MONTHS_RU[m]} {y}"


def _handle_manager_finance_summary(body: dict[str, Any]) -> dict[str, Any]:
    """Финансовая сводка менеджера: замеры, сборки, выручка, выплаты, доп работы.
    body: {initData, period: 'current_month'|'prev_month'|'quarter'}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    period = (body.get("period") or "current_month").strip()
    now = datetime.now(timezone.utc)
    prefixes, period_label = _month_prefixes(period, now)

    def _in_period(ts: str) -> bool:
        return bool(ts) and any(ts.startswith(p) for p in prefixes)

    # ── Замеры ──────────────────────────────────────────────────────
    meas_total = 0
    meas_done = 0
    try:
        ws_m = sheets.sheet("Measurements")
        m_rows = ws_m.get_all_values()
        if m_rows and len(m_rows) > 1:
            hdrs = m_rows[0]
            for r in m_rows[1:]:
                row = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                if _in_period(row.get("ts", "")):
                    meas_total += 1
                    if row.get("status") == "completed":
                        meas_done += 1
    except Exception as e:
        log.warning("finance_summary measurements: %s", e)

    # ── Сборки ──────────────────────────────────────────────────────
    asm_total = 0
    asm_done_count = 0
    asm_active_count = 0
    revenue_client = 0.0    # выручка (клиент платит)
    payout_assembler = 0.0  # выплата сборщику
    asm_list: list[dict] = []

    try:
        _ensure_assemblies_sheet()
        ws_a = sheets.sheet("Assemblies")
        a_rows = ws_a.get_all_values()
        if a_rows and len(a_rows) > 1:
            hdrs = a_rows[0]
            _ensure_rates_sheet()
            for r in a_rows[1:]:
                row = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                # Фильтр по периоду: created or completed in period
                ts_use = row.get("completed_at") or row.get("ts", "")
                if not _in_period(ts_use):
                    continue
                asm_total += 1
                status = row.get("status", "")
                if status == "done":
                    asm_done_count += 1
                elif status in ("created", "scheduled", "in_progress"):
                    asm_active_count += 1

                # Финансы только для done-сборок с указанной ценой кухни
                kp = 0.0
                try:
                    kp = float(row.get("kitchen_price") or 0)
                except (ValueError, TypeError):
                    pass

                if kp and status == "done":
                    atg = str(row.get("assigned_to_tg_id") or "")
                    cr, ar = _resolve_rates(atg, scope="*")
                    client_pay = round(kp * cr / 100, 2)
                    asm_pay    = round(kp * ar / 100, 2)
                    revenue_client    += client_pay
                    payout_assembler  += asm_pay
                    asm_list.append({
                        "id":           row.get("id", ""),
                        "client_name":  row.get("client_name", ""),
                        "address":      row.get("address", ""),
                        "completed_at": row.get("completed_at", ""),
                        "kitchen_price": kp,
                        "client_pay":   client_pay,
                        "asm_pay":      asm_pay,
                        "margin":       round(client_pay - asm_pay, 2),
                    })
    except Exception as e:
        log.warning("finance_summary assemblies: %s", e)

    # ── Доп работы (approved) ────────────────────────────────────────
    extras_total = 0.0
    extras_count = 0
    try:
        _ensure_assembly_extras_sheet()
        ws_ex = sheets.sheet("AssemblyExtras")
        ex_rows = ws_ex.get_all_values()
        if ex_rows and len(ex_rows) > 1:
            hdrs = ex_rows[0]
            for r in ex_rows[1:]:
                rd = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                if rd.get("status") == "approved" and _in_period(rd.get("ts", "")):
                    extras_count += 1
                    try:
                        extras_total += float(rd.get("amount") or 0)
                    except (ValueError, TypeError):
                        pass
    except Exception as e:
        log.warning("finance_summary extras: %s", e)

    margin = round(revenue_client - payout_assembler, 2)

    return {
        "ok": True,
        "period": period,
        "period_label": period_label,
        # Замеры
        "meas_total":   meas_total,
        "meas_done":    meas_done,
        # Сборки
        "asm_total":    asm_total,
        "asm_done":     asm_done_count,
        "asm_active":   asm_active_count,
        # Финансы
        "revenue_client":   round(revenue_client, 2),
        "payout_assembler": round(payout_assembler, 2),
        "margin":           margin,
        # Доп работы
        "extras_count":  extras_count,
        "extras_total":  round(extras_total, 2),
        # Детали сборок с деньгами
        "asm_list":      sorted(asm_list, key=lambda x: x["completed_at"], reverse=True),
    }


# =================================================================
# Согласование даты сборки с клиентом
# =================================================================

def _auth_manager_only(body: dict) -> tuple[Any, dict | None]:
    """Возвращает (tg_id, None) при успехе или (None, error_dict) при ошибке."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return None, {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return None, {"error": "only_manager"}
    return tg_id, None


def _auth_any_user(body: dict) -> tuple[Any, Any, dict | None]:
    """Возвращает (tg_id, user, None) при успехе или (None, None, error_dict)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return None, None, {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return None, None, {"error": "user_not_found"}
    return tg_id, user, None


def _fmt_dt_ru(iso: str) -> str:
    """ISO → «19 мая, 14:00»"""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        months = ["", "января", "февраля", "марта", "апреля", "мая", "июня",
                  "июля", "августа", "сентября", "октября", "ноября", "декабря"]
        return f"{dt.day} {months[dt.month]}, {dt.hour:02d}:{dt.minute:02d}"
    except Exception:
        return iso[:16].replace("T", " ")


# =================================================================
# Система оценок (Feedback)
# =================================================================

_FEEDBACK_COLUMNS = [
    "id", "ts",
    "from_tg_id", "from_role",
    "target_tg_id", "target_role",  # target_role: assembler|measurer|manager|service
    "ref_id", "ref_type",            # ref_type: assembly|measurement
    "stars",                         # 1..5
    "comment",
]


def _ensure_feedback_sheet() -> None:
    try:
        sheets.sheet("Feedback").row_values(1)
    except Exception:
        sheets.ensure_sheet("Feedback", _FEEDBACK_COLUMNS)


def _get_avg_stars(target_tg_id: str) -> float | None:
    """Средний балл по всем оценкам для target_tg_id. None если оценок нет."""
    try:
        _ensure_feedback_sheet()
        rows = sheets.get_all_rows("Feedback")
        vals = [
            int(r["stars"]) for r in rows
            if r.get("target_tg_id") == str(target_tg_id)
            and str(r.get("stars", "")).isdigit()
            and 1 <= int(r["stars"]) <= 5
        ]
        return round(sum(vals) / len(vals), 1) if vals else None
    except Exception:
        return None


def _handle_feedback_submit(body: dict[str, Any]) -> dict[str, Any]:
    """Сохраняет набор оценок одним вызовом.
    body: {
      initData,
      ref_id,        # assembly_id или measurement_id
      ref_type,      # "assembly" | "measurement"
      ratings: [
        {target_tg_id?, target_role, stars, comment?}
      ]
    }"""
    tg_id, user, err = _auth_any_user(body)
    if err:
        return err

    ref_id   = (body.get("ref_id")   or "").strip()
    ref_type = (body.get("ref_type") or "").strip()
    ratings  = body.get("ratings") or []

    if not ref_id or not ref_type:
        return {"error": "missing_ref"}
    if not ratings:
        return {"error": "missing_ratings"}

    roles = sheets.parse_roles(user.get("role", ""))
    from_role = (
        "client"   if "client"   in roles else
        "measurer" if "measurer" in roles else
        "assembler" if "assembler" in roles else
        "manager"  if "manager"  in roles else
        "user"
    )

    _ensure_feedback_sheet()
    now = _now_iso()

    for r in ratings:
        stars = int(r.get("stars") or 0)
        if not (1 <= stars <= 5):
            continue
        target_role = (r.get("target_role") or "").strip()
        if not target_role:
            continue
        sheets.append_row("Feedback", _FEEDBACK_COLUMNS, {
            "id":           str(uuid.uuid4()),
            "ts":           now,
            "from_tg_id":   str(tg_id),
            "from_role":    from_role,
            "target_tg_id": str(r.get("target_tg_id") or ""),
            "target_role":  target_role,
            "ref_id":       ref_id,
            "ref_type":     ref_type,
            "stars":        str(stars),
            "comment":      str(r.get("comment") or ""),
        })

    # Отмечаем что отзыв оставлен
    if ref_type == "assembly":
        _ensure_assemblies_sheet()
        asm = sheets.find_row("Assemblies", "id", ref_id)
        if asm:
            # Определяем поле по роли отправителя
            if from_role == "client":
                sheets.update_cell_by_key("Assemblies", "id", ref_id, "client_feedback_at", now)
    elif ref_type == "measurement":
        _ensure_measurements_sheet()
        m = sheets.find_row("Measurements", "id", ref_id)
        if m:
            if from_role == "measurer":
                sheets.update_cell_by_key("Measurements", "id", ref_id, "measurer_feedback_at", now)
            elif from_role == "manager":
                sheets.update_cell_by_key("Measurements", "id", ref_id, "manager_feedback_at", now)

    return {"ok": True, "saved": len(ratings)}


def _handle_feedback_my(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает агрегированные оценки для текущего пользователя (или target_tg_id).
    body: {initData, target_tg_id?}"""
    tg_id, user, err = _auth_any_user(body)
    if err:
        return err

    target_id = str(body.get("target_tg_id") or tg_id)
    # Менеджер может смотреть любого; остальные — только себя
    if str(target_id) != str(tg_id):
        if not sheets.has_role(user, "manager"):
            return {"error": "forbidden"}

    try:
        _ensure_feedback_sheet()
        rows = sheets.get_all_rows("Feedback")
    except Exception as e:
        return {"error": "sheets_error", "msg": str(e)}

    my_rows = [r for r in rows if r.get("target_tg_id") == str(target_id)]

    # Группируем по target_role
    by_role: dict[str, list[int]] = {}
    comments: list[dict] = []
    for r in my_rows:
        role = r.get("target_role", "")
        try:
            s = int(r["stars"])
            if 1 <= s <= 5:
                by_role.setdefault(role, []).append(s)
        except (ValueError, TypeError):
            pass
        if r.get("comment"):
            comments.append({
                "ts":      r.get("ts", ""),
                "role":    r.get("from_role", ""),
                "comment": r["comment"],
                "stars":   r.get("stars", ""),
            })

    aggregated = []
    role_labels = {
        "assembler": "Как сборщик",
        "measurer":  "Как замерщик",
        "manager":   "Как менеджер",
        "service":   "Сервис компании",
    }
    for role, vals in by_role.items():
        aggregated.append({
            "target_role": role,
            "label": role_labels.get(role, role),
            "avg":   round(sum(vals) / len(vals), 1),
            "count": len(vals),
        })

    # Последние 5 комментариев
    comments.sort(key=lambda x: x["ts"], reverse=True)

    return {
        "ok": True,
        "target_tg_id": target_id,
        "aggregated": aggregated,
        "comments": comments[:5],
        "total": len(my_rows),
    }


def _handle_assembly_suggest_slots(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает свободные слоты сборщиков на 14 дней вперёд, отсортированных по рейтингу.
    body: {initData, assembly_id}"""
    tg_id, err = _auth_manager_only(body)
    if err:
        return err

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()

    # ── Занятость всех сборщиков (busy dates) ────────────────────────
    # busy_dates[tg_id] = set of date strings "YYYY-MM-DD"
    busy_dates: dict[str, set] = {}
    completed_count: dict[str, int] = {}
    active_count: dict[str, int] = {}

    try:
        ws_a = sheets.sheet("Assemblies")
        rows = ws_a.get_all_values()
        if rows and len(rows) > 1:
            hdrs = rows[0]
            for r in rows[1:]:
                row = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                atg = (row.get("assigned_to_tg_id") or "").strip()
                if not atg:
                    continue
                status = row.get("status", "")
                if status == "done":
                    completed_count[atg] = completed_count.get(atg, 0) + 1
                if status in ("created", "scheduled", "in_progress"):
                    active_count[atg] = active_count.get(atg, 0) + 1
                    # Занятая дата = день назначенной/активной сборки
                    sched = (row.get("scheduled_at") or "").strip()
                    if sched:
                        day = sched[:10]
                        if atg not in busy_dates:
                            busy_dates[atg] = set()
                        busy_dates[atg].add(day)
    except Exception as e:
        log.warning("suggest_slots busy_dates: %s", e)

    # ── Список сборщиков ──────────────────────────────────────────────
    assemblers = []
    EQUIPMENT_REQUIRED = {"tablet", "laser_tape", "angle_meter", "tape", "laser_level"}
    try:
        ws_u = sheets.sheet("Users")
        u_rows = ws_u.get_all_values()
        if u_rows and len(u_rows) > 1:
            hdrs = u_rows[0]
            for r in u_rows[1:]:
                row = dict(zip(hdrs, r + [""] * (len(hdrs) - len(r))))
                roles = sheets.parse_roles(row.get("role", ""))
                if "assembler" not in roles:
                    continue
                atg = (row.get("tg_id") or "").strip()
                if not atg:
                    continue
                full_name = (f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
                             or row.get("tg_username", "") or atg)
                on_probation = str(row.get("on_probation", "")).lower() in ("1", "true", "yes")

                # Рейтинг: звёзды × 15 + завершённые × 10 − активные × 5 − испытательный × 20
                comp = completed_count.get(atg, 0)
                active = active_count.get(atg, 0)
                avg_stars = _get_avg_stars(atg)
                star_bonus = round((avg_stars - 3) * 15) if avg_stars else 0
                score = star_bonus + comp * 10 - active * 5 - (20 if on_probation else 0)

                assemblers.append({
                    "tg_id": atg,
                    "name": full_name,
                    "tg_username": row.get("tg_username", ""),
                    "on_probation": on_probation,
                    "completed_count": comp,
                    "active_count": active,
                    "avg_stars": avg_stars,
                    "score": score,
                    "_busy": busy_dates.get(atg, set()),
                })
    except Exception as e:
        log.warning("suggest_slots assemblers: %s", e)
        return {"error": "sheets_error"}

    # Сортируем по убыванию рейтинга
    assemblers.sort(key=lambda x: x["score"], reverse=True)

    # ── Генерируем свободные слоты на 14 дней ────────────────────────
    now = datetime.now(timezone.utc)
    # Начинаем со следующего дня (сегодня уже поздно)
    slots_hours = [9, 14]  # 09:00 и 14:00
    result = []

    for asm in assemblers[:6]:  # не больше 6 сборщиков
        busy = asm.pop("_busy")
        free_slots = []
        for day_offset in range(1, 15):
            d = now + timedelta(days=day_offset)
            day_str = d.strftime("%Y-%m-%d")
            if day_str in busy:
                continue
            # Пропускаем воскресенье (6)
            if d.weekday() == 6:
                continue
            for h in slots_hours:
                slot_iso = f"{day_str}T{h:02d}:00"
                free_slots.append(slot_iso)
                if len(free_slots) >= 6:
                    break
            if len(free_slots) >= 6:
                break
        asm["free_slots"] = free_slots
        result.append(asm)

    return {"ok": True, "assemblers": result}


def _handle_assembly_propose_date(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер предлагает дату сборки клиенту.
    body: {initData, assembly_id, proposed_date: ISO, assign_assembler_tg_id?: str}"""
    tg_id, err = _auth_manager_only(body)
    if err:
        return err

    assembly_id = (body.get("assembly_id") or "").strip()
    proposed_date = (body.get("proposed_date") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}
    if not proposed_date:
        return {"error": "missing_proposed_date"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}
    if str(asm.get("manager_tg_id", "")) != str(tg_id):
        return {"error": "forbidden"}

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "proposed_date", proposed_date)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "client_date_status", "pending")
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "client_preferred_date", "")

    # Опционально: назначить сборщика одновременно с предложением даты
    assembler_name = ""
    assign_tg_id = (body.get("assign_assembler_tg_id") or "").strip()
    if assign_tg_id:
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "assigned_to_tg_id", assign_tg_id)
        # Уведомляем сборщика
        try:
            asm_user = sheets.find_user(int(assign_tg_id))
            if asm_user:
                assembler_name = (asm_user.get("full_name") or
                    f"{asm_user.get('first_name','')} {asm_user.get('last_name','')}".strip())
            cfg = get_config()
            tg.send_message(
                cfg.bot_token, int(assign_tg_id),
                f"🔨 Вас назначили на сборку кухни!\n"
                f"📍 {asm.get('address', '')}\n"
                f"📅 Предлагаемая дата: <b>{_fmt_dt_ru(proposed_date)}</b>\n"
                f"(ожидаем подтверждения клиента)",
                parse_mode="HTML",
            )
        except Exception as e:
            log.warning("propose_date tg notify assembler: %s", e)

    # Telegram клиенту
    client_tg_id = (asm.get("client_tg_id") or "").strip()
    if client_tg_id:
        date_str = _fmt_dt_ru(proposed_date)
        master_line = f"\n👷 Мастер: <b>{assembler_name}</b>" if assembler_name else ""
        try:
            cfg = get_config()
            tg.send_message(
                cfg.bot_token, int(client_tg_id),
                f"📅 Менеджер предлагает дату сборки кухни:\n"
                f"<b>{date_str}</b>{master_line}\n\n"
                f"📍 {asm.get('address', '')}\n\n"
                f"Откройте приложение, чтобы подтвердить или предложить другое время.",
                parse_mode="HTML",
            )
        except Exception as e:
            log.warning("propose_date tg notify client: %s", e)

    return {"ok": True, "proposed_date": proposed_date, "assembler_assigned": bool(assign_tg_id)}


def _handle_assembly_date_confirm(body: dict[str, Any]) -> dict[str, Any]:
    """Клиент подтверждает предложенную дату сборки.
    body: {initData, assembly_id}"""
    tg_id, user, err = _auth_any_user(body)
    if err:
        return err

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    # Только клиент этой сборки
    if str(asm.get("client_tg_id", "")) != str(tg_id):
        return {"error": "forbidden"}

    proposed = (asm.get("proposed_date") or "").strip()
    if not proposed:
        return {"error": "no_proposed_date"}

    # Подтверждаем — ставим scheduled_at
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "scheduled_at", proposed)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "client_date_status", "confirmed")
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "proposed_date", "")

    # Если сборка ещё в created — переводим в scheduled
    if asm.get("status") == "created":
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status", "scheduled")

    # Telegram менеджеру
    mgr_tg_id = (asm.get("manager_tg_id") or "").strip()
    if mgr_tg_id:
        date_str = _fmt_dt_ru(proposed)
        client_name = asm.get("client_name", "Клиент")
        try:
            cfg = get_config()
            tg.send_message(
                cfg.bot_token, int(mgr_tg_id),
                f"✅ <b>{client_name}</b> подтвердил дату сборки:\n"
                f"<b>{date_str}</b>\n"
                f"📍 {asm.get('address', '')}",
                parse_mode="HTML",
            )
        except Exception as e:
            log.warning("date_confirm tg notify manager: %s", e)

    return {"ok": True, "scheduled_at": proposed}


def _handle_assembly_date_decline(body: dict[str, Any]) -> dict[str, Any]:
    """Клиент отклоняет предложенную дату, предлагает своё время.
    body: {initData, assembly_id, preferred_date?: ISO}"""
    tg_id, user, err = _auth_any_user(body)
    if err:
        return err

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    if str(asm.get("client_tg_id", "")) != str(tg_id):
        return {"error": "forbidden"}

    preferred = (body.get("preferred_date") or "").strip()

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "client_date_status", "declined")
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "client_preferred_date", preferred)

    # Telegram менеджеру
    mgr_tg_id = (asm.get("manager_tg_id") or "").strip()
    if mgr_tg_id:
        client_name = asm.get("client_name", "Клиент")
        msg_lines = [f"❌ <b>{client_name}</b> не может в предложенное время."]
        if preferred:
            msg_lines.append(f"Предлагает: <b>{_fmt_dt_ru(preferred)}</b>")
        else:
            msg_lines.append("Альтернативное время не указано — свяжитесь с клиентом.")
        msg_lines.append(f"📍 {asm.get('address', '')}")
        try:
            cfg = get_config()
            tg.send_message(cfg.bot_token, int(mgr_tg_id), "\n".join(msg_lines), parse_mode="HTML")
        except Exception as e:
            log.warning("date_decline tg notify manager: %s", e)

    return {"ok": True}


# Маппинг тип фото → префикс имени файла (по чек-листу замера)
_PHOTO_KIND_PREFIX = {
    "wall1":   "w1",
    "wall2":   "w2",
    "wall3":   "w3",
    "wall4":   "w4",
    "plan":    "plan",
    "general": "general",
    "detail":  "detail",
}


def _save_measurement_photo(
    measurement_id: str, idx: int, data_url: str,
    kind: str | None = None, kind_seq: int = 0,
) -> str | None:
    """Сохраняет фото с осмысленным именем: `w1.jpg` / `plan.jpg` / `general_3.jpg`.
    Если несколько фото одного типа — добавляем суффикс _2, _3.
    Возвращает имя файла или None при ошибке."""
    if not isinstance(data_url, str):
        return None
    m = _DATA_URL_RE.match(data_url.strip())
    if not m:
        return None
    ext = "jpg" if m.group(1) in ("jpeg", "jpg") else m.group(1)
    try:
        raw = base64.b64decode(m.group(2), validate=False)
    except Exception:
        return None
    if len(raw) > 10 * 1024 * 1024:  # 10 MB hard cap
        return None
    if not _SAFE_ID_RE.match(measurement_id):
        return None
    target_dir = PHOTOS_DIR / measurement_id
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        prefix = _PHOTO_KIND_PREFIX.get(kind or "", "")
        if prefix:
            # wall1/2/3/4 — один на стену; если дубль — добавляем _2, _3...
            if prefix.startswith("w") and len(prefix) == 2:
                name_base = prefix
                candidate = f"{name_base}.{ext}"
                n = 2
                while (target_dir / candidate).exists():
                    candidate = f"{name_base}_{n}.{ext}"
                    n += 1
                name = candidate
            else:
                # plan / general / detail — могут быть множественные
                name = f"{prefix}_{kind_seq}.{ext}" if kind_seq > 0 else f"{prefix}.{ext}"
                if (target_dir / name).exists():
                    name = f"{prefix}_{kind_seq + 1}.{ext}"
        else:
            name = f"{idx}.{ext}"
        (target_dir / name).write_bytes(raw)
        return name
    except Exception:
        log.warning("Не удалось сохранить фото %d для замера %s", idx, measurement_id)
        return None


def _measurement_columns() -> list[str]:
    """Гарантирует что у листа Measurements есть все нужные колонки (расширенный набор)."""
    return [
        "id", "ts", "client_tg_id", "manager_tg_id", "filled_by",
        "layout", "area_m2", "ceiling_mm", "walls", "openings", "infra", "niches",
        "photos", "notes", "status",
        # Поля Commit B (workflow)
        "assigned_to_tg_id", "requested_by_tg_id", "scheduled_at",
        "address", "client_name", "client_phone",
        # Поля Commit C (структура замера по чек-листу)
        "zamer_no", "zamer_date", "floor_base", "photos_meta",
        # Поля для приблизительной даты от менеджера (Commit C2)
        # preferred_type: specific | this_week | next_week | tbd
        # preferred_date: ISO date если specific
        # preferred_time_of_day: morning | day | evening
        # preferred_note: «после звонка», «не раньше вторника», ...
        "preferred_type", "preferred_date", "preferred_time_of_day", "preferred_note",
        # Логистика — заполняет замерщик на месте (Commit C3), нужна также сборщику
        # parking_type: free | paid | street | none
        "entrance", "floor", "gps_lat", "gps_lng",
        "parking_type", "parking_note", "delivery_notes",
        # Google Calendar — событие при scheduled
        "gcal_event_id", "gcal_event_url",
        # Идентификаторы клиента и договора (нумерация)
        "client_no", "contract_no", "contract_date",
        # Soft-delete
        "archived_at",
        # Чертёж/документы — DWG, PDF, PNG превью (B)
        "design_files",
        # Решение менеджера про подбор техники после замера (E)
        # podbor_decision: pending | needed | not_needed | later | done
        # podbor_decision_at — когда зафиксировано решение
        "podbor_decision", "podbor_decision_at", "podbor_lead_id",
        # Оплата замера
        "measurement_fee",
        "rooms_count",  # количество помещений для замера
        # Обратная связь замерщика о менеджере
        "measurer_feedback_at",
        # Обратная связь менеджера о замерщике
        "manager_feedback_at",
    ]


def _ensure_measurements_sheet() -> None:
    """Канонизирует схему Measurements:
    1. Создаёт лист если отсутствует.
    2. Добавляет недостающие колонки.
    3. Если порядок колонок не совпадает с _measurement_columns() —
       перестраивает лист: читает все данные, переставляет колонки по канону,
       перезаписывает лист целиком. Данные не теряются.
    """
    want = _measurement_columns()

    # --- Создать лист если не существует ---
    try:
        ws = sheets.sheet("Measurements")
        existing = ws.row_values(1)
    except Exception:
        sheets.ensure_sheet("Measurements", want)
        log.info("Measurements: создан с каноническим заголовком")
        return

    if not existing:
        ws.update("A1", [want])
        log.info("Measurements: заголовок установлен (лист был пуст)")
        return

    # --- Добавить недостающие колонки (без нарушения порядка) ---
    missing = [c for c in want if c not in existing]
    if missing:
        # Дописываем только в конец; данные встают в те позиции,
        # которые append_named_row() потом найдёт по имени
        ws.update("A1", [existing + missing])
        existing = existing + missing
        log.info("Measurements: добавлены колонки %s", missing)

    # --- Канонизация порядка если он не совпадает ---
    # Берём только колонки из канона (extra-колонки вне канона сохраняем справа)
    canon_set = set(want)
    extra = [c for c in existing if c not in canon_set]
    canonical_order = want + extra           # канон + внекановые справа

    if existing == canonical_order:
        return  # уже в правильном порядке — ничего не делаем

    log.info("Measurements: обнаружен неканонический порядок колонок — запускаем миграцию")
    try:
        all_rows = ws.get_all_values()
        if len(all_rows) < 2:
            # Данных нет — просто переписать заголовок
            ws.update("A1", [canonical_order])
            log.info("Measurements: заголовок канонизирован (данных не было)")
            return

        old_headers = all_rows[0]
        data_rows   = all_rows[1:]

        # Перестраиваем каждую строку: читаем по имени, пишем в канонический порядок
        new_rows = []
        for r in data_rows:
            old_dict = dict(zip(old_headers, r + [""] * max(0, len(old_headers) - len(r))))
            new_rows.append([old_dict.get(col, "") for col in canonical_order])

        # Перезаписываем лист целиком
        ws.clear()
        ws.update("A1", [canonical_order] + new_rows, value_input_option="USER_ENTERED")
        log.info("Measurements: миграция завершена, %d строк пересортировано", len(new_rows))
    except Exception as e:
        log.error("Measurements: ошибка канонизации (данные не тронуты): %s", e)


def _row_for_measurement(measurement_id: str, ts: str, **fields) -> dict[str, str]:
    """Возвращает словарь колонка→значение для записи в Measurements.
    Используется с sheets.append_named_row() — безопасно к порядку колонок."""
    base: dict[str, Any] = {
        "id": measurement_id, "ts": ts,
        "client_tg_id": "", "manager_tg_id": "", "filled_by": "",
        "layout": "", "area_m2": "", "ceiling_mm": "",
        "walls": "{}", "openings": "{}", "infra": "{}", "niches": "{}",
        "photos": "", "notes": "", "status": "submitted",
        "assigned_to_tg_id": "", "requested_by_tg_id": "", "scheduled_at": "",
        "address": "", "client_name": "", "client_phone": "",
        "zamer_no": "", "zamer_date": "", "floor_base": "", "photos_meta": "",
        "preferred_type": "", "preferred_date": "", "preferred_time_of_day": "", "preferred_note": "",
        "entrance": "", "floor": "", "gps_lat": "", "gps_lng": "",
        "parking_type": "", "parking_note": "", "delivery_notes": "",
        "gcal_event_id": "", "gcal_event_url": "",
        "client_no": "", "contract_no": "", "contract_date": "",
        "archived_at": "",
        "design_files": "",
        "podbor_decision": "", "podbor_decision_at": "", "podbor_lead_id": "",
    }
    base.update(fields)
    # Нормализуем: None → "", всё приводим к str
    return {k: str(v) if v is not None else "" for k, v in base.items()}


def _handle_measurement(body: dict[str, Any]) -> dict[str, Any]:
    """Полная сдача замера (когда форма заполнена). Поддерживает 2 режима:
    1. Создать новый замер с данными (старый MVP-режим — сам менеджер сделал замер)
    2. Обновить существующий request — статус → completed (для замерщика после посещения)
    """
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    _ensure_measurements_sheet()
    m = body.get("measurement") or {}
    existing_id = (m.get("measurement_id") or body.get("measurement_id") or "").strip()
    update_mode = bool(existing_id)
    is_manager = sheets.has_role(user, "manager")
    is_measurer = sheets.has_role(user, "measurer")

    measurement_id = existing_id or _short_id()
    if is_manager and not is_measurer:
        filled_by = "manager_for_client"
    elif is_measurer:
        filled_by = "measurer"
    else:
        filled_by = "client_self"

    # При update-mode загружаем существующую заявку и проверяем права
    existing_row = None
    if update_mode:
        existing_row = sheets.find_row("Measurements", "id", measurement_id)
        if not existing_row:
            return {"error": "measurement_not_found"}
        # Только назначенный замерщик или менеджер-владелец могут завершить
        if str(existing_row.get("assigned_to_tg_id")) != str(tg_id) and \
           str(existing_row.get("manager_tg_id")) != str(tg_id):
            return {"error": "forbidden"}

    client_tg_id = (existing_row or {}).get("client_tg_id") or \
                   (m.get("client_tg_id") if is_manager else tg_id) or ""
    manager_tg_id = (existing_row or {}).get("manager_tg_id") or (
        tg_id if is_manager else
        (sheets.find_row("Clients", "tg_id", tg_id) or {}).get("manager_tg_id", "")
    )

    client_name = m.get("client_name") or (existing_row or {}).get("client_name", "")
    client_phone = m.get("client_phone") or (existing_row or {}).get("client_phone", "")
    address = m.get("address") or (existing_row or {}).get("address", "")
    assigned_to = (existing_row or {}).get("assigned_to_tg_id", "")
    requested_by = (existing_row or {}).get("requested_by_tg_id", manager_tg_id or "")
    scheduled_at = (existing_row or {}).get("scheduled_at", "")

    # Прикрепляем имя/телефон/адрес к notes (для совместимости со старым кодом)
    notes_full = m.get("notes", "")
    extras = []
    if client_name:
        extras.append(f"Клиент: {client_name}")
    if client_phone:
        extras.append(f"Тел: {client_phone}")
    if address:
        extras.append(f"Адрес: {address}")
    if extras:
        notes_full = " · ".join(extras) + ("\n" + notes_full if notes_full else "")

    # Сохраняем фотографии (data-URL → файлы), в Sheets кладём только имена.
    # Имена структурные: w1.jpg, plan.jpg, general_2.jpg — по типу из photos_meta.
    raw_photos = m.get("photos") or []
    photos_meta = m.get("photos_meta") or []
    saved_photos: list[str] = []
    kind_counter: dict[str, int] = {}  # сколько раз уже встречался каждый kind
    if isinstance(raw_photos, list):
        for i, p in enumerate(raw_photos[:30]):  # хард-кап 30 фото на замер
            kind = ""
            if i < len(photos_meta) and isinstance(photos_meta[i], dict):
                kind = photos_meta[i].get("kind", "")
            kind_counter[kind] = kind_counter.get(kind, 0) + 1
            seq = kind_counter[kind] - 1  # 0 для первого, 1 для второго и т.д.
            if isinstance(p, str) and p.startswith("data:"):
                fn = _save_measurement_photo(measurement_id, i, p, kind=kind, kind_seq=seq)
                if fn:
                    saved_photos.append(fn)
            elif isinstance(p, str) and p and not p.startswith("data:"):
                saved_photos.append(p)

    status_new = "completed"
    walls_json = json.dumps(m.get("walls") or {}, ensure_ascii=False)
    openings_json = json.dumps(m.get("openings") or {}, ensure_ascii=False)
    infra_json = json.dumps(m.get("infra") or {}, ensure_ascii=False)
    niches_json = json.dumps(m.get("niches") or {}, ensure_ascii=False)
    photos_str = ",".join(saved_photos)
    photos_meta_json = json.dumps(m.get("photos_meta") or [], ensure_ascii=False)

    zamer_no   = (m.get("zamer_no")   or "").strip()
    zamer_date = (m.get("zamer_date") or "").strip()
    floor_base = (m.get("floor_base") or "").strip()

    if update_mode:
        # Обновляем существующую заявку — статус → completed, плюс заполняем поля
        updates = {
            "filled_by": filled_by,
            "layout": m.get("layout", ""),
            "area_m2": m.get("area_m2", ""),
            "ceiling_mm": m.get("ceiling_mm", ""),
            "walls": walls_json,
            "openings": openings_json,
            "infra": infra_json,
            "niches": niches_json,
            "photos": photos_str,
            "photos_meta": photos_meta_json,
            "notes": notes_full,
            "status": status_new,
            "zamer_no": zamer_no,
            "zamer_date": zamer_date,
            "floor_base": floor_base,
        }
        for col, val in updates.items():
            sheets.update_cell_by_key("Measurements", "id", measurement_id, col, val)
    else:
        sheets.append_named_row("Measurements", _row_for_measurement(
            measurement_id, _now_iso(),
            client_tg_id=client_tg_id or "",
            manager_tg_id=manager_tg_id or "",
            filled_by=filled_by,
            layout=m.get("layout", ""),
            area_m2=m.get("area_m2", ""),
            ceiling_mm=m.get("ceiling_mm", ""),
            walls=walls_json,
            openings=openings_json,
            infra=infra_json,
            niches=niches_json,
            photos=photos_str,
            photos_meta=photos_meta_json,
            notes=notes_full,
            status=status_new,
            assigned_to_tg_id=assigned_to,
            requested_by_tg_id=requested_by,
            scheduled_at=scheduled_at,
            address=address,
            client_name=client_name,
            client_phone=client_phone,
            zamer_no=zamer_no,
            zamer_date=zamer_date,
            floor_base=floor_base,
        ))

    if client_tg_id:
        sheets.update_cell_by_key("Clients", "tg_id", client_tg_id, "last_measurement_id", measurement_id)

    # Уведомления
    if update_mode and existing_row:
        # Замерщик завершил — пишем менеджеру который создавал заявку
        notify_to = requested_by or manager_tg_id
        if notify_to and str(notify_to) != str(tg_id):
            tg.send_message(
                notify_to,
                f"✅ <b>Замер выполнен</b>\n"
                f"Клиент: <b>{client_name or '—'}</b>\n"
                f"Замерщик: {user.get('full_name') or tg_id}\n"
                f"Фото: {len(saved_photos)} шт\n\n"
                f"❓ <b>Клиенту потребуется помощь с подбором техники?</b>\n"
                f"Откройте кабинет — на главной появится карточка с этим вопросом."
            )
    elif filled_by == "client_self" and manager_tg_id:
        tg.send_message(
            manager_tg_id,
            f"📐 Новый замер от клиента <b>{user.get('full_name') or tg_id}</b>.\n"
            f"Площадь: {m.get('area_m2', '?')} м², форма: {m.get('layout', '?')}.\n"
            f"Открыть в кабинете для просмотра."
        )

    sheets.log_event("measurement_submitted", tg_id, {
        "id": measurement_id, "filled_by": filled_by, "update_mode": update_mode,
    })
    return {"ok": True, "id": measurement_id, "photos": saved_photos, "status": status_new}


def _handle_podbor(body: dict[str, Any]) -> dict[str, Any]:
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not sheets.has_role(user, "manager"):
        return {"error": "only_manager_can_request_podbor"}

    checklist = body.get("checklist") or {}
    client_name = body.get("client_name", "")
    client_tg_id = body.get("client_tg_id", "")
    measurement_id = body.get("measurement_id", "")
    lead_id = _short_id()

    sheets.append_row("Leads", [
        lead_id, _now_iso(), tg_id, client_tg_id, client_name, measurement_id,
        json.dumps(checklist, ensure_ascii=False),
        "", "", 0, False, "new", 0,
    ])

    # Загружаем релевантный каталог моделей и передаём AI как «доступный пул»
    catalog_text = _build_catalog_context(checklist)

    user_prompt = (
        f"Подбери технику для следующего клиента:\n\n"
        f"{json.dumps({'client': {'name': client_name}, 'checklist': checklist}, ensure_ascii=False, indent=2)}"
    )
    if catalog_text:
        user_prompt += (
            "\n\n═══ ДОСТУПНЫЙ КАТАЛОГ МОДЕЛЕЙ (выбирай ТОЛЬКО из этого списка) ═══\n"
            + catalog_text
            + "\n\nВАЖНО: если модель не из этого списка — НЕ возвращай её. "
              "Каталог собран парсерами с реальных маркетплейсов РФ — это гарантия что артикул существует."
        )
    ai_result = ai.call_ai(user_prompt)

    # Обогащение моделей данными с маркетплейсов (WB / Я.Маркет / OZON / DNS)
    enrich_enabled = body.get("enrich", True)
    if enrich_enabled:
        try:
            _enrich_ai_marketplaces(ai_result)
        except Exception as e:
            log.warning("Marketplace enrich failed: %s", e)

    # Update lead row with AI response
    sheets.update_cell_by_key("Leads", "id", lead_id, "ai_response",
                              json.dumps(ai_result.get("json") or ai_result.get("text", ""), ensure_ascii=False))
    sheets.update_cell_by_key("Leads", "id", lead_id, "ai_model", ai_result.get("model", ""))
    sheets.update_cell_by_key("Leads", "id", lead_id, "ai_tokens_used", ai_result.get("tokens", 0))
    sheets.update_cell_by_key("Leads", "id", lead_id, "sent_to_tg", True)

    summary_text = _format_podbor_for_telegram(ai_result, client_name, lead_id)
    tg.send_message(tg_id, summary_text)

    sheets.log_event("podbor_completed", tg_id, {"id": lead_id, "tokens": ai_result.get("tokens", 0)})
    return {"ok": True, "id": lead_id, "summary": summary_text, "ai": ai_result.get("json")}


def _build_catalog_context(checklist: dict[str, Any]) -> str:
    """Готовит компактный текст-каталог для AI prompt.

    Берёт только релевантные категории + тиры (по budget_preset).
    """
    cats = checklist.get("categories") or []
    if not cats:
        return ""

    # Маппинг бюджет-пресета → тиры каталога
    bp = checklist.get("budget_preset") or ""
    tier_map = {
        "luxe":    ["premium"],
        "premium": ["premium", "middle"],
        "middle":  ["middle"],
        "budget":  ["middle", "budget"],  # средний и ниже
        "exact":   None,
    }
    tiers = tier_map.get(bp)  # None = все тиры

    try:
        return catalog.list_for_ai(cats, tiers=tiers, limit_per_cat=25)
    except Exception as e:
        log.warning("Cannot build catalog context: %s", e)
        return ""


def _enrich_ai_marketplaces(ai_result: dict[str, Any]) -> None:
    """Обогащает каждую модель из ai_result['json']['by_category'] данными
    с маркетплейсов (WB / Я.Маркет / OZON / DNS). Если PROXY6_TOKEN не задан —
    скорее всего вернёт пустые данные (Qrator блокирует прямые HTTP)."""
    j = ai_result.get("json")
    if not j or not isinstance(j, dict):
        return
    by_cat = j.get("by_category") or {}
    for cat_key, cat_data in by_cat.items():
        if not isinstance(cat_data, dict):
            continue
        models = cat_data.get("models") or []
        cat_data["models"] = parsers.enrich_models(models, delay_sec=0.4)


def _handle_clients(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает список клиентов менеджера со сводкой по подборам.
    Агрегирует клиентов из Leads И Measurements (включая draft-карточки)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    by_client: dict[str, dict[str, Any]] = {}

    def _ensure_client(key: str, name: str, phone: str, ctg_id: str):
        if key not in by_client:
            by_client[key] = {
                "client_name": name or "Без имени",
                "client_tg_id": ctg_id or None,
                "client_phone": phone or "",
                "address": "",
                "gps_lat": "",
                "gps_lng": "",
                "client_no": "",
                "contract_no": "",
                "contract_date": "",
                "leads_count": 0,
                "measurements_count": 0,
                "last_lead_at": "",
                "last_lead_id": "",
                "leads": [],
                # in_work=True если есть хотя бы один лид или замер не-draft
                "in_work": False,
            }
        else:
            # Заполним пустые поля если в этой записи есть данные
            c = by_client[key]
            if name and not c.get("client_name"): c["client_name"] = name
            if phone and not c.get("client_phone"): c["client_phone"] = phone
        return by_client[key]

    # 1. Из Leads — собираем подборы
    try:
        ws = sheets.sheet("Leads")
        rows = ws.get_all_values()
        if rows and len(rows) >= 2:
            headers = rows[0]
            for r in rows[1:]:
                row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
                if str(row.get("manager_tg_id", "")) != str(tg_id):
                    continue
                client_name = (row.get("client_name") or "").strip()
                client_tg_id = (row.get("client_tg_id") or "").strip()
                phone = ""
                checklist_str = row.get("checklist", "")
                if checklist_str:
                    try:
                        cl = json.loads(checklist_str)
                        phone = cl.get("client_phone", "") or ""
                    except (ValueError, TypeError):
                        pass
                key = client_tg_id or client_name.lower()
                if not key:
                    continue
                c = _ensure_client(key, client_name, phone, client_tg_id)
                c["leads_count"] += 1
                c["in_work"] = True  # есть подбор — клиент в работе
                lead_id = row.get("id", "")
                created_at = row.get("created_at", "")
                status = row.get("status", "")
                c["leads"].append({"id": lead_id, "created_at": created_at, "status": status})
                if created_at > c["last_lead_at"]:
                    c["last_lead_at"] = created_at
                    c["last_lead_id"] = lead_id
                    if phone and not c.get("client_phone"): c["client_phone"] = phone
    except Exception as e:
        log.warning("Failed to read Leads: %s", e)

    # 2. Из Measurements — для draft-карточек и заявок без подборов
    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
        if rows and len(rows) >= 2:
            headers = rows[0]
            for r in rows[1:]:
                row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
                if str(row.get("manager_tg_id", "")) != str(tg_id):
                    continue
                if row.get("archived_at"):
                    continue  # soft-deleted клиент
                client_name = (row.get("client_name") or "").strip()
                client_phone = (row.get("client_phone") or "").strip()
                client_tg_id = (row.get("client_tg_id") or "").strip()
                client_no = (row.get("client_no") or "").strip()
                contract_no = (row.get("contract_no") or "").strip()
                contract_date = (row.get("contract_date") or "").strip()
                address = (row.get("address") or "").strip()
                m_status = (row.get("status") or "").strip()
                key = client_tg_id or client_name.lower()
                if not key:
                    continue
                c = _ensure_client(key, client_name, client_phone, client_tg_id)
                if client_no and not c.get("client_no"): c["client_no"] = client_no
                if contract_no and not c.get("contract_no"): c["contract_no"] = contract_no
                if contract_date and not c.get("contract_date"): c["contract_date"] = contract_date
                if address and not c.get("address"): c["address"] = address
                if client_phone and not c.get("client_phone"): c["client_phone"] = client_phone
                gps_lat = (row.get("gps_lat") or "").strip()
                gps_lng = (row.get("gps_lng") or "").strip()
                if gps_lat and gps_lng and not c.get("gps_lat"): c["gps_lat"] = gps_lat; c["gps_lng"] = gps_lng
                c["measurements_count"] = c.get("measurements_count", 0) + 1
                # Замер не-draft = клиент в работе (requested/scheduled/completed)
                if m_status and m_status != "draft":
                    c["in_work"] = True
                # Если у клиента нет ни одного лида — last_at берём из measurement.ts
                ts = row.get("ts") or row.get("created_at") or ""
                if ts > c["last_lead_at"]:
                    c["last_lead_at"] = ts
    except Exception as e:
        log.warning("Failed to read Measurements for clients: %s", e)

    # 3. Из Assemblies — есть сборка = клиент в работе
    try:
        ws = sheets.sheet("Assemblies")
        rows = ws.get_all_values()
        if rows and len(rows) >= 2:
            headers = rows[0]
            for r in rows[1:]:
                row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
                if str(row.get("manager_tg_id", "")) != str(tg_id):
                    continue
                if row.get("archived_at"):
                    continue
                a_name = (row.get("client_name") or "").strip()
                a_ctg = (row.get("client_tg_id") or "").strip()
                a_phone = (row.get("client_phone") or "").strip()
                key = a_ctg or a_name.lower()
                if not key:
                    continue
                c = _ensure_client(key, a_name, a_phone, a_ctg)
                c["in_work"] = True
    except Exception:
        # Лист может ещё не существовать — не критично
        pass

    # Сортируем по дате последней активности (новые сверху)
    clients = sorted(by_client.values(), key=lambda x: x.get("last_lead_at") or "", reverse=True)
    for c in clients:
        c["leads"].sort(key=lambda x: x.get("created_at", ""), reverse=True)

    return {"ok": True, "count": len(clients), "clients": clients}


def _handle_lead(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает детали одного лида (включая AI-ответ и checklist)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]

    lead_id = body.get("lead_id") or body.get("id")
    if not lead_id:
        return {"error": "missing_lead_id"}

    row = sheets.find_row("Leads", "id", lead_id)
    if not row:
        return {"error": "lead_not_found"}

    # Проверяем что это лид этого менеджера
    if str(row.get("manager_tg_id", "")) != str(tg_id):
        return {"error": "forbidden"}

    # Парсим JSONы
    try:
        checklist = json.loads(row.get("checklist") or "{}")
    except (ValueError, TypeError):
        checklist = {}
    ai_response = row.get("ai_response") or ""
    try:
        ai_json = json.loads(ai_response) if ai_response else None
    except (ValueError, TypeError):
        ai_json = None

    return {
        "ok": True,
        "id": lead_id,
        "created_at": row.get("created_at"),
        "client_name": row.get("client_name"),
        "client_tg_id": row.get("client_tg_id"),
        "checklist": checklist,
        "ai": ai_json,
        "ai_text": ai_response if not ai_json else None,
        "status": row.get("status", ""),
    }


def _handle_grant_role(body: dict[str, Any]) -> dict[str, Any]:
    """Только админ может выдавать/отзывать роли."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    if int(tg_id) != int(cfg.admin_tg_id):
        return {"error": "admin_only"}

    target = body.get("target_tg_id")
    role = body.get("role", "").strip()
    action = body.get("action", "grant")
    if not target or not role:
        return {"error": "missing_fields"}
    try:
        target_int = int(target)
    except (TypeError, ValueError):
        return {"error": "bad_target"}

    if role not in sheets.VALID_ROLES:
        return {"error": "unknown_role", "valid": sorted(sheets.VALID_ROLES)}

    if action == "revoke":
        changed = sheets.revoke_role(target_int, role)
    else:
        changed = sheets.grant_role(target_int, role)

    sheets.log_event("role_changed", tg_id, {"target": target_int, "role": role, "action": action, "changed": changed})

    updated_user = sheets.find_user(target_int) or {}
    return {
        "ok": True,
        "target_tg_id": target_int,
        "changed": changed,
        "roles": sheets.parse_roles(updated_user.get("role", "")),
    }


def _handle_staff_list(body: dict[str, Any]) -> dict[str, Any]:
    """Список сотрудников с указанной ролью — для dropdown «выбрать замерщика»."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    role = (body.get("role") or "").strip()
    if role not in sheets.VALID_ROLES:
        return {"error": "unknown_role"}

    return {"ok": True, "role": role, "staff": sheets.list_users_with_role(role)}


def _handle_managers_list(body: dict[str, Any]) -> dict[str, Any]:
    """Список всех менеджеров — для dropdown «передать менеджеру»."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}
    managers = sheets.list_users_with_role("manager")
    # Исключаем самого себя (нет смысла передавать себе)
    managers = [m for m in managers if str(m.get("tg_id", "")) != str(tg_id)]
    return {"ok": True, "managers": managers}


def _handle_measurement_request(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер создаёт ЗАЯВКУ на замер (без замеров — пустая заготовка).
    body: {initData, client_name, client_phone, address, assigned_to_tg_id?, notes?, urgent?}
    urgent=True → немедленный push назначенному замерщику (или всем measurer-ам если не назначен)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    _ensure_measurements_sheet()
    client_name = (body.get("client_name") or "").strip()
    client_phone = (body.get("client_phone") or "").strip()
    address = (body.get("address") or "").strip()
    assigned_to = str(body.get("assigned_to_tg_id") or "").strip()
    notes = (body.get("notes") or "").strip()
    urgent = bool(body.get("urgent", False))

    # Приблизительная дата визита (Commit C2)
    rooms_count_req = body.get("rooms_count")

    # Приблизительная дата визита (Commit C2)
    preferred_type = (body.get("preferred_type") or "tbd").strip()
    preferred_date = (body.get("preferred_date") or "").strip()
    preferred_time_of_day = (body.get("preferred_time_of_day") or "").strip()
    preferred_note = (body.get("preferred_note") or "").strip()
    if preferred_type not in ("specific", "this_week", "next_week", "tbd"):
        preferred_type = "tbd"
    if preferred_time_of_day not in ("morning", "day", "evening", ""):
        preferred_time_of_day = ""

    if not client_name or not client_phone:
        return {"error": "missing_client_info", "hint": "client_name and client_phone are required"}

    # Если назначен — проверим что у него есть роль measurer
    if assigned_to:
        try:
            assigned_user = sheets.find_user(int(assigned_to))
        except (TypeError, ValueError):
            assigned_user = None
        if not assigned_user or not sheets.has_role(assigned_user, "measurer"):
            return {"error": "assigned_not_measurer"}

    # Передать другому менеджеру (любой менеджер может переслать заявку коллеге)
    target_manager_id = str(body.get("target_manager_tg_id") or "").strip()
    effective_manager_tg_id = tg_id  # по умолчанию — текущий
    if target_manager_id and target_manager_id != str(tg_id):
        try:
            target_mgr_user = sheets.find_user(int(target_manager_id))
        except (TypeError, ValueError):
            target_mgr_user = None
        if target_mgr_user and sheets.has_role(target_mgr_user, "manager"):
            effective_manager_tg_id = target_manager_id
        # Если целевой пользователь не найден или не менеджер — молча игнорируем

    measurement_id = _short_id()
    rooms_count_val = None
    if rooms_count_req is not None:
        try:
            rooms_count_val = str(max(1, int(rooms_count_req)))
        except (TypeError, ValueError):
            pass

    sheets.append_named_row("Measurements", _row_for_measurement(
        measurement_id, _now_iso(),
        manager_tg_id=effective_manager_tg_id,
        filled_by="request",
        status="requested",
        assigned_to_tg_id=assigned_to,
        requested_by_tg_id=tg_id,
        address=address,
        client_name=client_name,
        client_phone=client_phone,
        notes=notes,
        preferred_type=preferred_type,
        preferred_date=preferred_date,
        preferred_time_of_day=preferred_time_of_day,
        preferred_note=preferred_note,
        rooms_count=rooms_count_val,
    ))

    # Уведомляем целевого менеджера если передали ему
    if effective_manager_tg_id != str(tg_id):
        tg.send_message(
            int(effective_manager_tg_id),
            f"📋 <b>Вам передана заявка на замер</b>\n\n"
            f"Клиент: <b>{client_name}</b>\n"
            f"Телефон: <code>{client_phone}</code>\n"
            f"Адрес: {address or '—'}\n"
            f"От: {user.get('full_name') or tg_id}\n\n"
            f"Откройте кабинет — заявка уже в вашем списке."
        )

    # Уведомление назначенному замерщику
    if assigned_to:
        note_line = f"\nПримечание: {preferred_note}" if preferred_note else ""
        tg.send_message(
            int(assigned_to),
            f"📐 <b>Новая заявка на замер</b>\n\n"
            f"Клиент: <b>{client_name}</b>\n"
            f"Телефон: <code>{client_phone}</code>\n"
            f"Адрес: {address or '—'}\n"
            f"От менеджера: {user.get('full_name') or tg_id}"
            f"{note_line}\n\n"
            f"Откройте кабинет — согласуйте точную дату с клиентом."
        )

    # Срочный замер: push всем замерщикам или конкретному
    if urgent:
        scheduled_line = (
            f"📅 {preferred_date}" if preferred_type == "specific" and preferred_date
            else "📅 дата уточняется"
        )
        urgent_text = (
            f"⚡ <b>СРОЧНЫЙ ЗАМЕР</b>\n\n"
            f"📍 Адрес: {address or '—'}\n"
            f"{scheduled_line}\n"
            f"👤 {client_name}\n\n"
            f"Откройте MiniApp → Входящие"
        )
        if assigned_to:
            tg.send_message(int(assigned_to), urgent_text)
        else:
            measurers = sheets.find_users_by_role("measurer")
            for m in measurers:
                try:
                    m_tg_id = int(m.get("tg_id", 0))
                except (TypeError, ValueError):
                    continue
                if m_tg_id:
                    tg.send_message(m_tg_id, urgent_text)

    sheets.log_event("measurement_requested", tg_id, {
        "id": measurement_id, "assigned_to": assigned_to, "client": client_name, "urgent": urgent,
    })
    return {"ok": True, "id": measurement_id, "status": "requested", "assigned_to_tg_id": assigned_to}


def _handle_measurement_inbox(body: dict[str, Any]) -> dict[str, Any]:
    """Замерщик: список назначенных мне заявок (requested/scheduled/in_progress)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "measurer"):
        return {"error": "only_measurer"}

    _ensure_measurements_sheet()
    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception as e:
        log.warning("inbox read failed: %s", e)
        return {"ok": True, "measurements": []}

    if not rows or len(rows) < 2:
        return {"ok": True, "measurements": []}
    headers = rows[0]
    active_statuses = {"requested", "scheduled", "in_progress"}
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if str(row.get("assigned_to_tg_id", "")) != str(tg_id):
            continue
        if row.get("status") not in active_statuses:
            continue
        out.append({
            "id": row.get("id"),
            "created_at": row.get("ts"),
            "status": row.get("status"),
            "scheduled_at": row.get("scheduled_at", ""),
            "client_name": row.get("client_name", ""),
            "client_phone": row.get("client_phone", ""),
            "address": row.get("address", ""),
            "notes": row.get("notes", ""),
            "manager_tg_id": row.get("manager_tg_id", ""),
            "requested_by_tg_id": row.get("requested_by_tg_id", ""),
            "preferred_type": row.get("preferred_type", ""),
            "preferred_date": row.get("preferred_date", ""),
            "preferred_time_of_day": row.get("preferred_time_of_day", ""),
            "preferred_note": row.get("preferred_note", ""),
        })
    # Назначенная дата → первая; затем requested без даты
    def _sort_key(item):
        sched = item.get("scheduled_at") or ""
        return (0 if sched else 1, sched, item.get("created_at") or "")
    out.sort(key=_sort_key)
    return {"ok": True, "count": len(out), "measurements": out}


def _handle_measurement_schedule(body: dict[str, Any]) -> dict[str, Any]:
    """Замерщик назначает дату посещения. body: {initData, measurement_id, scheduled_at}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "measurer"):
        return {"error": "only_measurer"}

    measurement_id = (body.get("measurement_id") or "").strip()
    scheduled_at = (body.get("scheduled_at") or "").strip()
    if not measurement_id or not scheduled_at:
        return {"error": "missing_fields"}

    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}
    if str(row.get("assigned_to_tg_id")) != str(tg_id):
        return {"error": "forbidden"}

    sheets.update_cell_by_key("Measurements", "id", measurement_id, "scheduled_at", scheduled_at)
    sheets.update_cell_by_key("Measurements", "id", measurement_id, "status", "scheduled")

    # Google Calendar — создаём или обновляем событие
    gcal_url = ""
    try:
        from . import gcalendar
        existing_event_id = row.get("gcal_event_id") or ""
        client_name = row.get("client_name") or "—"
        address = row.get("address") or ""
        client_phone = row.get("client_phone") or ""
        descr_parts = [f"Клиент: {client_name}"]
        if client_phone: descr_parts.append(f"Телефон: {client_phone}")
        if row.get("preferred_note"): descr_parts.append(f"Примечание: {row.get('preferred_note')}")
        descr_parts.append(f"Замерщик: {user.get('full_name') or tg_id}")
        descr_parts.append(f"\nЗаявка: {measurement_id}")
        summary = f"Замер: {client_name}"
        description = "\n".join(descr_parts)

        if existing_event_id:
            ev = gcalendar.update_event(
                event_id=existing_event_id,
                summary=summary, description=description,
                start_iso=scheduled_at, location=address,
            )
        else:
            ev = gcalendar.create_event(
                summary=summary, description=description,
                start_iso=scheduled_at, location=address,
            )
        if ev:
            sheets.update_cell_by_key("Measurements", "id", measurement_id, "gcal_event_id", ev.get("id", ""))
            sheets.update_cell_by_key("Measurements", "id", measurement_id, "gcal_event_url", ev.get("html_link", ""))
            gcal_url = ev.get("html_link", "")
    except Exception as e:
        log.warning("GCal integration error: %s", e)

    # Уведомляем менеджера
    notify_to = row.get("requested_by_tg_id") or row.get("manager_tg_id")
    if notify_to and str(notify_to) != str(tg_id):
        try:
            cal_line = f"\n📅 <a href=\"{gcal_url}\">В календаре</a>" if gcal_url else ""
            tg.send_message(
                int(notify_to),
                f"📅 <b>Замер назначен</b>\n\n"
                f"Клиент: <b>{row.get('client_name') or '—'}</b>\n"
                f"Дата: {_format_date_human(scheduled_at)}\n"
                f"Замерщик: {user.get('full_name') or tg_id}\n"
                f"Адрес: {row.get('address') or '—'}"
                f"{cal_line}"
            )
        except Exception:
            pass

    sheets.log_event("measurement_scheduled", tg_id, {
        "id": measurement_id, "scheduled_at": scheduled_at, "gcal": bool(gcal_url),
    })
    return {
        "ok": True, "id": measurement_id, "status": "scheduled",
        "scheduled_at": scheduled_at, "gcal_event_url": gcal_url,
    }


def _handle_measurement_logistics(body: dict[str, Any]) -> dict[str, Any]:
    """Замерщик/сборщик/менеджер обновляет логистику замера —
    подъезд, этаж, GPS, парковка, заметки для логистов.
    body: {initData, measurement_id, entrance, floor, gps_lat, gps_lng,
           parking_type, parking_note, delivery_notes}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"], "_unsafe": True}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    measurement_id = (body.get("measurement_id") or "").strip()
    if not measurement_id:
        return {"error": "missing_measurement_id"}

    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}

    # Право редактировать — назначенный замерщик, менеджер-заказчик, или админ
    is_assigned_measurer = str(row.get("assigned_to_tg_id", "")) == str(tg_id)
    is_owner_manager = str(row.get("manager_tg_id", "")) == str(tg_id) or \
                       str(row.get("requested_by_tg_id", "")) == str(tg_id)
    is_assembler = sheets.has_role(user, "assembler")
    if not (is_assigned_measurer or is_owner_manager or is_assembler):
        return {"error": "forbidden"}

    # Валидация значений
    parking_type = (body.get("parking_type") or "").strip()
    if parking_type not in ("free", "paid", "street", "none", ""):
        parking_type = ""

    def _num_or_empty(v):
        if v is None or v == "":
            return ""
        try:
            return str(float(v))
        except (TypeError, ValueError):
            return ""

    updates = {
        "entrance":       (body.get("entrance")        or "").strip()[:80],
        "floor":          (body.get("floor")           or "").strip()[:20],
        "gps_lat":        _num_or_empty(body.get("gps_lat")),
        "gps_lng":        _num_or_empty(body.get("gps_lng")),
        "parking_type":   parking_type,
        "parking_note":   (body.get("parking_note")    or "").strip()[:200],
        "delivery_notes": (body.get("delivery_notes")  or "").strip()[:500],
    }
    for col, val in updates.items():
        sheets.update_cell_by_key("Measurements", "id", measurement_id, col, val)

    sheets.log_event("measurement_logistics_updated", tg_id, {"id": measurement_id})
    return {"ok": True, "id": measurement_id, "logistics": updates}


_DESIGN_DATA_URL_RE = re.compile(r"^data:([\w/\-+.]+);base64,(.+)$", re.DOTALL)
_DESIGN_ALLOWED_EXT = {"dwg", "dxf", "pdf", "png", "jpg", "jpeg", "webp"}
_DESIGN_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9_.\-]+")


def _save_design_file(measurement_id: str, data_url: str, raw_filename: str = "") -> str | None:
    """Сохраняет чертёж/документ (DWG, PDF, PNG) в PHOTOS_DIR/<id>/design_<n>.<ext>.
    Возвращает имя файла или None."""
    if not isinstance(data_url, str):
        return None
    m = _DESIGN_DATA_URL_RE.match(data_url.strip())
    if not m:
        return None
    mime = m.group(1).lower()
    try:
        raw = base64.b64decode(m.group(2), validate=False)
    except Exception:
        return None
    if len(raw) > 30 * 1024 * 1024:  # 30 MB cap (DWG могут быть тяжёлыми)
        return None
    if not _SAFE_ID_RE.match(measurement_id):
        return None

    # Расширение из mime, fallback на filename
    ext_map = {
        "application/x-autocad": "dwg",
        "image/vnd.dwg": "dwg",
        "application/acad": "dwg",
        "application/dxf": "dxf",
        "application/pdf": "pdf",
        "image/png": "png",
        "image/jpeg": "jpg",
        "image/jpg": "jpg",
        "image/webp": "webp",
    }
    ext = ext_map.get(mime, "")
    if not ext and raw_filename:
        rname = raw_filename.lower().rsplit(".", 1)
        if len(rname) == 2 and rname[1] in _DESIGN_ALLOWED_EXT:
            ext = rname[1]
    if not ext or ext not in _DESIGN_ALLOWED_EXT:
        return None

    target_dir = PHOTOS_DIR / measurement_id
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        # Подбираем уникальное имя design_1.ext, design_2.ext...
        n = 1
        while (target_dir / f"design_{n}.{ext}").exists():
            n += 1
        name = f"design_{n}.{ext}"
        # Если был передан осмысленный filename — используем его (sanitized)
        if raw_filename:
            base = raw_filename.rsplit(".", 1)[0]
            safe = _DESIGN_SAFE_NAME_RE.sub("_", base)[:60].strip("_")
            if safe:
                name = f"design_{safe}.{ext}"
                k = 1
                while (target_dir / name).exists():
                    k += 1
                    name = f"design_{safe}_{k}.{ext}"
        (target_dir / name).write_bytes(raw)
        return name
    except Exception:
        log.warning("Не удалось сохранить design-файл для %s", measurement_id)
        return None


def _handle_measurement_design_upload(body: dict[str, Any]) -> dict[str, Any]:
    """Загрузка чертежа/документа к замеру.
    body: {initData, measurement_id, files: [{name, data_url}, ...]}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    measurement_id = (body.get("measurement_id") or "").strip()
    if not measurement_id:
        return {"error": "missing_measurement_id"}
    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}
    # Право: менеджер-владелец, замерщик, технолог, админ
    is_owner = str(row.get("manager_tg_id")) == str(tg_id) or \
               str(row.get("requested_by_tg_id")) == str(tg_id) or \
               str(row.get("assigned_to_tg_id")) == str(tg_id)
    if not is_owner and not sheets.has_role(user, "manager"):
        return {"error": "forbidden"}

    files = body.get("files") or []
    if not isinstance(files, list) or not files:
        return {"error": "no_files"}

    saved = []
    for f in files[:10]:  # хард-кап 10 файлов за раз
        if not isinstance(f, dict):
            continue
        data_url = f.get("data_url") or ""
        name = f.get("name") or ""
        fn = _save_design_file(measurement_id, data_url, name)
        if fn:
            saved.append(fn)

    if not saved:
        return {"error": "no_valid_files"}

    # Объединяем с уже сохранёнными
    existing = [s for s in (row.get("design_files") or "").split(",") if s]
    combined = existing + saved
    sheets.update_cell_by_key("Measurements", "id", measurement_id, "design_files", ",".join(combined))

    sheets.log_event("design_uploaded", tg_id, {"id": measurement_id, "count": len(saved)})
    return {
        "ok": True,
        "id": measurement_id,
        "saved": saved,
        "total": len(combined),
        "design_files": combined,
    }


def _handle_measurement_decision(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер фиксирует решение про подбор техники после замера.
    body: {initData, measurement_id, decision: needed|not_needed|later|done, lead_id?}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    measurement_id = (body.get("measurement_id") or "").strip()
    decision = (body.get("decision") or "").strip()
    if decision not in ("needed", "not_needed", "later", "done"):
        return {"error": "bad_decision"}
    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}
    if str(row.get("manager_tg_id")) != str(tg_id) and str(row.get("requested_by_tg_id")) != str(tg_id):
        return {"error": "forbidden"}

    sheets.update_cell_by_key("Measurements", "id", measurement_id, "podbor_decision", decision)
    sheets.update_cell_by_key("Measurements", "id", measurement_id, "podbor_decision_at", _now_iso())
    lead_id = (body.get("lead_id") or "").strip()
    if lead_id:
        sheets.update_cell_by_key("Measurements", "id", measurement_id, "podbor_lead_id", lead_id)

    sheets.log_event("podbor_decision", tg_id, {"id": measurement_id, "decision": decision})
    return {"ok": True, "id": measurement_id, "decision": decision}


def _handle_measurement_add_photos(body: dict[str, Any]) -> dict[str, Any]:
    """Дозагрузка фото к существующему замеру.
    body: {initData, measurement_id, photos: [{data_url, label?}, ...]}
    label: before | after | general | extra (необязательно)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    measurement_id = (body.get("measurement_id") or "").strip()
    if not measurement_id or not _SAFE_ID_RE.match(measurement_id):
        return {"error": "missing_measurement_id"}

    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}

    is_owner = (str(row.get("manager_tg_id")) == str(tg_id)
                or str(row.get("requested_by_tg_id")) == str(tg_id)
                or str(row.get("assigned_to_tg_id")) == str(tg_id))
    if not is_owner and not sheets.has_role(user, "manager"):
        return {"error": "forbidden"}

    photos_input = body.get("photos") or []
    if not isinstance(photos_input, list) or not photos_input:
        return {"error": "no_photos"}

    existing = [p for p in (row.get("photos") or "").split(",") if p]
    saved: list[str] = []
    for i, p in enumerate(photos_input[:20]):
        if not isinstance(p, dict):
            continue
        data_url = p.get("data_url") or ""
        label = (p.get("label") or "extra").strip()
        if not isinstance(data_url, str) or not data_url.startswith("data:"):
            continue
        fn = _save_measurement_photo(measurement_id, len(existing) + i, data_url, kind=label)
        if fn:
            saved.append(fn)

    if not saved:
        return {"error": "no_photos_saved"}

    all_photos = existing + saved
    sheets.update_cell_by_key("Measurements", "id", measurement_id, "photos", ",".join(all_photos))
    sheets.log_event("measurement_photos_added", tg_id, {"id": measurement_id, "count": len(saved)})
    return {"ok": True, "id": measurement_id, "saved": saved, "total": len(all_photos)}


def _handle_measurement_set_status(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер меняет статус замера из карточки.
    body: {initData, measurement_id, status}
    Допустимые целевые статусы: cancelled, completed.
    Из draft/completed/cancelled — изменения запрещены."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    measurement_id = (body.get("measurement_id") or "").strip()
    new_status = (body.get("status") or "").strip()

    if not measurement_id:
        return {"error": "missing_measurement_id"}
    if new_status not in ("cancelled", "completed"):
        return {"error": "bad_status", "msg": "Допустимо: cancelled, completed"}

    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}

    # Только владелец-менеджер
    if str(row.get("manager_tg_id", "")) != str(tg_id):
        return {"error": "forbidden"}

    current = (row.get("status") or "").strip()
    if current in ("draft", "completed", "cancelled"):
        return {"error": "cannot_change", "msg": f"Статус «{current}» нельзя изменить"}
    # requested / scheduled → cancelled или completed
    sheets.update_cell_by_key("Measurements", "id", measurement_id, "status", new_status)
    sheets.log_event("measurement_status_changed", tg_id, {
        "id": measurement_id, "from": current, "to": new_status,
    })
    return {"ok": True, "id": measurement_id, "status": new_status, "prev_status": current}


def _handle_manager_pending(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает actionable карты для менеджера на главной:
    завершённые замеры где ещё не зафиксировано решение про подбор."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception:
        return {"ok": True, "pending": []}
    if not rows or len(rows) < 2:
        return {"ok": True, "pending": []}

    headers = rows[0]
    out = []
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if str(row.get("manager_tg_id", "")) != str(tg_id) and \
           str(row.get("requested_by_tg_id", "")) != str(tg_id):
            continue
        if row.get("archived_at"):
            continue
        if row.get("status") != "completed":
            continue
        decision = row.get("podbor_decision") or ""
        # Показываем pending (нет решения) + later (отложено) — для повторного предложения
        if decision in ("needed", "not_needed", "done"):
            continue
        out.append({
            "id": row.get("id", ""),
            "client_name": row.get("client_name", ""),
            "client_phone": row.get("client_phone", ""),
            "address": row.get("address", ""),
            "ts": row.get("ts", ""),
            "decision": decision,  # пусто или "later"
        })
    # Сортируем самые свежие сверху
    out.sort(key=lambda x: x.get("ts", ""), reverse=True)
    return {"ok": True, "count": len(out), "pending": out}


# =================================================================
# Сборки (Phase 4) — workflow от подписанного договора до приёмки
# =================================================================

def _assembly_columns() -> list[str]:
    return [
        "id", "ts",
        # Связи
        "manager_tg_id", "assigned_to_tg_id",
        "client_name", "client_phone", "address",
        "measurement_id", "lead_id", "client_tg_id",
        # Скоуп и расписание
        "scope_of_work",       # текстовое описание
        "scheduled_at",        # ISO
        # Статус: created | scheduled | in_progress | completed | cancelled
        "status",
        "started_at", "completed_at",
        # Фото-отчёт: списки имён файлов через запятую (внутри PHOTOS_DIR/<assembly_id>/)
        "photos_before", "photos_in_progress", "photos_after",
        # Приёмка / подпись (SignRequest)
        "sign_token", "sign_token_expires_at",
        "signed_via",          # canvas | code | proxy | absent
        "signed_by_name", "signed_by_tg_id", "signed_by_phone",
        "signature_file", "signed_at",
        # Google Calendar
        "gcal_event_id", "gcal_event_url",
        # Планирование: менеджер задаёт диапазон, мастер подтверждает конкретное время
        "date_range",          # текстовая подсказка от менеджера: "20–22 мая, утро"
        "confirm_by",          # ISO — дедлайн для подтверждения (назначение + 3 ч)
        "confirmed_at",        # ISO — когда мастер подтвердил время
        # Экспедитор (приёмка товара)
        "expeditor_tg_id",
        # Прочее
        "manager_note",
        "assembler_notes",      # заметки сборщика в процессе работы
        "kitchen_price",
        # Счёт клиенту на сборку
        "assembly_invoice_amount", "assembly_invoice_date",
        # Согласование даты с клиентом
        "proposed_date",         # ISO — дата предложенная менеджером клиенту
        "client_date_status",    # "pending" | "confirmed" | "declined"
        "client_preferred_date", # ISO — альтернатива от клиента
        # Обратная связь
        "client_feedback_at",    # ISO — когда клиент оставил оценку
        "archived_at",
        # Логистика (диспетчер)
        "shipment_date", "packages_count",
        "arrival_date", "arrival_packages_count", "arrival_confirmed_by_tg_id",
    ]


def _ensure_assemblies_sheet() -> None:
    """Догоняет схему Assemblies (добавляет недостающие колонки)."""
    try:
        ws = sheets.sheet("Assemblies")
        existing = ws.row_values(1)
    except Exception:
        sheets.ensure_sheet("Assemblies", _assembly_columns())
        return
    want = _assembly_columns()
    missing = [c for c in want if c not in existing]
    if missing:
        new_headers = existing + missing
        ws.update("A1", [new_headers])
        log.info("Assemblies: дополнили колонки: %s", missing)


def _row_for_assembly(assembly_id: str, ts: str, **fields) -> list[str]:
    cols = _assembly_columns()
    base = {c: "" for c in cols}
    base["id"] = assembly_id
    base["ts"] = ts
    base["status"] = "created"
    base.update(fields)
    return [str(base.get(c, "")) for c in cols]


def _handle_assembly_create(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер заводит сборку.
    body: {initData, client_name, client_phone?, address, scope_of_work,
           measurement_id?, lead_id?, scheduled_at?, manager_note?}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    _ensure_assemblies_sheet()

    client_name = (body.get("client_name") or "").strip()
    address = (body.get("address") or "").strip()
    scope = (body.get("scope_of_work") or "").strip()
    if not client_name:
        return {"error": "missing_client_name"}
    if not address:
        return {"error": "missing_address"}
    if not scope:
        return {"error": "missing_scope"}

    phone_raw = (body.get("client_phone") or "").strip()
    phone_norm, _ = _normalize_phone(phone_raw) if phone_raw else ("", False)

    assembly_id = _short_id()
    ts = _now_iso()
    scheduled_at = (body.get("scheduled_at") or "").strip()
    status = "scheduled" if scheduled_at else "created"

    assigned_to = (body.get("assigned_to_tg_id") or "").strip()
    date_range  = (body.get("date_range") or "").strip()
    # Дедлайн подтверждения: 3 часа с момента создания (если есть назначенный мастер)
    from datetime import timedelta
    confirm_by = (datetime.utcnow() + timedelta(hours=3)).isoformat() if assigned_to else ""

    fields = {
        "manager_tg_id": tg_id,
        "assigned_to_tg_id": assigned_to,
        "client_name": client_name,
        "client_phone": phone_norm or phone_raw,
        "address": address,
        "scope_of_work": scope,
        "measurement_id": (body.get("measurement_id") or "").strip(),
        "lead_id": (body.get("lead_id") or "").strip(),
        "client_tg_id": (body.get("client_tg_id") or "").strip(),
        "scheduled_at": scheduled_at,
        "status": status,
        "manager_note": (body.get("manager_note") or "").strip(),
        "date_range": date_range,
        "confirm_by": confirm_by,
    }

    # Google Calendar — если дата назначена
    if scheduled_at:
        try:
            from . import gcalendar
            ev = gcalendar.create_event(
                summary=f"🔨 Сборка: {client_name}",
                description=f"{scope}\n\nКлиент: {client_name}\nТел: {phone_norm or phone_raw}\nАдрес: {address}",
                start_iso=scheduled_at,
                duration_min=240,  # 4 часа на сборку
                location=address,
            )
            if ev:
                fields["gcal_event_id"] = ev.get("id", "")
                fields["gcal_event_url"] = ev.get("html_link", "")
        except Exception as e:
            log.warning("Не удалось создать событие Calendar для сборки: %s", e)

    sheets.append_row("Assemblies", _row_for_assembly(assembly_id, ts, **fields))
    sheets.log_event("assembly_created", tg_id, {"id": assembly_id, "client": client_name})

    return {"ok": True, "id": assembly_id, "status": status}


def _handle_assembly_list(body: dict[str, Any]) -> dict[str, Any]:
    """Список сборок.
    Менеджер: видит свои (manager_tg_id == self).
    Мастер: видит назначенные ему (assigned_to_tg_id == self) + неназначенные status='created'."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    is_manager = sheets.has_role(user, "manager")
    is_master = sheets.is_master(user)
    is_client = sheets.has_role(user, "client")
    if not is_manager and not is_master and not is_client:
        return {"error": "forbidden"}

    _ensure_assemblies_sheet()
    try:
        ws = sheets.sheet("Assemblies")
        rows = ws.get_all_values()
    except Exception:
        return {"ok": True, "assemblies": []}
    if not rows or len(rows) < 2:
        return {"ok": True, "assemblies": []}

    headers = rows[0]
    out = []
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if row.get("archived_at"):
            continue
        # Фильтр по роли
        visible = False
        if is_manager and str(row.get("manager_tg_id")) == str(tg_id):
            visible = True
        if is_master:
            if str(row.get("assigned_to_tg_id")) == str(tg_id):
                visible = True
            elif not row.get("assigned_to_tg_id") and row.get("status") in ("created", "scheduled"):
                visible = True
        if is_client and str(row.get("client_tg_id")) == str(tg_id):
            visible = True
        if not visible:
            continue
        out.append({
            "id": row.get("id", ""),
            "ts": row.get("ts", ""),
            "client_name": row.get("client_name", ""),
            "client_phone": row.get("client_phone", ""),
            "address": row.get("address", ""),
            "scope_of_work": row.get("scope_of_work", ""),
            "scheduled_at": row.get("scheduled_at", ""),
            "status": row.get("status", ""),
            "assigned_to_tg_id": row.get("assigned_to_tg_id", ""),
            "manager_tg_id": row.get("manager_tg_id", ""),
            "gcal_event_url": row.get("gcal_event_url", ""),
            "measurement_id": row.get("measurement_id", ""),
            "lead_id": row.get("lead_id", ""),
            "kitchen_price": row.get("kitchen_price", ""),
        })
    out.sort(key=lambda x: x.get("scheduled_at") or x.get("ts", ""), reverse=True)
    return {"ok": True, "count": len(out), "assemblies": out}


def _calc_assembly_prices(row: dict, viewer_tg_id) -> dict:
    """Вычисляет стоимости сборки с учётом ставок из Assembly_Rates.
    Возвращает словарь полей для добавления в ответ assembly_detail."""
    assembler_tg_id = str(row.get("assigned_to_tg_id") or "")
    client_rate, assembler_rate = _resolve_rates(assembler_tg_id, scope="*")
    is_assembler = str(viewer_tg_id) == assembler_tg_id

    try:
        kp = float(row.get("kitchen_price") or 0)
    except (ValueError, TypeError):
        kp = 0.0

    result: dict[str, Any] = {
        "client_rate_pct": client_rate,
        "assembler_rate_pct": assembler_rate,
        "assembly_price_for_client": round(kp * client_rate / 100, 2) if kp else None,
        "viewer_is_assembler": is_assembler,
    }
    # Сборщик видит свой заработок; менеджер и клиент — только цену для клиента
    if is_assembler or sheets.has_role(sheets.find_user(viewer_tg_id), "manager"):
        result["assembler_payout"] = round(kp * assembler_rate / 100, 2) if kp else None
    return result


def _handle_assembly_detail(body: dict[str, Any]) -> dict[str, Any]:
    """Детальная карточка сборки."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}
    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    # Право: менеджер-владелец, назначенный мастер, клиент-владелец, неназначенная сборка
    is_owner = str(row.get("manager_tg_id")) == str(tg_id) or \
               str(row.get("assigned_to_tg_id")) == str(tg_id) or \
               str(row.get("client_tg_id")) == str(tg_id)
    is_open_slot = (not row.get("assigned_to_tg_id")) and row.get("status") in ("created", "scheduled")
    if not is_owner and not is_open_slot:
        return {"error": "forbidden"}

    def _list(s: str) -> list[str]:
        return [x for x in (s or "").split(",") if x]

    # Контакт назначенного мастера (для клиента) + испытательный срок
    assigned_tg_id_str = row.get("assigned_to_tg_id", "")
    assigned_to_name = ""
    assigned_to_username = ""
    assigned_on_probation = False
    assigned_user = None
    if assigned_tg_id_str:
        try:
            assigned_user = sheets.find_user(int(assigned_tg_id_str))
            if assigned_user:
                assigned_to_name = assigned_user.get("full_name") or (
                    f"{assigned_user.get('first_name', '')} {assigned_user.get('last_name', '')}".strip())
                assigned_to_username = assigned_user.get("tg_username", "")
                assigned_on_probation = str(assigned_user.get("on_probation", "")).lower() in ("1", "true", "yes")
        except Exception:
            pass

    # Испытательный срок самого просматривающего (актуально для сборщика)
    viewer_on_probation = str(user.get("on_probation", "")).lower() in ("1", "true", "yes")

    # Act №4 summary (не блокирует при ошибке)
    act4_total = 0
    act4_damaged = 0
    act4_signed = False
    act4_signed_by = ""
    try:
        _ensure_act4_sheet()
        act4_row = sheets.find_row("Act4s", "assembly_id", assembly_id)
        if act4_row:
            items_raw = act4_row.get("items_json", "")
            if items_raw:
                items_parsed = json.loads(items_raw)
                act4_total   = sum(int(it.get("qty", 1)) for it in items_parsed)
                act4_damaged = sum(int(it.get("qty", 1)) for it in items_parsed if it.get("condition") == "damaged")
            act4_signed    = bool(act4_row.get("signed_by_name"))
            act4_signed_by = act4_row.get("signed_by_name", "")
    except Exception:
        pass

    return {
        "ok": True,
        "id": row.get("id", ""),
        "ts": row.get("ts", ""),
        "manager_tg_id": row.get("manager_tg_id", ""),
        "assigned_to_tg_id": row.get("assigned_to_tg_id", ""),
        "assigned_to_name": assigned_to_name,
        "assigned_to_username": assigned_to_username,
        "assigned_on_probation": assigned_on_probation,
        "expeditor_tg_id": row.get("expeditor_tg_id", ""),
        "viewer_tg_id": str(tg_id),
        "viewer_is_assembler": sheets.has_role(user, "assembler"),
        "viewer_is_manager": sheets.has_role(user, "manager"),
        "viewer_on_probation": viewer_on_probation,
        "client_name": row.get("client_name", ""),
        "client_phone": row.get("client_phone", ""),
        "address": row.get("address", ""),
        "measurement_id": row.get("measurement_id", ""),
        "lead_id": row.get("lead_id", ""),
        "scope_of_work": row.get("scope_of_work", ""),
        "scheduled_at": row.get("scheduled_at", ""),
        "status": row.get("status", ""),
        "started_at": row.get("started_at", ""),
        "completed_at": row.get("completed_at", ""),
        "photos_before": _list(row.get("photos_before", "")),
        "photos_in_progress": _list(row.get("photos_in_progress", "")),
        "photos_after": _list(row.get("photos_after", "")),
        "signature_file": row.get("signature_file", ""),
        "signed_via": row.get("signed_via", ""),
        "signed_by_name": row.get("signed_by_name", ""),
        "signed_by_tg_id": row.get("signed_by_tg_id", ""),
        "signed_by_phone": row.get("signed_by_phone", ""),
        "signed_at": row.get("signed_at", ""),
        "sign_token_expires_at": row.get("sign_token_expires_at", ""),
        "gcal_event_id": row.get("gcal_event_id", ""),
        "gcal_event_url": row.get("gcal_event_url", ""),
        "manager_note": row.get("manager_note", ""),
        "assembler_notes": row.get("assembler_notes", ""),
        "kitchen_price": row.get("kitchen_price", ""),
        "assembly_invoice_amount": row.get("assembly_invoice_amount", ""),
        "assembly_invoice_date":   row.get("assembly_invoice_date", ""),
        "client_tg_id": row.get("client_tg_id", ""),
        # Act4 summary
        "act4_total":     act4_total,
        "act4_damaged":   act4_damaged,
        "act4_signed":    act4_signed,
        "act4_signed_by": act4_signed_by,
        # Согласование даты с клиентом
        "proposed_date":         row.get("proposed_date", ""),
        "client_date_status":    row.get("client_date_status", ""),
        "client_preferred_date": row.get("client_preferred_date", ""),
        # Оценки
        "client_feedback_at":    row.get("client_feedback_at", ""),
        # Ставки — подсчёт в реальном времени
        **_calc_assembly_prices(row, tg_id),
    }


def _handle_assembly_set_kitchen_price(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер устанавливает стоимость кухни для сборки.
    body: {initData, assembly_id, kitchen_price}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    try:
        kitchen_price = float(body.get("kitchen_price") or 0)
    except (TypeError, ValueError):
        return {"error": "bad_kitchen_price", "msg": "kitchen_price должен быть числом"}
    if kitchen_price < 0:
        return {"error": "bad_kitchen_price", "msg": "kitchen_price не может быть отрицательным"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    if str(row.get("manager_tg_id")) != str(tg_id):
        return {"error": "forbidden"}

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "kitchen_price", str(kitchen_price))
    sheets.log_event("assembly_kitchen_price_set", tg_id, {
        "id": assembly_id, "kitchen_price": kitchen_price,
    })

    client_rate, assembler_rate = _resolve_rates(
        row.get("assigned_to_tg_id") or "", scope="*"
    )
    assembly_price = round(kitchen_price * client_rate / 100, 2)
    return {
        "ok": True,
        "kitchen_price": kitchen_price,
        "assembly_price": assembly_price,
        "client_rate_pct": client_rate,
        "assembler_rate_pct": assembler_rate,
    }


# =================================================================
# Assembly Rates — настройка % сборки (клиент / сборщик)
# =================================================================

_RATES_COLUMNS = [
    "rule_id", "assembler_tg_id", "assembler_name",
    "scope", "client_rate_pct", "assembler_rate_pct",
    "note", "active", "updated_by", "updated_at",
]
_DEFAULT_CLIENT_RATE    = 10.0
_DEFAULT_ASSEMBLER_RATE = 9.0
_rates_cache: dict = {"data": None, "ts": 0.0}
_RATES_CACHE_TTL = 120  # секунд


def _ensure_contracts_sheet() -> None:
    """Создаёт лист Contracts если не существует."""
    HEADERS = [
        "contract_id", "assembly_id", "contract_num", "contract_date",
        "travel_spb", "travel_outside", "tech_list",
        "created_at", "created_by_tg_id", "updated_at",
    ]
    try:
        wb = sheets._get_workbook()
        if "Contracts" not in [ws.title for ws in wb.worksheets()]:
            ws = wb.add_worksheet("Contracts", rows=200, cols=len(HEADERS))
            ws.append_row(HEADERS)
    except Exception as e:
        log.warning("_ensure_contracts_sheet: %s", e)


def _ensure_rates_sheet() -> None:
    try:
        ws = sheets._ws("Assembly_Rates")
        existing = ws.row_values(1)
    except Exception:
        sheets.ensure_sheet("Assembly_Rates", _RATES_COLUMNS)
        # seed default rule
        _seed_default_rate()
        return
    missing = [c for c in _RATES_COLUMNS if c not in existing]
    if missing:
        for col in missing:
            ws.update_cell(1, len(existing) + 1, col)
            existing.append(col)
    # Seed если лист пуст (только заголовок)
    try:
        all_rows = sheets.get_all_rows("Assembly_Rates")
        if not all_rows:
            _seed_default_rate()
    except Exception:
        pass


def _seed_default_rate() -> None:
    sheets.append_row("Assembly_Rates", _RATES_COLUMNS, {
        "rule_id": str(uuid.uuid4()),
        "assembler_tg_id": "*",
        "assembler_name": "Все сборщики",
        "scope": "*",
        "client_rate_pct": str(_DEFAULT_CLIENT_RATE),
        "assembler_rate_pct": str(_DEFAULT_ASSEMBLER_RATE),
        "note": "Базовая ставка по умолчанию",
        "active": "TRUE",
        "updated_by": "system",
        "updated_at": _now_iso(),
    })


def _get_rates_cached() -> list[dict]:
    now = time.time()
    if _rates_cache["data"] is None or (now - _rates_cache["ts"]) > _RATES_CACHE_TTL:
        try:
            _ensure_rates_sheet()
            rows = sheets.get_all_rows("Assembly_Rates")
            _rates_cache["data"] = [r for r in rows if r.get("active", "").upper() != "FALSE"]
            _rates_cache["ts"] = now
        except Exception as e:
            log.warning("_get_rates_cached error: %s", e)
            _rates_cache["data"] = _rates_cache["data"] or []
    return _rates_cache["data"] or []


def _resolve_rates(assembler_tg_id: str, scope: str = "*") -> tuple[float, float]:
    """Ищет наиболее специфичное правило для сборщика и типа работ.
    Приоритет: конкретный сборщик+scope > сборщик+* > *+scope > *+* > дефолт."""
    rules = _get_rates_cached()
    best_score = -1
    best_rule = None
    tid = str(assembler_tg_id).strip()
    for r in rules:
        rtid = str(r.get("assembler_tg_id", "*")).strip()
        rscope = str(r.get("scope", "*")).strip()
        score = 0
        if rtid == tid:
            score += 2
        elif rtid != "*":
            continue
        if rscope == scope:
            score += 1
        elif rscope != "*":
            continue
        if score > best_score:
            best_score = score
            best_rule = r
    if best_rule:
        try:
            cpct = float(best_rule.get("client_rate_pct", _DEFAULT_CLIENT_RATE))
            apct = float(best_rule.get("assembler_rate_pct", _DEFAULT_ASSEMBLER_RATE))
            return (cpct, apct)
        except (ValueError, TypeError):
            pass
    return (_DEFAULT_CLIENT_RATE, _DEFAULT_ASSEMBLER_RATE)


def _handle_assembly_rates_list(body: dict[str, Any]) -> dict[str, Any]:
    """Список всех правил ставок (включая неактивные). Доступен менеджеру."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "forbidden"}
    _ensure_rates_sheet()
    rows = sheets.get_all_rows("Assembly_Rates")
    return {"ok": True, "rates": rows}


def _handle_assembly_rate_save(body: dict[str, Any]) -> dict[str, Any]:
    """Создать или обновить правило ставки.
    body: {initData, rule_id?, assembler_tg_id, assembler_name,
           scope, client_rate_pct, assembler_rate_pct, note}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "forbidden"}

    try:
        cpct = float(body.get("client_rate_pct", _DEFAULT_CLIENT_RATE))
        apct = float(body.get("assembler_rate_pct", _DEFAULT_ASSEMBLER_RATE))
    except (ValueError, TypeError):
        return {"error": "bad_rate", "msg": "Ставка должна быть числом"}
    if not (0 < cpct <= 100) or not (0 < apct <= 100):
        return {"error": "bad_rate", "msg": "Ставка должна быть от 0.1 до 100"}
    if apct > cpct:
        return {"error": "bad_rate", "msg": "Ставка сборщика не может быть больше ставки клиента"}

    _ensure_rates_sheet()
    rule_id = (body.get("rule_id") or "").strip()
    now = _now_iso()

    if rule_id:
        # Обновляем существующее
        for field, val in [
            ("assembler_tg_id", str(body.get("assembler_tg_id") or "*")),
            ("assembler_name",  str(body.get("assembler_name") or "")),
            ("scope",           str(body.get("scope") or "*")),
            ("client_rate_pct", str(cpct)),
            ("assembler_rate_pct", str(apct)),
            ("note",            str(body.get("note") or "")),
            ("active",          "TRUE"),
            ("updated_by",      str(tg_id)),
            ("updated_at",      now),
        ]:
            sheets.update_cell_by_key("Assembly_Rates", "rule_id", rule_id, field, val)
    else:
        # Создаём новое
        rule_id = str(uuid.uuid4())
        sheets.append_row("Assembly_Rates", _RATES_COLUMNS, {
            "rule_id": rule_id,
            "assembler_tg_id": str(body.get("assembler_tg_id") or "*"),
            "assembler_name":  str(body.get("assembler_name") or ""),
            "scope":           str(body.get("scope") or "*"),
            "client_rate_pct": str(cpct),
            "assembler_rate_pct": str(apct),
            "note":            str(body.get("note") or ""),
            "active":          "TRUE",
            "updated_by":      str(tg_id),
            "updated_at":      now,
        })

    # Сбрасываем кеш
    _rates_cache["data"] = None
    return {"ok": True, "rule_id": rule_id}


def _handle_assembly_rate_delete(body: dict[str, Any]) -> dict[str, Any]:
    """Деактивирует правило ставки.
    body: {initData, rule_id}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "forbidden"}

    rule_id = (body.get("rule_id") or "").strip()
    if not rule_id:
        return {"error": "missing_rule_id"}
    _ensure_rates_sheet()
    sheets.update_cell_by_key("Assembly_Rates", "rule_id", rule_id, "active", "FALSE")
    sheets.update_cell_by_key("Assembly_Rates", "rule_id", rule_id, "updated_by", str(tg_id))
    sheets.update_cell_by_key("Assembly_Rates", "rule_id", rule_id, "updated_at", _now_iso())
    _rates_cache["data"] = None
    return {"ok": True}


# =================================================================
# Assembler Analytics — парсинг таблицы занятости сборщиков
# =================================================================

# Кэш распарсенного Excel в памяти (drive bytes → parse → aggregate)
_schedule_cache: dict = {"data": None, "ts": 0.0, "etag": ""}
_SCHEDULE_CACHE_TTL = 600  # 10 минут


_LOCAL_SCHEDULE_PATH = os.environ.get(
    "ASSEMBLER_SCHEDULE_PATH",
    "/app/data/assembler_schedule.xlsx"
)


def _get_schedule_data() -> dict:
    """Парсит таблицу занятости сборщиков. Кэш 10 мин.
    Источник: локальный файл (LOCAL) или Google Drive (DRIVE)."""
    import time as _time
    now = _time.monotonic()

    if _schedule_cache["data"] and (now - _schedule_cache["ts"]) < _SCHEDULE_CACHE_TTL:
        return _schedule_cache["data"]

    # Пробуем локальный файл
    if os.path.exists(_LOCAL_SCHEDULE_PATH):
        log.info("assembler_schedule: using local file %s", _LOCAL_SCHEDULE_PATH)
        parsed = assembler_parser.parse_file(_LOCAL_SCHEDULE_PATH)
    else:
        # Fallback: Google Drive
        cfg = get_config()
        file_id = cfg.assembler_schedule_file_id
        if not file_id:
            return {"error": "Файл не найден локально и ASSEMBLER_SCHEDULE_FILE_ID не задан"}
        try:
            xlsx_bytes = drive.download_file_bytes(file_id)
        except Exception as e:
            log.warning("assembler_schedule download error: %s", e)
            return {"error": f"download_failed: {e}"}
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tf.write(xlsx_bytes)
            tmp_path = tf.name
        try:
            parsed = assembler_parser.parse_file(tmp_path)
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    if "error" in parsed:
        return parsed

    records = parsed.get("records", [])
    agg = assembler_parser.aggregate(records)

    data = {
        "ok": True,
        "parsed_sheets": parsed.get("parsed_sheets", []),
        "total_records": len(records),
        "elapsed_s": parsed.get("elapsed_s"),
        "parsed_at": parsed.get("parsed_at"),
        "by_assembler": agg["by_assembler"],
        "by_month":     agg["by_month"],
    }
    _schedule_cache["data"] = data
    _schedule_cache["ts"] = now
    return data


def _handle_assembler_analytics(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает аналитику занятости/стоимостей сборщиков.
    body: {initData, year?: '2026', assembler_name?: 'Иванов И.И.'}
    Доступен менеджеру."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not (sheets.has_role(user, "manager") or sheets.has_role(user, "admin")):
        return {"error": "forbidden"}

    data = _get_schedule_data()
    if "error" in data:
        return data

    # Фильтр по году (если указан)
    year = str(body.get("year") or "").strip()
    if year and year.isdigit():
        by_month = {k: v for k, v in data["by_month"].items() if k.startswith(year)}
        by_assembler = {}
        for name, months in data["by_assembler"].items():
            filtered = {k: v for k, v in months.items() if k.startswith(year)}
            if filtered:
                by_assembler[name] = filtered
    else:
        by_month = data["by_month"]
        by_assembler = data["by_assembler"]

    # Топ-5 сборщиков по итоговой сумме
    assembler_totals = [
        {
            "name": name,
            "total_amount": sum(m["total_amount"] for m in months.values()),
            "total_orders": sum(m["orders"] for m in months.values()),
            "months": months,
        }
        for name, months in by_assembler.items()
    ]
    assembler_totals.sort(key=lambda x: x["total_amount"], reverse=True)

    return {
        "ok": True,
        "parsed_at": data.get("parsed_at"),
        "total_records": data.get("total_records"),
        "by_month": by_month,
        "assemblers": assembler_totals,
    }


@app.post("/api/assembler_analytics")
async def api_assembler_analytics(request: Request):
    body = await _safe_json(request)
    return _handle_assembler_analytics(body)


def _name_match_score(excel_name: str, full_name: str) -> int:
    """Возвращает score (0-3) схожести имени из Excel с full_name из Users."""
    en = excel_name.strip().lower()
    fn = full_name.strip().lower()
    if not en or not fn:
        return 0
    if en == fn:
        return 3
    # Первое слово (фамилия) совпадает
    en_first = en.split()[0] if en.split() else ""
    fn_first = fn.split()[0] if fn.split() else ""
    if en_first and fn_first and en_first == fn_first:
        # Дополнительно: совпадает второе слово или инициал
        en_parts = en.split()
        fn_parts = fn.split()
        if len(en_parts) > 1 and len(fn_parts) > 1:
            if en_parts[1] == fn_parts[1] or en_parts[1][:1] == fn_parts[1][:1]:
                return 2
        return 1
    return 0


def _handle_assembler_earnings(body: dict[str, Any]) -> dict[str, Any]:
    """Личная аналитика сборщика — его заработки из Excel-расписания.
    body: {initData, year?: '2026'}
    Доступен сборщику, замерщику, менеджеру."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "assembler") or sheets.has_role(user, "measurer") or
            sheets.has_role(user, "manager") or sheets.has_role(user, "admin")):
        return {"error": "forbidden"}

    full_name = (user.get("full_name") or "").strip()
    if not full_name:
        return {"error": "no_name", "message": "Имя не задано в профиле"}

    data = _get_schedule_data()
    if "error" in data:
        return data

    year = str(body.get("year") or "").strip()

    # Находим лучшее совпадение по имени
    best_name = None
    best_score = 0
    for excel_name in data.get("by_assembler", {}).keys():
        score = _name_match_score(excel_name, full_name)
        if score > best_score:
            best_score = score
            best_name = excel_name

    if not best_name or best_score == 0:
        return {
            "ok": True,
            "matched_name": None,
            "full_name": full_name,
            "months": {},
            "total_amount": 0,
            "total_orders": 0,
            "message": "Данные по вашему имени не найдены в таблице занятости",
        }

    months_raw = data["by_assembler"][best_name]
    if year and year.isdigit():
        months_raw = {k: v for k, v in months_raw.items() if k.startswith(year)}

    total_amount = sum(m["total_amount"] for m in months_raw.values())
    total_orders = sum(m["orders"] for m in months_raw.values())

    # Сортируем по дате desc
    months_sorted = dict(sorted(months_raw.items(), reverse=True))

    return {
        "ok": True,
        "matched_name": best_name,
        "full_name": full_name,
        "match_score": best_score,
        "months": months_sorted,
        "total_amount": total_amount,
        "total_orders": total_orders,
        "parsed_at": data.get("parsed_at"),
    }


@app.post("/api/assembler_earnings")
async def api_assembler_earnings(request: Request):
    body = await _safe_json(request)
    return _handle_assembler_earnings(body)


def _handle_staff_clients(body: dict[str, Any]) -> dict[str, Any]:
    """Список клиентов для сборщика / замерщика.
    Assembler: все сборки где assigned_to_tg_id == self.
    Measurer:  все замеры  где assigned_to_tg_id == self.
    Оба: объединённый список, сгруппированный по клиенту.
    body: {initData, initDataUnsafe, filter?: 'active'|'done'|'all'}
    """
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    is_assembler = sheets.has_role(user, "assembler") or sheets.is_master(user)
    is_measurer  = sheets.has_role(user, "measurer")
    if not is_assembler and not is_measurer and not sheets.has_role(user, "manager"):
        return {"error": "forbidden"}

    flt = str(body.get("filter") or "active").lower()  # active | done | all

    def _row_visible_active(status: str) -> bool:
        if flt == "all":   return True
        if flt == "done":  return status in ("done", "completed", "cancelled")
        return status not in ("done", "completed", "cancelled", "archived")

    clients: dict[str, dict] = {}  # key → client card

    # ── Сборки ──────────────────────────────────────────────────────
    if is_assembler or sheets.has_role(user, "manager"):
        try:
            _ensure_assemblies_sheet()
            ws = sheets.sheet("Assemblies")
            rows = ws.get_all_values()
            if rows and len(rows) > 1:
                headers = rows[0]
                for r in rows[1:]:
                    row = dict(zip(headers, r + [""] * max(0, len(headers) - len(r))))
                    if row.get("archived_at"):
                        continue
                    if sheets.has_role(user, "manager"):
                        if str(row.get("manager_tg_id")) != str(tg_id):
                            continue
                    else:
                        if str(row.get("assigned_to_tg_id")) != str(tg_id):
                            continue
                    status = row.get("status", "")
                    if not _row_visible_active(status):
                        continue
                    ckey = row.get("client_tg_id") or row.get("client_name", "").lower().strip()
                    if ckey not in clients:
                        clients[ckey] = {
                            "client_name":  row.get("client_name", ""),
                            "client_phone": row.get("client_phone", ""),
                            "client_tg_id": row.get("client_tg_id", ""),
                            "assemblies":   [],
                            "measurements": [],
                        }
                    clients[ckey]["assemblies"].append({
                        "id":           row.get("id", ""),
                        "address":      row.get("address", ""),
                        "status":       status,
                        "scheduled_at": row.get("scheduled_at", ""),
                        "scope_of_work": row.get("scope_of_work", ""),
                        "signed_by_name": row.get("signed_by_name", ""),
                        "manager_tg_id": row.get("manager_tg_id", ""),
                        "date_range":   row.get("date_range", ""),
                        "confirm_by":   row.get("confirm_by", ""),
                        "confirmed_at": row.get("confirmed_at", ""),
                    })
        except Exception as e:
            log.warning("staff_clients assemblies error: %s", e)

    # ── Замеры ───────────────────────────────────────────────────────
    if is_measurer or sheets.has_role(user, "manager"):
        try:
            ws2 = sheets.sheet("Measurements")
            rows2 = ws2.get_all_values()
            if rows2 and len(rows2) > 1:
                headers2 = rows2[0]
                for r in rows2[1:]:
                    row = dict(zip(headers2, r + [""] * max(0, len(headers2) - len(r))))
                    if row.get("archived_at"):
                        continue
                    if sheets.has_role(user, "manager"):
                        if str(row.get("manager_tg_id")) != str(tg_id):
                            continue
                    else:
                        if str(row.get("assigned_to_tg_id")) != str(tg_id):
                            continue
                    status = row.get("status", "")
                    if not _row_visible_active(status):
                        continue
                    ckey = row.get("client_tg_id") or row.get("client_name", "").lower().strip()
                    if ckey not in clients:
                        clients[ckey] = {
                            "client_name":  row.get("client_name", ""),
                            "client_phone": row.get("client_phone", ""),
                            "client_tg_id": row.get("client_tg_id", ""),
                            "assemblies":   [],
                            "measurements": [],
                        }
                    clients[ckey]["measurements"].append({
                        "id":              row.get("id", ""),
                        "address":         row.get("address", ""),
                        "status":          status,
                        "scheduled_at":    row.get("scheduled_at", ""),
                        "zamer_no":        row.get("zamer_no", ""),
                        "layout":          row.get("layout", ""),
                        "preferred_date":  row.get("preferred_date", ""),
                        "preferred_time_of_day": row.get("preferred_time_of_day", ""),
                    })
        except Exception as e:
            log.warning("staff_clients measurements error: %s", e)

    # ── Сортировка: сначала с ближайшей датой ───────────────────────
    def _latest_date(c: dict) -> str:
        dates = (
            [a["scheduled_at"] for a in c["assemblies"]  if a["scheduled_at"]] +
            [m["scheduled_at"] for m in c["measurements"] if m["scheduled_at"]]
        )
        return max(dates) if dates else ""

    result = sorted(clients.values(), key=_latest_date, reverse=True)

    return {
        "ok": True,
        "is_assembler": is_assembler,
        "is_measurer":  is_measurer,
        "count": len(result),
        "clients": result,
    }


@app.post("/api/staff_clients")
async def api_staff_clients(request: Request):
    body = await _safe_json(request)
    return _handle_staff_clients(body)


def _handle_assembly_schedule(body: dict[str, Any]) -> dict[str, Any]:
    """Мастер подтверждает конкретную дату/время сборки после созвона с клиентом.
    body: {initData, assembly_id, scheduled_at: ISO, note?}
    После подтверждения → уведомление менеджеру."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.is_master(user) or sheets.has_role(user, "assembler") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    assembly_id  = str(body.get("assembly_id") or "").strip()
    scheduled_at = str(body.get("scheduled_at") or "").strip()
    note         = str(body.get("note") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}
    if not scheduled_at:
        return {"error": "missing_scheduled_at"}

    _ensure_assemblies_sheet()
    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    # Только назначенный мастер или менеджер могут подтверждать
    is_assigned = str(asm.get("assigned_to_tg_id", "")) == str(tg_id)
    is_mgr = sheets.has_role(user, "manager") and str(asm.get("manager_tg_id", "")) == str(tg_id)
    if not is_assigned and not is_mgr:
        return {"error": "not_assigned"}

    now_iso = datetime.utcnow().isoformat()
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "scheduled_at",  scheduled_at)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "confirmed_at",  now_iso)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status",        "scheduled")
    if note:
        existing_note = asm.get("manager_note", "")
        new_note = f"{existing_note}\n[Подтверждение {now_iso[:10]}]: {note}".strip()
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, "manager_note", new_note)

    # Google Calendar — обновляем/создаём событие
    try:
        from . import gcalendar
        ev_id = asm.get("gcal_event_id", "")
        client_name = asm.get("client_name", "")
        address     = asm.get("address", "")
        scope       = asm.get("scope_of_work", "")
        phone       = asm.get("client_phone", "")
        staff_name  = user.get("full_name") or f"{user.get('first_name','')} {user.get('last_name','')}".strip() or str(tg_id)
        if ev_id:
            gcalendar.update_event(ev_id, start_iso=scheduled_at)
        else:
            ev = gcalendar.create_event(
                summary=f"🔨 Сборка: {client_name}",
                description=f"{scope}\n\nКлиент: {client_name}\nТел: {phone}\nАдрес: {address}\nМастер: {staff_name}",
                start_iso=scheduled_at,
                duration_min=240,
                location=address,
            )
            if ev:
                sheets.update_cell_by_key("Assemblies", "id", assembly_id, "gcal_event_id",  ev.get("id", ""))
                sheets.update_cell_by_key("Assemblies", "id", assembly_id, "gcal_event_url", ev.get("html_link", ""))
    except Exception as e:
        log.warning("assembly_schedule gcal error: %s", e)

    # Уведомление менеджеру
    manager_tg_id = asm.get("manager_tg_id", "")
    if manager_tg_id and str(manager_tg_id) != str(tg_id):
        try:
            staff_name = user.get("full_name") or f"{user.get('first_name','')} {user.get('last_name','')}".strip() or str(tg_id)
            dt_str = scheduled_at[:16].replace("T", " ")
            tg.send_message(
                int(manager_tg_id),
                f"✅ <b>Дата сборки согласована</b>\n\n"
                f"Клиент: <b>{asm.get('client_name','')}</b>\n"
                f"Адрес: {asm.get('address','')}\n"
                f"Дата: <b>{dt_str}</b>\n"
                f"Мастер: {staff_name}\n\n"
                f"Лид закреплён 🎯",
            )
        except Exception as e:
            log.warning("assembly_schedule notify error: %s", e)

    sheets.log_event("assembly_scheduled", tg_id, {"id": assembly_id, "scheduled_at": scheduled_at})
    return {"ok": True, "scheduled_at": scheduled_at}


@app.post("/api/assembly_schedule")
async def api_assembly_schedule(request: Request):
    body = await _safe_json(request)
    return _handle_assembly_schedule(body)


def _handle_measurement_schedule(body: dict[str, Any]) -> dict[str, Any]:
    """Замерщик подтверждает дату замера после созвона с клиентом.
    body: {initData, measurement_id, scheduled_at: ISO, note?}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "measurer") or sheets.is_master(user) or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    meas_id      = str(body.get("measurement_id") or "").strip()
    scheduled_at = str(body.get("scheduled_at") or "").strip()
    note         = str(body.get("note") or "").strip()
    if not meas_id or not scheduled_at:
        return {"error": "missing_fields"}

    meas = sheets.find_row("Measurements", "id", meas_id)
    if not meas:
        return {"error": "measurement_not_found"}

    is_assigned = str(meas.get("assigned_to_tg_id", "")) == str(tg_id)
    is_mgr = sheets.has_role(user, "manager") and str(meas.get("manager_tg_id", "")) == str(tg_id)
    if not is_assigned and not is_mgr:
        return {"error": "not_assigned"}

    now_iso = datetime.utcnow().isoformat()
    sheets.update_cell_by_key("Measurements", "id", meas_id, "scheduled_at",  scheduled_at)
    sheets.update_cell_by_key("Measurements", "id", meas_id, "status",        "scheduled")

    # Уведомление менеджеру
    manager_tg_id = meas.get("manager_tg_id", "")
    if manager_tg_id and str(manager_tg_id) != str(tg_id):
        try:
            staff_name = user.get("full_name") or f"{user.get('first_name','')} {user.get('last_name','')}".strip() or str(tg_id)
            dt_str = scheduled_at[:16].replace("T", " ")
            tg.send_message(
                int(manager_tg_id),
                f"📐 <b>Дата замера согласована</b>\n\n"
                f"Клиент: <b>{meas.get('client_name','')}</b>\n"
                f"Адрес: {meas.get('address','')}\n"
                f"Дата: <b>{dt_str}</b>\n"
                f"Замерщик: {staff_name}\n\n"
                f"Лид закреплён 🎯",
            )
        except Exception as e:
            log.warning("measurement_schedule notify error: %s", e)

    sheets.log_event("measurement_scheduled", tg_id, {"id": meas_id, "scheduled_at": scheduled_at})
    return {"ok": True, "scheduled_at": scheduled_at}


@app.post("/api/measurement_schedule")
async def api_measurement_schedule(request: Request):
    body = await _safe_json(request)
    return _handle_measurement_schedule(body)


def _handle_contract_preview(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает данные сборки + сохранённые поля контракта для предпросмотра акта.
    body: {initData, initDataUnsafe, assembly_id}
    Доступен менеджеру и сборщику."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "manager") or sheets.has_role(user, "assembler") or
            sheets.has_role(user, "admin")):
        return {"error": "forbidden"}

    assembly_id = str(body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    asm = sheets.find_row("Assemblies", "id", assembly_id)
    if not asm:
        return {"error": "assembly_not_found"}

    # Загружаем сохранённые поля контракта (если есть)
    contract = sheets.find_row("Contracts", "assembly_id", assembly_id) or {}

    return {
        "ok": True,
        "assembly": {
            "id":           asm.get("id", ""),
            "client_name":  asm.get("client_name", ""),
            "client_tg_id": asm.get("client_tg_id", ""),
            "address":      asm.get("address", ""),
            "scheduled_at": asm.get("scheduled_at", ""),
            "assembly_price_for_client": asm.get("assembly_price_for_client") or asm.get("kitchen_price", ""),
            "signed_by_name": asm.get("signed_by_name", ""),
            "signed_at":    asm.get("signed_at", ""),
            "signed_via":   asm.get("signed_via", ""),
            "status":       asm.get("status", ""),
        },
        "contract": {
            "contract_num":    contract.get("contract_num", assembly_id),
            "contract_date":   contract.get("contract_date", ""),
            "travel_spb":      contract.get("travel_spb", "0"),
            "travel_outside":  contract.get("travel_outside", "0"),
            "tech_list":       contract.get("tech_list", ""),
        },
    }


@app.post("/api/contract_preview")
async def api_contract_preview(request: Request):
    body = await _safe_json(request)
    return _handle_contract_preview(body)


def _handle_contract_save(body: dict[str, Any]) -> dict[str, Any]:
    """Сохраняет дополнительные поля акта в лист Contracts.
    body: {initData, initDataUnsafe, assembly_id, contract_num, contract_date, travel_spb, travel_outside, tech_list}
    Доступен менеджеру."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "manager") or sheets.has_role(user, "admin")):
        return {"error": "forbidden"}

    assembly_id = str(body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_contracts_sheet()

    contract_num   = str(body.get("contract_num")  or assembly_id).strip()
    contract_date  = str(body.get("contract_date") or "").strip()
    travel_spb     = str(body.get("travel_spb")    or "0").strip()
    travel_outside = str(body.get("travel_outside") or "0").strip()
    tech_list      = str(body.get("tech_list")     or "").strip()
    now_iso        = datetime.utcnow().isoformat()

    existing = sheets.find_row("Contracts", "assembly_id", assembly_id)
    if existing:
        # Обновляем существующую запись
        sheets.update_cell_by_key("Contracts", "assembly_id", assembly_id, "contract_num",   contract_num)
        sheets.update_cell_by_key("Contracts", "assembly_id", assembly_id, "contract_date",  contract_date)
        sheets.update_cell_by_key("Contracts", "assembly_id", assembly_id, "travel_spb",     travel_spb)
        sheets.update_cell_by_key("Contracts", "assembly_id", assembly_id, "travel_outside", travel_outside)
        sheets.update_cell_by_key("Contracts", "assembly_id", assembly_id, "tech_list",      tech_list)
        sheets.update_cell_by_key("Contracts", "assembly_id", assembly_id, "updated_at",     now_iso)
    else:
        # Создаём новую запись
        import uuid
        contract_id = str(uuid.uuid4())[:8]
        sheets.append_row("Contracts", [
            contract_id, assembly_id, contract_num, contract_date,
            travel_spb, travel_outside, tech_list,
            now_iso, str(tg_id), "",
        ])

    return {"ok": True}


@app.post("/api/contract_save")
async def api_contract_save(request: Request):
    body = await _safe_json(request)
    return _handle_contract_save(body)


# =================================================================
# Счёт на оплату замера (с QR-кодом ГОСТ Р 56042-2014 / СБП)
# =================================================================

_IP_NAME  = "ИП Васильев Руслан Геннадьевич"
_IP_INN   = "781909921730"
_IP_RS    = "40802810355710022284"
_IP_BANK  = "Северо-Западный банк ПАО Сбербанк"
_IP_BIC   = "044030653"
_IP_KS    = "30101810500000000653"


def _invoice_qr_b64(amount_rub: float, purpose: str) -> str:
    """Генерирует QR ГОСТ Р 56042-2014 и возвращает base64 PNG."""
    import qrcode
    amount_kopecks = int(round(amount_rub * 100))
    qr_data = (
        f"ST00012|Name={_IP_NAME}|PersonalAcc={_IP_RS}"
        f"|BankName={_IP_BANK}|BIC={_IP_BIC}|CorrespAcc={_IP_KS}"
        f"|PayeeINN={_IP_INN}|Sum={amount_kopecks}|Purpose={purpose}"
    )
    img = qrcode.make(qr_data)
    import io, base64
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


def _handle_invoice_create(body: dict[str, Any]) -> dict[str, Any]:
    """Создаёт счёт на оплату замера.
    body: {initData, measurement_id, amount}
    Доступно: измеряющий (measurer) или менеджер."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "measurer") or sheets.has_role(user, "manager")):
        return {"error": "forbidden"}

    measurement_id = (body.get("measurement_id") or "").strip()
    amount_raw = body.get("amount")
    rooms_count_raw = body.get("rooms_count")

    MEASUREMENT_FEE_BASE = 2500
    MEASUREMENT_FEE_EXTRA = 1000

    if not measurement_id:
        return {"error": "missing_measurement_id"}

    _ensure_measurements_sheet()
    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}

    if amount_raw is not None:
        try:
            amount = float(amount_raw)
            if amount <= 0:
                raise ValueError
        except (TypeError, ValueError):
            return {"error": "invalid_amount"}
    else:
        # Авто-расчёт по rooms_count
        if rooms_count_raw is not None:
            try:
                rooms = max(1, int(rooms_count_raw))
            except (TypeError, ValueError):
                rooms = 1
        else:
            try:
                rooms = max(1, int(row.get("rooms_count") or 1))
            except (TypeError, ValueError):
                rooms = 1
        amount = MEASUREMENT_FEE_BASE + max(0, rooms - 1) * MEASUREMENT_FEE_EXTRA

    # Сохраняем rooms_count если передан
    if rooms_count_raw is not None:
        try:
            rooms_to_save = max(1, int(rooms_count_raw))
            sheets.update_cell_by_key("Measurements", "id", measurement_id, "rooms_count", str(rooms_to_save))
        except Exception as e:
            log.warning("invoice_create: rooms_count save error: %s", e)

    client_name  = row.get("client_name", "Клиент")
    client_phone = row.get("client_phone", "")
    address      = row.get("address", "")
    sched_date   = (row.get("scheduled_at") or row.get("ts") or "")[:10]
    purpose = f"Оплата услуг замера кухни {address or measurement_id}"

    try:
        qr_b64 = _invoice_qr_b64(amount, purpose)
    except Exception as e:
        log.warning("invoice qr error: %s", e)
        qr_b64 = ""

    # Сохраняем fee в Measurements для статистики заработков
    try:
        _ensure_measurements_sheet()
        sheets.update_cell_by_key("Measurements", "id", measurement_id, "measurement_fee", str(amount))
    except Exception as e:
        log.warning("invoice_create: fee save error: %s", e)

    return {
        "ok": True,
        "measurement_id": measurement_id,
        "client_name":   client_name,
        "client_phone":  client_phone,
        "address":       address,
        "date":          sched_date,
        "amount":        amount,
        "purpose":       purpose,
        "ip_name":       _IP_NAME,
        "ip_inn":        _IP_INN,
        "bank_name":     _IP_BANK,
        "bic":           _IP_BIC,
        "rs":            _IP_RS,
        "ks":            _IP_KS,
        "qr_b64":        qr_b64,
    }


@app.post("/api/invoice_create")
async def api_invoice_create(request: Request):
    body = await _safe_json(request)
    return _handle_invoice_create(body)


# =================================================================
# SignRequest — цифровая подпись акта сборки (ФЗ-63 ПЭП)
# =================================================================

def _handle_sign_request_create(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер/сборщик инициирует подпись акта.
    body: {initData, initDataUnsafe, assembly_id, mode: canvas|code|proxy|absent}
    Для code-режима: генерирует OTP и отправляет клиенту через бот."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = str(auth["user"]["id"])
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}
    if not (sheets.has_role(user, "manager") or sheets.has_role(user, "admin")
            or sheets.has_role(user, "assembler") or sheets.has_role(user, "measurer")):
        return {"error": "forbidden"}

    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}
    mode = (body.get("mode") or "canvas").strip()
    if mode not in ("canvas", "code", "proxy", "absent"):
        return {"error": "invalid_mode"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    is_owner = (str(row.get("manager_tg_id")) == tg_id
                or str(row.get("assigned_to_tg_id")) == tg_id)
    if not is_owner:
        return {"error": "forbidden"}

    now = datetime.now(timezone.utc)
    expires_at = (now + timedelta(hours=72)).isoformat()
    # 6-значный OTP (надёжнее строковым генератором)
    otp = str(secrets.randbelow(900000) + 100000)

    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "sign_token", otp)
    sheets.update_cell_by_key("Assemblies", "id", assembly_id, "sign_token_expires_at", expires_at)

    client_tg_id = row.get("client_tg_id", "")
    client_sent = False
    if mode == "code" and client_tg_id:
        msg = (
            "🔐 <b>Код подтверждения акта сборки</b>\n\n"
            f"Адрес: {row.get('address', '—')}\n\n"
            f"Ваш код: <code>{otp}</code>\n\n"
            "Сообщите код мастеру или введите в приложении. "
            "Код действителен 72 часа."
        )
        client_sent = tg.send_message(int(client_tg_id), msg)

    sheets.log_event("sign_request_created", tg_id,
                     {"assembly_id": assembly_id, "mode": mode})
    return {
        "ok": True,
        "sign_token": otp,
        "expires_at": expires_at,
        "mode": mode,
        "client_tg_id": client_tg_id,
        "client_name": row.get("client_name", ""),
        "client_sent": client_sent,
    }


def _handle_sign_request_submit(body: dict[str, Any]) -> dict[str, Any]:
    """Фиксирует подпись. Режимы:
    canvas  — {initData, assembly_id, mode, signature_data(base64 PNG), signed_by_name, signed_by_phone?}
    code    — {assembly_id, mode, code, signed_by_name, signed_by_phone?, initData?}
    proxy   — {initData, assembly_id, mode, signed_by_name, signed_by_phone?}
    absent  — {initData, assembly_id, mode, absent_reason?}
    """
    cfg = get_config()
    mode = (body.get("mode") or "canvas").strip()
    if mode not in ("canvas", "code", "proxy", "absent"):
        return {"error": "invalid_mode"}
    assembly_id = (body.get("assembly_id") or "").strip()
    if not assembly_id:
        return {"error": "missing_assembly_id"}

    _ensure_assemblies_sheet()
    row = sheets.find_row("Assemblies", "id", assembly_id)
    if not row:
        return {"error": "assembly_not_found"}

    now_iso = _now_iso()
    signed_by_name = (body.get("signed_by_name") or "").strip()
    signed_by_phone = (body.get("signed_by_phone") or "").strip()
    signed_by_tg_id = ""
    signature_file = ""

    if mode == "code":
        code = (body.get("code") or "").strip()
        stored = (row.get("sign_token") or "").strip()
        expires_str = (row.get("sign_token_expires_at") or "").strip()
        if not stored:
            return {"error": "no_sign_token"}
        if not code:
            return {"error": "missing_code"}
        if code != stored:
            return {"error": "invalid_code"}
        if expires_str:
            try:
                # Парсим ISO без dateutil
                exp = datetime.fromisoformat(expires_str.replace("Z", "+00:00"))
                if datetime.now(timezone.utc) > exp:
                    return {"error": "code_expired"}
            except Exception:
                pass  # не блокируем если парсинг упал
        # Берём tg_id из initData если есть (клиент авторизован в боте)
        auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
        if auth and auth.get("user"):
            signed_by_tg_id = str(auth["user"]["id"])

    elif mode in ("canvas", "proxy", "absent"):
        auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
        if not auth or not auth.get("user"):
            unsafe = body.get("initDataUnsafe") or {}
            if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
                auth = {"user": unsafe["user"]}
            else:
                return {"error": "invalid_init_data"}
        tg_id = str(auth["user"]["id"])
        is_owner = (str(row.get("manager_tg_id")) == tg_id
                    or str(row.get("assigned_to_tg_id")) == tg_id)
        if not is_owner:
            return {"error": "forbidden"}
        signed_by_tg_id = tg_id

        if mode == "canvas":
            sig_data = (body.get("signature_data") or "").strip()
            if not sig_data:
                return {"error": "missing_signature_data"}
            # data:image/png;base64,<data> → bytes
            raw = sig_data.split(",", 1)[-1]
            try:
                sig_bytes = base64.b64decode(raw)
            except Exception:
                return {"error": "invalid_signature_data"}
            sig_dir = PHOTOS_DIR / assembly_id
            sig_dir.mkdir(parents=True, exist_ok=True)
            sig_filename = f"sign_{int(time.time())}.png"
            (sig_dir / sig_filename).write_bytes(sig_bytes)
            signature_file = sig_filename

        elif mode == "absent":
            reason = (body.get("absent_reason") or "Клиент отсутствовал").strip()
            signed_by_name = reason

    # Пишем все поля за один проход (каждый update_cell — отдельный запрос к Sheets)
    updates = {
        "signed_via": mode,
        "signed_by_name": signed_by_name,
        "signed_by_phone": signed_by_phone,
        "signed_by_tg_id": signed_by_tg_id,
        "signed_at": now_iso,
        "signature_file": signature_file,
    }
    for col, val in updates.items():
        sheets.update_cell_by_key("Assemblies", "id", assembly_id, col, val)

    # Автоматика: акт №3 подписан → сборка завершена
    try:
        if row.get("status") not in ("done", "cancelled"):
            sheets.update_cell_by_key("Assemblies", "id", assembly_id, "status", "done")
            sheets.update_cell_by_key("Assemblies", "id", assembly_id, "completed_at", now_iso)
            log.info("sign_request signed → assembly %s done", assembly_id)
        # Уведомить менеджера
        mgr_id = row.get("manager_tg_id")
        if mgr_id:
            tg.send_message(int(mgr_id),
                f"✅ <b>Акт №3 подписан — сборка завершена</b>\n"
                f"Сборка: <code>{assembly_id}</code>\n"
                f"Клиент: {row.get('client_name','')}\n"
                f"Подписал: {signed_by_name}")
    except Exception as e:
        log.warning("sign_submit status update error: %s", e)

    sheets.log_event("assembly_signed", signed_by_tg_id or "anon",
                     {"assembly_id": assembly_id, "mode": mode, "by": signed_by_name})
    return {"ok": True, "signed_at": now_iso, "mode": mode, "signed_by_name": signed_by_name}


@app.post("/api/sign_request_create")
async def api_sign_request_create(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_sign_request_create(body))


@app.post("/api/sign_request_submit")
async def api_sign_request_submit(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_sign_request_submit(body))


@app.post("/api/managers_list")
async def api_managers_list(request: Request):
    body = await _safe_json(request)
    return JSONResponse(_handle_managers_list(body))


def _normalize_phone(raw: str) -> tuple[str, bool]:
    """Нормализует RU-телефон в формат +7XXXXXXXXXX.
    Возвращает (нормализованный, валиден ли)."""
    if not raw:
        return "", False
    digits = "".join(c for c in raw if c.isdigit())
    # Если начинается с 8 — заменяем на 7
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    # Если 10 цифр — добавляем 7 в начало
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) != 11 or not digits.startswith("7"):
        return raw, False
    return "+" + digits, True


def _next_client_no(manager_tg_id: str) -> int:
    """Следующий порядковый номер клиента для менеджера (1, 2, 3, ...)."""
    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception:
        return 1
    if not rows or len(rows) < 2:
        return 1
    headers = rows[0]
    if "client_no" not in headers or "manager_tg_id" not in headers:
        return 1
    no_idx = headers.index("client_no")
    mgr_idx = headers.index("manager_tg_id")
    max_n = 0
    for r in rows[1:]:
        if mgr_idx < len(r) and str(r[mgr_idx]).strip() == str(manager_tg_id):
            try:
                n = int(str(r[no_idx]).strip()) if no_idx < len(r) and r[no_idx] else 0
                if n > max_n:
                    max_n = n
            except (ValueError, TypeError):
                pass
    return max_n + 1


def _handle_client_create(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер заводит клиента без замера/подбора.
    body: {initData, full_name, phone, address?, note?, contract_no?, contract_date?}
    Создаёт пустую заявку-карточку (status='draft') чтобы клиент появился
    в списке клиентов менеджера и был доступен в карточке."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    full_name = (body.get("full_name") or "").strip()
    phone_raw = (body.get("phone") or "").strip()
    address = (body.get("address") or "").strip()
    note = (body.get("note") or "").strip()
    contract_no = (body.get("contract_no") or "").strip()
    contract_date = (body.get("contract_date") or "").strip()
    gps_lat = body.get("gps_lat") or ""
    gps_lng = body.get("gps_lng") or ""

    # Валидация
    if not full_name:
        return {"error": "missing_name", "field": "full_name", "msg": "Укажите ФИО клиента"}
    if len(full_name) < 2:
        return {"error": "bad_name", "field": "full_name", "msg": "Имя слишком короткое"}

    phone, phone_ok = _normalize_phone(phone_raw)
    if not phone_ok:
        return {"error": "bad_phone", "field": "phone", "msg": "Телефон в формате +7XXXXXXXXXX (10 цифр после +7)"}

    if address and len(address) < 5:
        return {"error": "bad_address", "field": "address", "msg": "Адрес слишком короткий"}

    _ensure_measurements_sheet()
    measurement_id = _short_id()
    client_no = _next_client_no(str(tg_id))

    # Создаём «карточку клиента» как заявку со статусом draft
    sheets.append_named_row("Measurements", _row_for_measurement(
        measurement_id, _now_iso(),
        manager_tg_id=str(tg_id),
        requested_by_tg_id=str(tg_id),
        filled_by="client_card",
        status="draft",  # карточка клиента, без активного замера
        address=address,
        client_name=full_name,
        client_phone=phone,
        notes=note,
        preferred_note=note,
        client_no=str(client_no),
        contract_no=contract_no,
        contract_date=contract_date,
        gps_lat=gps_lat,
        gps_lng=gps_lng,
    ))

    # Сохраняем заметку в ClientNotes если она передана
    if note:
        try:
            _ensure_client_notes_sheet()
            key = _normalize_client_key(full_name, phone)
            sheets.append_row("ClientNotes", [str(tg_id), key, note, _now_iso()])
        except Exception:
            pass

    sheets.log_event("client_created", tg_id, {
        "id": measurement_id, "client": full_name, "phone": phone, "client_no": client_no,
    })
    # client_key — формат совместимый с _handle_clients (которое использует name.lower())
    return {
        "ok": True,
        "id": measurement_id,
        "client_name": full_name,
        "client_phone": phone,
        "client_no": client_no,
        "contract_no": contract_no,
        "client_key": full_name.lower(),
    }


def _handle_client_delete(body: dict[str, Any]) -> dict[str, Any]:
    """Soft-delete всех записей Measurements по клиенту (для текущего менеджера).
    Удаление разрешено ТОЛЬКО если у клиента нет реальной работы:
    нет лидов и все его замеры в статусе 'draft' (карточка не использована).
    body: {initData, client_key} — client_key это name.lower() как в _handle_clients."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    client_key = (body.get("client_key") or "").strip().lower()
    if not client_key:
        return {"error": "missing_client_key"}

    # Проверка: есть ли у клиента работа в Leads
    try:
        ws_l = sheets.sheet("Leads")
        rows_l = ws_l.get_all_values()
        if rows_l and len(rows_l) >= 2:
            headers_l = rows_l[0]
            for r in rows_l[1:]:
                row = dict(zip(headers_l, r + [""] * (len(headers_l) - len(r))))
                if str(row.get("manager_tg_id", "")) != str(tg_id):
                    continue
                if (row.get("client_name") or "").strip().lower() != client_key:
                    continue
                return {"error": "in_work", "msg": "У клиента есть подбор — удаление запрещено. Используйте редактирование."}
    except Exception:
        pass

    # Проверка: есть ли у клиента сборки
    try:
        ws_a = sheets.sheet("Assemblies")
        rows_a = ws_a.get_all_values()
        if rows_a and len(rows_a) >= 2:
            headers_a = rows_a[0]
            for r in rows_a[1:]:
                row = dict(zip(headers_a, r + [""] * (len(headers_a) - len(r))))
                if str(row.get("manager_tg_id", "")) != str(tg_id):
                    continue
                if row.get("archived_at"):
                    continue
                if (row.get("client_name") or "").strip().lower() != client_key:
                    continue
                return {"error": "in_work", "msg": "У клиента есть сборка — удаление запрещено. Используйте редактирование."}
    except Exception:
        pass

    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception as e:
        return {"error": f"sheets: {e}"}
    if not rows or len(rows) < 2:
        return {"ok": True, "archived": 0}

    headers = rows[0]
    if "archived_at" not in headers or "client_name" not in headers or "manager_tg_id" not in headers:
        return {"error": "schema_missing"}

    # Проверка: есть ли не-draft замеры?
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if str(row.get("manager_tg_id", "")) != str(tg_id):
            continue
        if (row.get("client_name") or "").strip().lower() != client_key:
            continue
        if row.get("archived_at"):
            continue
        status = (row.get("status") or "").strip()
        if status and status != "draft":
            return {"error": "in_work", "msg": "У клиента есть замер в работе — удаление запрещено. Используйте редактирование."}

    archived_idx = headers.index("archived_at") + 1
    now = _now_iso()
    count = 0
    for i, r in enumerate(rows[1:], start=2):
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if str(row.get("manager_tg_id", "")) != str(tg_id):
            continue
        if (row.get("client_name") or "").strip().lower() != client_key:
            continue
        if row.get("archived_at"):
            continue
        ws.update_cell(i, archived_idx, now)
        count += 1

    sheets.log_event("client_deleted", tg_id, {"client_key": client_key, "count": count})
    return {"ok": True, "archived": count}


def _handle_client_update(body: dict[str, Any]) -> dict[str, Any]:
    """Менеджер обновляет данные клиента (имя, телефон, адрес, договор).
    Обновляет ВСЕ строки Measurements этого менеджера для этого клиента.
    body: {initData, client_key, full_name?, phone?, address?, contract_no?, contract_date?}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    client_key = (body.get("client_key") or "").strip().lower()
    if not client_key:
        return {"error": "missing_client_key"}

    new_name = (body.get("full_name") or "").strip()
    new_phone_raw = (body.get("phone") or "").strip()
    new_address = body.get("address")
    new_contract_no = body.get("contract_no")
    new_contract_date = body.get("contract_date")
    new_gps_lat = body.get("gps_lat")
    new_gps_lng = body.get("gps_lng")

    if new_name and len(new_name) < 2:
        return {"error": "bad_name", "msg": "Имя слишком короткое"}

    new_phone = ""
    if new_phone_raw:
        norm, ok = _normalize_phone(new_phone_raw)
        if not ok:
            return {"error": "bad_phone", "msg": "Телефон в формате +7XXXXXXXXXX"}
        new_phone = norm

    if isinstance(new_address, str) and new_address.strip() and len(new_address.strip()) < 5:
        return {"error": "bad_address", "msg": "Адрес слишком короткий"}

    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception as e:
        return {"error": f"sheets: {e}"}
    if not rows or len(rows) < 2:
        return {"ok": True, "updated": 0}

    headers = rows[0]
    if "client_name" not in headers or "manager_tg_id" not in headers:
        return {"error": "schema_missing"}

    def col_idx(name: str) -> int | None:
        return headers.index(name) + 1 if name in headers else None

    name_col = col_idx("client_name")
    phone_col = col_idx("client_phone")
    address_col = col_idx("address")
    contract_no_col = col_idx("contract_no")
    contract_date_col = col_idx("contract_date")
    gps_lat_col = col_idx("gps_lat")
    gps_lng_col = col_idx("gps_lng")

    updated = 0
    for i, r in enumerate(rows[1:], start=2):
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if str(row.get("manager_tg_id", "")) != str(tg_id):
            continue
        if (row.get("client_name") or "").strip().lower() != client_key:
            continue
        if row.get("archived_at"):
            continue
        if new_name and name_col:
            ws.update_cell(i, name_col, new_name)
        if new_phone and phone_col:
            ws.update_cell(i, phone_col, new_phone)
        if isinstance(new_address, str) and address_col:
            ws.update_cell(i, address_col, new_address.strip())
        if isinstance(new_contract_no, str) and contract_no_col:
            ws.update_cell(i, contract_no_col, new_contract_no.strip())
        if isinstance(new_contract_date, str) and contract_date_col:
            ws.update_cell(i, contract_date_col, new_contract_date.strip())
        if new_gps_lat is not None and gps_lat_col:
            ws.update_cell(i, gps_lat_col, new_gps_lat)
        if new_gps_lng is not None and gps_lng_col:
            ws.update_cell(i, gps_lng_col, new_gps_lng)
        updated += 1

    sheets.log_event("client_updated", tg_id, {"client_key": client_key, "updated": updated})
    new_key = new_name.lower() if new_name else client_key
    return {"ok": True, "updated": updated, "client_key": new_key}


def _normalize_client_key(name: str, phone: str) -> str:
    """Стабильный ключ клиента: телефон в цифрах либо имя в lower."""
    digits = "".join(c for c in (phone or "") if c.isdigit())
    if len(digits) >= 10:
        # Нормализуем +7/8 → 7XXXXXXXXXX (последние 10 цифр)
        return "p:" + digits[-10:]
    return "n:" + (name or "").strip().lower()


_CLIENT_NOTES_HEADERS = ["manager_tg_id", "client_key", "note", "updated_at"]


def _ensure_client_notes_sheet():
    try:
        sheets.ensure_sheet("ClientNotes", _CLIENT_NOTES_HEADERS)
    except Exception as e:
        log.warning("Не удалось убедиться что ClientNotes есть: %s", e)


def _handle_client_note(body: dict[str, Any]) -> dict[str, Any]:
    """Чтение/запись примечаний менеджера по клиенту (append-only история).
    body: {initData, client_name, client_phone, note?}

    Если note передано — добавляем новую запись (append).
    Возвращает {ok, notes: [{note, updated_at}, ...], note, updated_at} (notes = все записи, новые сверху)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if not (isinstance(unsafe, dict) and unsafe.get("user", {}).get("id")):
            return {"error": "invalid_init_data"}
        auth = {"user": unsafe["user"]}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    client_name = (body.get("client_name") or "").strip()
    client_phone = (body.get("client_phone") or "").strip()
    if not client_name and not client_phone:
        return {"error": "missing_client_id"}
    key = _normalize_client_key(client_name, client_phone)

    _ensure_client_notes_sheet()

    # Если note передано — пишем новую запись (append-only, история не перезаписывается)
    if "note" in body and body.get("note") is not None:
        new_note = str(body.get("note") or "").strip()[:4000]
        if not new_note:
            return {"error": "empty_note"}
        now_iso = _now_iso()
        sheets.append_row("ClientNotes", [str(tg_id), key, new_note, now_iso])
        sheets.log_event("client_note_added", tg_id, {"key": key, "len": len(new_note)})

    # Читаем все заметки этого менеджера по этому клиенту
    try:
        ws = sheets.sheet("ClientNotes")
        rows = ws.get_all_values()
    except Exception as e:
        log.warning("ClientNotes read failed: %s", e)
        rows = []

    notes: list[dict[str, str]] = []
    if rows and len(rows) >= 2:
        headers = rows[0]
        try:
            idx_mgr  = headers.index("manager_tg_id")
            idx_key  = headers.index("client_key")
            idx_note = headers.index("note")
            idx_upd  = headers.index("updated_at")
        except ValueError:
            idx_mgr = idx_key = idx_note = idx_upd = -1
        if idx_mgr >= 0 and idx_key >= 0:
            for r in rows[1:]:
                if (r[idx_mgr] if idx_mgr < len(r) else "") == str(tg_id) \
                        and (r[idx_key] if idx_key < len(r) else "") == key:
                    note_text = r[idx_note] if idx_note < len(r) else ""
                    upd = r[idx_upd] if idx_upd < len(r) else ""
                    if note_text:
                        notes.append({"note": note_text, "updated_at": upd})

    # Новые сверху
    notes.sort(key=lambda x: x.get("updated_at", ""), reverse=True)
    latest = notes[0] if notes else {}
    return {
        "ok": True,
        "notes": notes,
        "note": latest.get("note", ""),
        "updated_at": latest.get("updated_at", ""),
        "client_key": key,
    }


def _handle_geocode(body: dict[str, Any]) -> dict[str, Any]:
    """Прямое геокодирование: текст адреса → lat/lon.
    Использует Yandex (если есть YANDEX_GEOCODER_API_KEY в env) с fallback на OSM.
    body: {initData, address, city?}"""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if not (isinstance(unsafe, dict) and unsafe.get("user", {}).get("id")):
            return {"error": "invalid_init_data"}

    address = (body.get("address") or "").strip()
    if not address:
        return {"error": "missing_address"}
    city = (body.get("city") or "Санкт-Петербург").strip()
    result = geocoder.geocode(address, city=city)
    if not result:
        return {"ok": False, "error": "not_found", "address": address}
    return {
        "ok": True,
        "address": address,
        "result": {
            "lat": result.get("lat"),
            "lng": result.get("lon"),  # фронт использует lng
            "formatted": result.get("formatted"),
            "precision": result.get("precision"),
            "kind": result.get("kind"),
            "source": result.get("source"),
        },
        "yandex_maps_url": geocoder.build_yandex_maps_url(
            result["lat"], result["lon"], text=result.get("formatted") or address,
        ),
    }


def _handle_measurement_next_no(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает следующий свободный номер замера (max существующих + 1).
    Если в Sheets ничего нет — стартуем с 1. Менеджер может скорректировать вручную
    (например первый раз поставить 158, если до этого замеры были вне системы)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if not (isinstance(unsafe, dict) and unsafe.get("user", {}).get("id")):
            return {"error": "invalid_init_data"}

    _ensure_measurements_sheet()
    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception:
        return {"ok": True, "next_no": 1}
    if not rows or len(rows) < 2:
        return {"ok": True, "next_no": 1}
    headers = rows[0]
    if "zamer_no" not in headers:
        return {"ok": True, "next_no": 1}
    idx = headers.index("zamer_no")
    max_n = 0
    for r in rows[1:]:
        if idx >= len(r):
            continue
        try:
            n = int(str(r[idx]).strip())
            if n > max_n:
                max_n = n
        except (ValueError, TypeError):
            pass
    return {"ok": True, "next_no": max_n + 1}


def _format_date_human(iso: str) -> str:
    """ISO datetime → '15.05.2026 14:00' для уведомлений."""
    if not iso:
        return "—"
    try:
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return d.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return iso


def _format_preferred_human(p_type: str, p_date: str, p_tod: str, p_note: str) -> str:
    """Приблизительная дата от менеджера в человекочитаемом виде."""
    tod_map = {"morning": "утром", "day": "днём", "evening": "вечером"}
    if p_type == "specific":
        date_part = p_date
        if p_date:
            try:
                from datetime import datetime as _dt
                date_part = _dt.strptime(p_date, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                pass
        parts = []
        if date_part:
            parts.append(date_part)
        if p_tod in tod_map:
            parts.append(tod_map[p_tod])
        s = " ".join(parts) if parts else "конкретная дата"
    elif p_type == "this_week":
        s = "эта неделя"
    elif p_type == "next_week":
        s = "следующая неделя"
    else:
        s = "согласовать с клиентом"
    if p_note:
        s += f" · {p_note}"
    return s


def _handle_measurement_detail(body: dict[str, Any]) -> dict[str, Any]:
    """Возвращает один замер целиком — для детальной страницы и печати."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user:
        return {"error": "user_not_found"}

    measurement_id = body.get("measurement_id") or body.get("id")
    if not measurement_id:
        return {"error": "missing_measurement_id"}

    row = sheets.find_row("Measurements", "id", measurement_id)
    if not row:
        return {"error": "measurement_not_found"}

    # Доступ: владелец-менеджер, любой мастер (замерщик или сборщик), клиент-владелец
    is_owner_manager = sheets.has_role(user, "manager") and str(row.get("manager_tg_id", "")) == str(tg_id)
    is_master = sheets.is_master(user)  # measurer или assembler — оба видят фото замера
    is_client = str(row.get("client_tg_id", "")) == str(tg_id)
    if not (is_owner_manager or is_master or is_client):
        return {"error": "forbidden"}

    def _safe_json(s: str) -> Any:
        try:
            return json.loads(s) if s else {}
        except (ValueError, TypeError):
            return {}

    photo_files = [p for p in (row.get("photos") or "").split(",") if p]

    return {
        "ok": True,
        "id": row.get("id"),
        "created_at": row.get("ts") or row.get("created_at"),
        "client_tg_id": row.get("client_tg_id", ""),
        "manager_tg_id": row.get("manager_tg_id", ""),
        "filled_by": row.get("filled_by", ""),
        "layout": row.get("layout", ""),
        "area_m2": row.get("area_m2", ""),
        "ceiling_mm": row.get("ceiling_mm", ""),
        "walls": _safe_json(row.get("walls", "")),
        "openings": _safe_json(row.get("openings", "")),
        "infra": _safe_json(row.get("infra", "")),
        "niches": _safe_json(row.get("niches", "")),
        "photos": photo_files,
        "notes": row.get("notes", ""),
        "status": row.get("status", ""),
        # Поля Commit B (workflow)
        "assigned_to_tg_id": row.get("assigned_to_tg_id", ""),
        "requested_by_tg_id": row.get("requested_by_tg_id", ""),
        "scheduled_at": row.get("scheduled_at", ""),
        "address": row.get("address", ""),
        "client_name": row.get("client_name", ""),
        "client_phone": row.get("client_phone", ""),
        # Поля Commit C (структура замера)
        "zamer_no": row.get("zamer_no", ""),
        "zamer_date": row.get("zamer_date", ""),
        "floor_base": row.get("floor_base", ""),
        "photos_meta": _safe_json(row.get("photos_meta", "")),
        # Приблизительная дата от менеджера (Commit C2)
        "preferred_type": row.get("preferred_type", ""),
        "preferred_date": row.get("preferred_date", ""),
        "preferred_time_of_day": row.get("preferred_time_of_day", ""),
        "preferred_note": row.get("preferred_note", ""),
        # Логистика — заполняет замерщик (Commit C3)
        "entrance":       row.get("entrance", ""),
        "floor":          row.get("floor", ""),
        "gps_lat":        row.get("gps_lat", ""),
        "gps_lng":        row.get("gps_lng", ""),
        "parking_type":   row.get("parking_type", ""),
        "parking_note":   row.get("parking_note", ""),
        "delivery_notes": row.get("delivery_notes", ""),
        # Google Calendar
        "gcal_event_id":  row.get("gcal_event_id", ""),
        "gcal_event_url": row.get("gcal_event_url", ""),
        # Чертёж и решение про подбор
        "design_files":   [f for f in (row.get("design_files") or "").split(",") if f],
        "podbor_decision":    row.get("podbor_decision", ""),
        "podbor_decision_at": row.get("podbor_decision_at", ""),
        "podbor_lead_id":     row.get("podbor_lead_id", ""),
        # Номера
        "client_no":      row.get("client_no", ""),
        "contract_no":    row.get("contract_no", ""),
        "contract_date":  row.get("contract_date", ""),
        # Оплата замера
        "measurement_fee": row.get("measurement_fee", ""),
        "rooms_count": row.get("rooms_count", ""),
        # Оценки
        "measurer_feedback_at": row.get("measurer_feedback_at", ""),
        "manager_feedback_at":  row.get("manager_feedback_at", ""),
        # Для замерщика: кто менеджер
        "viewer_is_measurer": sheets.has_role(user, "measurer"),
        "viewer_is_manager":  sheets.has_role(user, "manager"),
        "viewer_tg_id": str(tg_id),
    }


def _handle_measurements_list(body: dict[str, Any]) -> dict[str, Any]:
    """Список замеров менеджера, опционально отфильтрованный по client_tg_id / client_name."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return {"error": "only_manager"}

    client_tg_id = body.get("client_tg_id") or ""
    client_name = (body.get("client_name") or "").strip().lower()

    try:
        ws = sheets.sheet("Measurements")
        rows = ws.get_all_values()
    except Exception as e:
        log.warning("Failed to read Measurements: %s", e)
        return {"ok": True, "measurements": []}

    if not rows or len(rows) < 2:
        return {"ok": True, "measurements": []}

    headers = rows[0]
    out = []
    for r in rows[1:]:
        row = dict(zip(headers, r + [""] * (len(headers) - len(r))))
        if str(row.get("manager_tg_id", "")) != str(tg_id):
            continue
        # Скрываем soft-deleted
        if row.get("archived_at"):
            continue
        # Опциональные фильтры по клиенту
        if client_tg_id and str(row.get("client_tg_id", "")) != str(client_tg_id):
            continue
        if client_name and (row.get("client_name") or "").strip().lower() != client_name:
            continue
        photo_files = [p for p in (row.get("photos") or "").split(",") if p]
        out.append({
            "id": row.get("id", ""),
            "created_at": row.get("ts") or row.get("created_at", ""),
            "client_tg_id": row.get("client_tg_id", ""),
            "manager_tg_id": row.get("manager_tg_id", ""),
            "filled_by": row.get("filled_by", ""),
            "layout": row.get("layout", ""),
            "area_m2": row.get("area_m2", ""),
            "ceiling_mm": row.get("ceiling_mm", ""),
            "notes": row.get("notes", ""),
            "status": row.get("status", ""),
            "photos": photo_files,
            "photo_count": len(photo_files),
            # Ключевые поля для рендера карточки клиента и таймлайна
            "client_name": row.get("client_name", ""),
            "client_phone": row.get("client_phone", ""),
            "address": row.get("address", ""),
            "scheduled_at": row.get("scheduled_at", ""),
            "client_no": row.get("client_no", ""),
            "contract_no": row.get("contract_no", ""),
            "contract_date": row.get("contract_date", ""),
            "assigned_to_tg_id": row.get("assigned_to_tg_id", ""),
        })

    # Сортируем по дате desc
    out.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return {"ok": True, "count": len(out), "measurements": out}


_CONTRACT_SYSTEM = """\
Ты — помощник клиента мебельной фабрики ЗОВ (Россия). \
Клиент получил договор на покупку и монтаж кухни и хочет понять его содержание.

Твоя задача — проанализировать текст и дать ответ строго в формате JSON без Markdown-оберток. \
Структура ответа:
{
  "summary": "2-3 предложения — о чём договор",
  "payment": {
    "total": "итоговая сумма если есть",
    "schedule": "схема оплаты (предоплата %, доплата когда)",
    "prepayment_pct": число_или_null
  },
  "deadlines": [
    {"label": "Изготовление", "value": "дата или срок", "note": "подробности"}
  ],
  "risks": [
    {"level": "high|medium|low", "title": "заголовок", "description": "что именно"}
  ],
  "recommendations": ["что уточнить у менеджера или на что обратить внимание"],
  "missing_clauses": ["важные пункты которых нет в тексте"]
}

Риски level:
- high — условие явно невыгодно клиенту или может привести к потере денег
- medium — спорное, зависит от ситуации
- low — мелочь, но лучше знать

Если есть конкретный вопрос (поле question) — добавь поле "question_answer": "ответ на вопрос".
Отвечай на русском. Будь конкретным, избегай юридического жаргона.
"""


def _handle_contract_review(body: dict[str, Any]) -> dict[str, Any]:
    """Клиент вставляет текст договора — AI анализирует его простым языком.
    body: { initData, text: str, question?: str }
    """
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth:
        unsafe = body.get("initDataUnsafe") or {}
        if not (isinstance(unsafe, dict) and unsafe.get("user", {}).get("id")):
            return {"error": "invalid_init_data"}

    text = str(body.get("text") or "").strip()
    question = str(body.get("question") or "").strip()

    if not text:
        return {"error": "text_required"}
    if len(text) > 16_000:
        text = text[:16_000]

    user_prompt = f"Текст договора:\n\n{text}"
    if question:
        user_prompt += f"\n\nВопрос клиента: {question}"

    import asyncio
    result = ai.call_ai(
        user_prompt,
        system_prompt=_CONTRACT_SYSTEM,
        temperature=0.2,
        max_tokens=3000,
    )

    if result.get("error"):
        return {"error": result.get("text", "AI error")}

    analysis = result.get("json") or {}
    return {
        "ok": True,
        "analysis": analysis,
        "raw_text": result.get("text", ""),
        "tokens": result.get("tokens", 0),
        "model": result.get("model", ""),
    }


def _handle_test_ai() -> dict[str, Any]:
    cfg = get_config()
    res = ai.call_ai("Скажи одной фразой: что за фабрика ЗОВ?",
                     system_prompt="Ты — кратко и по делу отвечаешь. Без markdown.")
    return {
        "ok": not res.get("error"),
        "provider": "GigaChat",
        "model": res.get("model", cfg.gigachat_model),
        "response_text": (res.get("text") or "")[:500],
        "tokens": res.get("tokens", 0),
    }


def _handle_test_telegram() -> dict[str, Any]:
    cfg = get_config()
    ok = tg.send_message(
        cfg.admin_tg_id,
        "🟢 Привет из Python-бэкенда на VPS! Связка backend↔бот работает.",
    )
    return {"ok": ok, "sent_to": cfg.admin_tg_id}


def _handle_seed_admin() -> dict[str, Any]:
    cfg = get_config()
    admin_id = cfg.admin_tg_id
    if sheets.find_row("Managers", "tg_id", admin_id):
        return {"ok": True, "status": "already_seeded", "admin_id": admin_id}
    sheets.append_row("Managers", [
        admin_id, "Руслан Васильев", "vasrusgen@gmail.com", "",
        "ЗОВ — куратор сети", "Санкт-Петербург",
        True, "active", "", "", 0, 0, 0, "MGR_ADMIN",
    ])
    if not sheets.find_user(admin_id):
        sheets.append_row("Users", [
            admin_id, "VASRUSGEN", "Руслан", "Васильев", "manager",
            _now_iso(), _now_iso(), "",
        ])
    return {"ok": True, "status": "seeded", "admin_id": admin_id, "full_name": "Руслан Васильев"}


# =================================================================
# Helpers
# =================================================================

_CAT_LABELS = {
    "fridge": "❄️ Холодильник",
    "hob": "🔥 Варочная панель",
    "oven": "🔥 Духовой шкаф",
    "dw": "💧 Посудомоечная",
    "hood": "💨 Вытяжка",
    "microwave": "📻 СВЧ",
    "coffee": "☕ Кофемашина",
    "washer": "🧺 Стиральная машина",
}


def _format_podbor_for_telegram(ai_result: dict[str, Any], client_name: str, lead_id: str = "") -> str:
    if ai_result.get("error"):
        return f"❌ Не удалось получить подбор от AI.\n{ai_result.get('text', '')}"
    j = ai_result.get("json")
    if not j:
        return "<b>Подбор готов</b>\n\n" + (ai_result.get("text") or "")[:3500]

    lines = ["✅ <b>Подбор готов</b>"]
    if client_name:
        lines.append(f"Клиент: <b>{client_name}</b>")
    lines.append("")
    if j.get("summary"):
        lines.append(j["summary"])
        lines.append("")

    # Новая структура: by_category
    by_cat = j.get("by_category") or {}
    if by_cat:
        for cat_key, cat_data in by_cat.items():
            cat_label = _CAT_LABELS.get(cat_key, cat_key.upper())
            lines.append(f"━━━ <b>{cat_label}</b> ━━━")
            # Анализ категории от AI
            analysis = (cat_data or {}).get("analysis")
            if analysis:
                lines.append(f"<i>{analysis}</i>")
                lines.append("")
            models = (cat_data or {}).get("models") or []
            for i, m in enumerate(models, 1):
                lines.append(f"<b>{i}. {m.get('brand', '')} {m.get('model', '')}</b>")
                # Цены и магазины из enrichment
                enriched = m.get("enriched") or {}
                pmin = enriched.get("price_min_rub") or m.get("price_min_rub")
                pmax = enriched.get("price_max_rub") or m.get("price_max_rub")
                if pmin and pmax and pmin != pmax:
                    lines.append(f"💰 {_format_price(pmin)} — {_format_price(pmax)} ₽")
                elif pmin:
                    lines.append(f"💰 {_format_price(pmin)} ₽")
                # Отзывы и рейтинг (если есть)
                rating = enriched.get("rating_max")
                reviews = enriched.get("reviews_total")
                meta_parts = []
                if rating: meta_parts.append(f"★ {rating:.1f}")
                if reviews: meta_parts.append(f"{reviews} отзыв.")
                stores = enriched.get("stores_count")
                if stores: meta_parts.append(f"{stores} магаз.")
                if meta_parts:
                    lines.append("📊 " + " · ".join(meta_parts))
                # Источники где нашли товар
                sources_found = [
                    src.upper() for src in ("ozon", "citilink", "wb", "yamarket", "dns")
                    if enriched.get(src)
                ]
                if sources_found:
                    lines.append(f"🛒 Нашли в: {' · '.join(sources_found)}")
                if m.get("highlights"):
                    lines.append("✓ " + ", ".join(m["highlights"]))
                if m.get("pros"):
                    lines.append("<b>⊕ Плюсы:</b>")
                    for p in m["pros"][:4]:
                        lines.append(f"  • {p}")
                if m.get("cons"):
                    lines.append("<b>⊖ Минусы:</b>")
                    for c in m["cons"][:3]:
                        lines.append(f"  • {c}")
                if m.get("reasoning"):
                    lines.append(f"<i>💡 {m['reasoning']}</i>")
                # Ссылка на «лучший» магазин
                best_url = enriched.get("best_url")
                if best_url:
                    lines.append(f"🔗 <a href=\"{best_url}\">Открыть в магазине</a>")
                lines.append("")
            lines.append("")
    else:
        # Fallback: старая структура items[]
        for item in (j.get("items") or []):
            lines.append(f"<b>{item.get('brand', '')} {item.get('model', '')}</b>")
            if item.get("price_rub"):
                lines.append(f"💰 {_format_price(item['price_rub'])} ₽")
            if item.get("highlights"):
                lines.append("✓ " + ", ".join(item["highlights"]))
            if item.get("caveats"):
                lines.append(f"⚠️ {item['caveats']}")
            lines.append("")

    # Итого
    tpe = j.get("total_price_estimate_rub") or {}
    if isinstance(tpe, dict) and (tpe.get("min") or tpe.get("max")):
        tmin = tpe.get("min", 0)
        tmax = tpe.get("max", 0)
        if tmin and tmax and tmin != tmax:
            lines.append(f"<b>ИТОГО: {_format_price(tmin)} — {_format_price(tmax)} ₽</b> · {j.get('budget_status', '')}")
        else:
            lines.append(f"<b>ИТОГО: {_format_price(tmin or tmax)} ₽</b> · {j.get('budget_status', '')}")
    elif j.get("total_price_rub"):
        lines.append(f"<b>ИТОГО: {_format_price(j['total_price_rub'])} ₽</b> · {j.get('budget_status', '')}")

    if j.get("warnings"):
        lines.append("\n⚠️ " + "; ".join(j["warnings"]))
    if lead_id:
        lines.append(f"\n<i>ID: {lead_id[:8]}</i>")
    return "\n".join(lines)


def _format_price(n: int | float) -> str:
    if n is None:
        return "—"
    s = str(int(round(float(n))))
    # Разделители тысяч пробелом
    return " ".join([s[max(0, len(s) - 3 * (i + 1)):len(s) - 3 * i] for i in range((len(s) + 2) // 3)][::-1]).strip()


def _initial(name: str) -> str:
    return ((name or "").strip()[:1] or "?").upper()


def _xlsx_auth_manager(body: dict[str, Any]) -> tuple[Any, dict[str, Any] | None]:
    """Проверяет initData и возвращает (tg_id, None) для менеджера или (None, error_dict)."""
    cfg = get_config()
    auth = verify_init_data(body.get("initData") or "", cfg.bot_token)
    if not auth or not auth.get("user"):
        unsafe = body.get("initDataUnsafe") or {}
        if isinstance(unsafe, dict) and unsafe.get("user", {}).get("id"):
            auth = {"user": unsafe["user"]}
        else:
            return None, {"error": "invalid_init_data"}
    tg_id = auth["user"]["id"]
    user = sheets.find_user(tg_id)
    if not user or not sheets.has_role(user, "manager"):
        return None, {"error": "only_manager"}
    return tg_id, None  # успешно: вернуть None-ошибку, чтобы caller мог сделать if err:


def _parse_xlsx_groups(file_bytes: bytes, source_label: str) -> list[dict[str, Any]]:
    """Общий парсер для ОТГРУЗКИ.xlsx и «Поступление заказов на склад СПб.xlsx».

    Оба файла содержат листы вида «ЗОВ ДД.ММ.ГГ» с одинаковыми столбцами:
    №, Товар (Заказ/Дозаказ), договор №, Срок, Кол мест, Фурн-ра СПБ,
    Панели/техника СПБ, (empty), Продавец, Сборщик, Примечание, Дата отгрузки, Кто забрал.

    «Поступление» дополнительно имеет:
    - 2 строки-шапки перед заголовком (Накладная от…, Поставщик:…)
    - разделители «Кухни» / «Дозаказы» между блоками данных
    Функция находит строку заголовков динамически (первая строка с «Товар»).
    """
    import io
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl_not_installed")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    groups: list[dict[str, Any]] = []

    # Метки секций, которые нужно пропускать как строки-данные
    _SECTION_LABELS = {"кухни", "дозаказы", "кухня"}

    for sheet_name in wb.sheetnames:
        if not sheet_name.strip().upper().startswith("ЗОВ "):
            continue
        date_part = sheet_name.strip()[4:].strip()
        try:
            if len(date_part) == 8:
                factory_date = datetime.strptime(date_part, "%d.%m.%y").date()
            elif len(date_part) == 10:
                factory_date = datetime.strptime(date_part, "%d.%m.%Y").date()
            else:
                continue
        except ValueError:
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue

        # Динамически находим строку заголовков — первая строка, содержащая «Товар»
        header_idx = None
        for i, row in enumerate(all_rows):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if "товар" in cells:
                header_idx = i
                break
        if header_idx is None:
            continue  # нет строки с заголовком

        raw_headers = all_rows[header_idx]
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        def _clean(v: Any) -> str:
            s = str(v or "").strip()
            return "" if s.lower() in ("none", "") else s

        items: list[dict[str, Any]] = []
        for row in all_rows[header_idx + 1:]:
            # Пропускаем полностью пустые строки
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if not non_empty:
                continue

            # Строим словарь
            rd: dict[str, Any] = {}
            for i, val in enumerate(row):
                key = headers[i] if i < len(headers) else f"_col{i}"
                rd[key] = val

            tovar_raw = str(rd.get("Товар") or "").strip()
            if not tovar_raw or tovar_raw.lower() in ("none", "товар"):
                continue
            # Пропускаем разделители-секции («Кухни», «Дозаказы»)
            if tovar_raw.lower() in _SECTION_LABELS:
                continue
            # Пропускаем шаблонные пустые строки — Заказ/Дозаказ без прочих данных
            if tovar_raw in ("Заказ", "Дозаказ") and len(non_empty) <= 2:
                continue

            # Договор № — колонка с заголовком «договор», или 3-я колонка (C)
            contract = ""
            for h in headers:
                if "договор" in h.lower() or "дог" in h.lower():
                    contract = _clean(rd.get(h))
                    break
            if not contract and len(headers) > 2:
                contract = _clean(rd.get(headers[2]))

            items.append({
                "num":           _clean(rd.get("№")),
                "tovar":         tovar_raw,
                "contract":      contract,
                "deadline":      _parse_xlsx_date(rd.get("Срок")),
                "places":        _clean(rd.get("Кол мест") or rd.get("Кол. мест") or rd.get("Кол.мест")),
                "furn_spb":      _clean(rd.get("Фурн-ра СПБ") or rd.get("фурн-ра СПБ") or rd.get("Фурн СПБ")),
                "panels_spb":    _clean(rd.get("Панели/техника СПБ") or rd.get("панели и техника СПБ в заказе") or rd.get("Панели СПБ")),
                "seller":        _clean(rd.get("Продавец")),
                "assembler":     _clean(rd.get("Сборщик")),
                "note":          _clean(rd.get("Примечание")),
                "delivery_date": _parse_xlsx_date(rd.get("Дата отгрузки")),
                "picked_up_by":  _clean(rd.get("Кто забрал")),
            })

        if items:
            groups.append({
                "factory_date":     factory_date.strftime("%d.%m.%Y"),
                "factory_date_iso": factory_date.isoformat(),
                "sheet_name":       sheet_name.strip(),
                "source":           source_label,
                "count":            len(items),
                "count_zakazov":    sum(1 for i in items if "Доз" not in i["tovar"]),
                "count_dozakazov":  sum(1 for i in items if "Доз" in i["tovar"]),
                "items":            items,
            })

    groups.sort(key=lambda x: x["factory_date_iso"])
    return groups


def _handle_shipments(body: dict[str, Any]) -> dict[str, Any]:
    """ОТГРУЗКИ.xlsx — отгрузки с завода. Только для менеджера."""
    tg_id, err = _xlsx_auth_manager(body)
    if err:
        return err

    cfg = get_config()
    file_id = cfg.shipments_file_id
    if not file_id:
        return {"ok": True, "shipments": [], "note": "file_not_configured"}
    try:
        file_bytes = drive.download_file_bytes(file_id)
    except Exception as e:
        log.exception("shipments: не удалось скачать drive=%s", file_id)
        return {"error": f"drive_error: {str(e)}"}
    try:
        groups = _parse_xlsx_groups(file_bytes, "shipments")
    except Exception as e:
        log.exception("shipments: ошибка парсинга xlsx")
        return {"error": f"parse_error: {str(e)}"}
    return {"ok": True, "shipments": groups}


def _handle_arrivals(body: dict[str, Any]) -> dict[str, Any]:
    """«Поступление заказов на склад СПб.xlsx» — приход на склад. Только для менеджера."""
    tg_id, err = _xlsx_auth_manager(body)
    if err:
        return err

    cfg = get_config()
    file_id = cfg.arrivals_file_id
    if not file_id:
        return {"ok": True, "shipments": [], "note": "file_not_configured"}
    try:
        file_bytes = drive.download_file_bytes(file_id)
    except Exception as e:
        log.exception("arrivals: не удалось скачать drive=%s", file_id)
        return {"error": f"drive_error: {str(e)}"}
    try:
        groups = _parse_xlsx_groups(file_bytes, "arrivals")
    except Exception as e:
        log.exception("arrivals: ошибка парсинга xlsx")
        return {"error": f"parse_error: {str(e)}"}
    return {"ok": True, "shipments": groups}


def _parse_xlsx_date(val: Any) -> str:
    """Парсит дату из ячейки Excel — datetime, date или строка."""
    if val is None:
        return ""
    from datetime import date as date_t
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, date_t):
        return val.strftime("%d.%m.%Y")
    s = str(val).strip()
    if not s or s.lower() in ("none", ""):
        return ""
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass
    return s  # вернём как есть если не распознали


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _format_date(d) -> str | None:
    if not d:
        return None
    if isinstance(d, datetime):
        return d.strftime("%d.%m.%Y")
    return str(d)


def _short_id() -> str:
    return uuid.uuid4().hex[:13]


async def _safe_json(request: Request) -> dict[str, Any]:
    try:
        return await request.json()
    except Exception:
        return {}
